from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, Response
import os
import sqlite3
import sys
from dotenv import load_dotenv
from werkzeug.utils import secure_filename
# Password disimpan sebagai plain text (tidak di-hash)
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from flask_wtf.csrf import CSRFProtect
import re
from PIL import Image, ImageOps, ImageStat
import io
import base64
import tempfile
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, KeepTogether, PageBreak
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Pastikan stdout mendukung UTF-8 (hindari UnicodeEncodeError di Windows)
try:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


# =========================
# Konfigurasi Database SQLite (WAJIB di PythonAnywhere)
# =========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.path.join(BASE_DIR, 'bgtk_db.db')

def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', os.urandom(24).hex())  # Secret key untuk session

# Konfigurasi session permanen (30 hari)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = False  # Set True jika menggunakan HTTPS
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Setup CSRF Protection
csrf = CSRFProtect(app)

# Konfigurasi upload folder
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'uploads')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Buat folder upload jika belum ada
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Konfigurasi database SQLite
DB_NAME = os.getenv('DB_NAME', 'bgtk_db.db')
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), DB_NAME)

# Konstanta untuk ukuran kertas PDF (F4)
F4_SIZE = (8.27 * inch, 13 * inch)

def get_db_connection():
    """Membuat koneksi ke database SQLite"""
    try:
        connection = sqlite3.connect(DB_PATH, check_same_thread=False)
        connection.row_factory = sqlite3.Row  # Enable dictionary-like access
        # Enable foreign key constraints
        connection.execute('PRAGMA foreign_keys = ON')
        return connection
    except sqlite3.Error as e:
        print(f"❌ Error connecting to SQLite: {e}")
        print(f"   Database path: {DB_PATH}")
        return None

def get_db_cursor_dict(connection):
    """Membuat cursor yang mengembalikan dictionary (untuk kompatibilitas dengan kode lama)"""
    connection.row_factory = sqlite3.Row
    return connection.cursor()

def row_to_dict(row):
    """Convert sqlite3.Row to dictionary"""
    if row is None:
        return None
    if isinstance(row, dict):
        return row
    return dict(row)

def column_exists(connection, table_name, column_name):
    """Cek apakah kolom ada di tabel (SQLite equivalent of INFORMATION_SCHEMA)"""
    try:
        # SQLite PRAGMA tidak support parameter binding untuk nama tabel, gunakan string formatting dengan hati-hati
        cursor = connection.cursor()
        # Sanitize table_name untuk mencegah SQL injection (hanya alphanumeric dan underscore)
        safe_table_name = ''.join(c for c in table_name if c.isalnum() or c == '_')
        cursor.execute(f"PRAGMA table_info({safe_table_name})")
        columns = cursor.fetchall()
        for col in columns:
            if col[1] == column_name:  # Column name is at index 1
                return True
        return False
    except sqlite3.Error:
        return False

def table_exists(connection, table_name):
    """Cek apakah tabel ada di database"""
    try:
        cursor = connection.cursor()
        cursor.execute("""
            SELECT name FROM sqlite_master
            WHERE type='table' AND name=?
        """, (table_name,))
        return cursor.fetchone() is not None
    except sqlite3.Error:
        return False

def init_database():
    """Menginisialisasi database dan membuat tabel users jika belum ada"""
    connection = get_db_connection()
    if connection is None:
        print("❌ Gagal membuat koneksi ke SQLite!")
        print("   Pastikan:")
        print("   1. Folder aplikasi memiliki permission write")
        print("   2. Konfigurasi di .env atau app.py sudah benar")
        return False

    try:
        cursor = connection.cursor()

        # Enable foreign keys
        cursor.execute('PRAGMA foreign_keys = ON')

        # Cek apakah tabel users sudah ada
        users_table_exists = table_exists(connection, 'users')

        # Membuat tabel users jika belum ada
        if not users_table_exists:
            print("📋 Membuat tabel 'users' jika belum ada...")
            create_table_query = """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nama VARCHAR(255) DEFAULT NULL,
                email VARCHAR(255) DEFAULT NULL,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                role TEXT DEFAULT 'user',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
            cursor.execute(create_table_query)
            connection.commit()
            print("✅ Tabel 'users' siap!")
        else:
            print("📋 Tabel 'users' sudah ada, mengecek kolom...")

            # Tambahkan kolom role jika tabel sudah ada tapi kolom belum ada
            if not column_exists(connection, 'users', 'role'):
                print("📝 Menambahkan kolom 'role' ke tabel 'users'...")
                cursor.execute("ALTER TABLE users ADD COLUMN role TEXT DEFAULT 'user'")
                connection.commit()
                print("✅ Kolom 'role' berhasil ditambahkan!")

            # Tambahkan kolom nama jika tabel sudah ada tapi kolom belum ada
            if not column_exists(connection, 'users', 'nama'):
                print("📝 Menambahkan kolom 'nama' ke tabel 'users'...")
                cursor.execute("ALTER TABLE users ADD COLUMN nama VARCHAR(255) NULL")
                connection.commit()
                print("✅ Kolom 'nama' berhasil ditambahkan!")

            # Tambahkan kolom email jika tabel sudah ada tapi kolom belum ada
            if not column_exists(connection, 'users', 'email'):
                print("📝 Menambahkan kolom 'email' ke tabel 'users'...")
                cursor.execute("ALTER TABLE users ADD COLUMN email VARCHAR(255) NULL")
                connection.commit()
                print("✅ Kolom 'email' berhasil ditambahkan!")

        # Buat akun admin default jika belum ada
        try:
            admin_username = 'admin'
            admin_password = 'admin123'
            cursor.execute("SELECT id FROM users WHERE username = ?", (admin_username,))
            admin_exists = cursor.fetchone()

            if not admin_exists:
                print("👤 Membuat akun admin default...")
                # Simpan password sebagai plain text
                cursor.execute(
                    "INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                    (admin_username, admin_password, 'admin')
                )
                connection.commit()
                print(f"✅ Akun admin default berhasil dibuat!")
                print(f"   Username: {admin_username}")
                print(f"   Password: {admin_password}")
            else:
                # Update role jika admin sudah ada tapi belum set role
                cursor.execute("UPDATE users SET role = 'admin' WHERE username = ? AND (role IS NULL OR role = 'user' OR role = 'operator')", (admin_username,))
                # Update password ke plain text jika masih hash
                cursor.execute("SELECT password FROM users WHERE username = ?", (admin_username,))
                existing_admin = cursor.fetchone()
                if existing_admin and existing_admin[0] and (existing_admin[0].startswith('$') or existing_admin[0].startswith('pbkdf2:') or existing_admin[0].startswith('scrypt:')):
                    # Password masih hash, update ke plain text
                    cursor.execute("UPDATE users SET password = ? WHERE username = ?", (admin_password, admin_username))
                connection.commit()
                if cursor.rowcount > 0:
                    print("✅ Role admin berhasil diupdate untuk user yang sudah ada!")
        except sqlite3.Error as e:
            print(f"⚠️  Perhatian saat membuat akun admin: {e}")

        # Membuat tabel biodata_kegiatan jika belum ada
        print("📋 Membuat tabel 'biodata_kegiatan' jika belum ada...")
        create_biodata_table_query = """
        CREATE TABLE IF NOT EXISTS biodata_kegiatan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nik VARCHAR(20) NOT NULL,
            user_id INTEGER NOT NULL,
            nama_lengkap VARCHAR(100) NOT NULL,
            nip_nippk VARCHAR(50) NOT NULL,
            tempat_lahir VARCHAR(100) NOT NULL,
            tanggal_lahir DATE NOT NULL,
            jenis_kelamin TEXT NOT NULL,
            agama TEXT NOT NULL,
            pendidikan_terakhir TEXT NOT NULL,
            jurusan VARCHAR(100) NOT NULL,
            alamat_domisili TEXT NOT NULL,
            alamat_email VARCHAR(100) NOT NULL,
            no_hp VARCHAR(20) NOT NULL,
            npwp VARCHAR(50) NOT NULL,
            status_asn TEXT NOT NULL,
            pangkat_golongan VARCHAR(100) NOT NULL,
            jabatan VARCHAR(100) NOT NULL,
            instansi VARCHAR(200) NOT NULL,
            alamat_instansi TEXT NOT NULL,
            kabupaten_kota VARCHAR(100) NOT NULL,
            kabko_lainnya VARCHAR(100) DEFAULT NULL,
            peran VARCHAR(100) NOT NULL,
            nama_kegiatan TEXT NOT NULL,
            waktu_pelaksanaan VARCHAR(100) NOT NULL,
            tempat_pelaksanaan VARCHAR(200) NOT NULL,
            nama_bank VARCHAR(100) NOT NULL,
            nama_bank_lainnya VARCHAR(100) DEFAULT NULL,
            no_rekening VARCHAR(50) NOT NULL,
            nama_pemilik_rekening VARCHAR(100) NOT NULL,
            buku_tabungan_path VARCHAR(255) DEFAULT NULL,
            tanda_tangan TEXT DEFAULT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        )
        """
        cursor.execute(create_biodata_table_query)
        connection.commit()
        print("✅ Tabel 'biodata_kegiatan' siap!")

        # Membuat tabel kegiatan_master jika belum ada
        print("📋 Membuat tabel 'kegiatan_master' jika belum ada...")
        create_kegiatan_master_table_query = """
        CREATE TABLE IF NOT EXISTS kegiatan_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nama_kegiatan TEXT NOT NULL,
            waktu_pelaksanaan VARCHAR(100) NOT NULL,
            tempat_pelaksanaan VARCHAR(200) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
        cursor.execute(create_kegiatan_master_table_query)
        connection.commit()
        print("✅ Tabel 'kegiatan_master' siap!")

        # Tambahkan kolom is_hidden jika belum ada
        if not column_exists(connection, 'kegiatan_master', 'is_hidden'):
            print("📝 Menambahkan kolom 'is_hidden' ke tabel 'kegiatan_master'...")
            try:
                cursor.execute("ALTER TABLE kegiatan_master ADD COLUMN is_hidden INTEGER DEFAULT 0")
                connection.commit()
                print("✅ Kolom 'is_hidden' berhasil ditambahkan!")
            except sqlite3.Error as e:
                print(f"⚠️  Perhatian saat menambahkan kolom is_hidden: {e}")

        # Membuat tabel operator_kegiatan untuk relasi many-to-many antara operator dan kegiatan
        print("📋 Membuat tabel 'operator_kegiatan' jika belum ada...")
        create_operator_kegiatan_table_query = """
        CREATE TABLE IF NOT EXISTS operator_kegiatan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            kegiatan_id INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (user_id, kegiatan_id),
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
            FOREIGN KEY (kegiatan_id) REFERENCES kegiatan_master(id) ON DELETE CASCADE
        )
        """
        cursor.execute(create_operator_kegiatan_table_query)
        connection.commit()
        print("✅ Tabel 'operator_kegiatan' siap!")

        # Tambahkan kolom user_id jika tabel sudah ada tapi kolom belum ada
        if not column_exists(connection, 'biodata_kegiatan', 'user_id'):
            print("📝 Menambahkan kolom 'user_id' ke tabel 'biodata_kegiatan'...")
            try:
                cursor.execute("ALTER TABLE biodata_kegiatan ADD COLUMN user_id INTEGER NOT NULL DEFAULT 1")
                connection.commit()
                print("✅ Kolom 'user_id' berhasil ditambahkan!")
            except sqlite3.Error as e:
                print(f"⚠️  Perhatian saat menambahkan kolom user_id: {e}")

        # Create index untuk user_id
        try:
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_biodata_user_id ON biodata_kegiatan (user_id)")
            connection.commit()
        except sqlite3.Error:
            pass

        print("🎉 Database berhasil diinisialisasi!")
        return True

    except sqlite3.Error as e:
        print(f"❌ Error initializing database: {e}")
        return False
    finally:
        if connection:
            cursor.close()
            connection.close()

# Inisialisasi database saat aplikasi dimulai
print("🚀 Memulai inisialisasi database...")
if not init_database():
    print("⚠️  Peringatan: Inisialisasi database gagal atau database belum siap!")
    print("   Silakan periksa file database dan coba refresh halaman.")
else:
    print("✅ Database siap digunakan!")

def allowed_file(filename):
    """Cek apakah file yang diupload memiliki ekstensi yang diizinkan"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def is_valid_image(file):
    """Validasi bahwa file benar-benar gambar dengan MIME type dan content verification"""
    if not file or not file.filename:
        return False

    # Cek ekstensi file
    if not allowed_file(file.filename):
        return False

    try:
        # Baca file content
        file.seek(0)  # Reset file pointer
        file_content = file.read()
        file.seek(0)  # Reset lagi untuk save

        if len(file_content) == 0:
            return False

        # Cek MIME type dari content (bukan dari filename)
        # Valid MIME types untuk gambar
        valid_mime_types = {
            b'\xff\xd8\xff': 'image/jpeg',  # JPEG
            b'\x89PNG\r\n\x1a\n': 'image/png',  # PNG
            b'GIF87a': 'image/gif',  # GIF87a
            b'GIF89a': 'image/gif',  # GIF89a
        }

        # Cek magic bytes
        is_valid_mime = False
        for magic_bytes, mime_type in valid_mime_types.items():
            if file_content.startswith(magic_bytes):
                is_valid_mime = True
                break

        if not is_valid_mime:
            return False

        # Verifikasi dengan PIL bahwa file benar-benar gambar yang valid
        try:
            img = Image.open(io.BytesIO(file_content))
            img.verify()  # Verify that it is, in fact, an image
            return True
        except Exception:
            return False

    except Exception as e:
        print(f"Error validating image: {e}")
        return False

def save_uploaded_file(file, nik):
    """Menyimpan file yang diupload dengan validasi ketat dan mengembalikan path-nya (relatif dari static folder)"""
    if not file or not file.filename:
        return None

    # Validasi file dengan MIME type dan image verification
    if not is_valid_image(file):
        return None

    try:
        # Buat nama file unik berdasarkan NIK dan timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_extension = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'
        filename = secure_filename(f"{nik}_{timestamp}.{file_extension}")
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        # Buka dan verifikasi gambar lagi, lalu save
        file.seek(0)
        img = Image.open(io.BytesIO(file.read()))

        # Convert ke RGB jika perlu (untuk format yang tidak support RGB langsung)
        if img.mode in ('RGBA', 'LA', 'P'):
            # Buat background putih untuk transparansi
            rgb_img = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            rgb_img.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
            img = rgb_img
        elif img.mode != 'RGB':
            img = img.convert('RGB')

        # Resize jika terlalu besar (max 1920x1080 untuk menghemat storage)
        max_size = (1920, 1080)
        if img.size[0] > max_size[0] or img.size[1] > max_size[1]:
            img.thumbnail(max_size, Image.Resampling.LANCZOS)

        # Save dengan quality 85% untuk kompresi
        img.save(filepath, 'JPEG', quality=85, optimize=True)

        # Kembalikan path relatif dari static folder (uploads/filename.jpg)
        return os.path.join('uploads', filename).replace('\\', '/')
    except Exception as e:
        print(f"Error saving uploaded file: {e}")
        return None

def save_tanda_tangan_file(tanda_tangan_base64, nik):
    """Menyimpan tanda tangan dari base64 ke file dan mengembalikan path-nya (relatif dari static folder)"""
    if not tanda_tangan_base64:
        print("❌ save_tanda_tangan_file: tanda_tangan_base64 is None or empty")
        return None

    try:
        print(f"🔍 save_tanda_tangan_file - NIK: {nik}")
        print(f"🔍 save_tanda_tangan_file - tanda_tangan_base64 type: {type(tanda_tangan_base64)}")
        print(f"🔍 save_tanda_tangan_file - tanda_tangan_base64 length: {len(str(tanda_tangan_base64))}")
        print(f"🔍 save_tanda_tangan_file - tanda_tangan_base64 preview: {str(tanda_tangan_base64)[:100]}")

        # Decode base64
        img_data = None
        if isinstance(tanda_tangan_base64, str) and tanda_tangan_base64.startswith('data:image'):
            # Format: data:image/png;base64,...
            print("🔍 save_tanda_tangan_file - Format: data:image")
            header, encoded = tanda_tangan_base64.split(',', 1)
            img_data = base64.b64decode(encoded)
            print(f"🔍 save_tanda_tangan_file - Decoded size: {len(img_data)} bytes")
        else:
            # Base64 langsung
            print("🔍 save_tanda_tangan_file - Format: base64 langsung")
            # Hapus whitespace jika ada
            clean_data = str(tanda_tangan_base64).strip().replace('\n', '').replace('\r', '').replace(' ', '')
            img_data = base64.b64decode(clean_data)
            print(f"🔍 save_tanda_tangan_file - Decoded size: {len(img_data)} bytes")

        if not img_data:
            print("❌ save_tanda_tangan_file: Failed to decode base64")
            return None

        # Buka gambar
        img = Image.open(io.BytesIO(img_data))
        print(f"🔍 save_tanda_tangan_file - Image opened: {img.size}, mode: {img.mode}")

        # Convert ke RGB jika perlu
        if img.mode in ('RGBA', 'LA', 'P'):
            rgb_img = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            rgb_img.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
            img = rgb_img
        elif img.mode != 'RGB':
            img = img.convert('RGB')

        # Buat nama file unik berdasarkan NIK dan timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = secure_filename(f"{nik}_ttd_{timestamp}.jpg")
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        print(f"🔍 save_tanda_tangan_file - Saving to: {filepath}")

        # Pastikan folder upload ada
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

        # Save sebagai JPEG dengan quality 90% untuk kualitas yang baik
        img.save(filepath, 'JPEG', quality=90, optimize=True)
        print(f"✅ save_tanda_tangan_file - File saved successfully: {filepath}")

        # Kembalikan path relatif dari static folder (uploads/filename.jpg)
        result_path = os.path.join('uploads', filename).replace('\\', '/')
        print(f"✅ save_tanda_tangan_file - Returning path: {result_path}")
        return result_path
    except Exception as e:
        print(f"❌ Error saving tanda tangan file: {e}")
        import traceback
        traceback.print_exc()
        return None

def process_tanda_tangan_for_pdf(tanda_tangan_data, temp_files_list=None):
    """
    Memproses tanda tangan untuk PDF export
    Returns: (RLImage object, error_message)
    """
    if not tanda_tangan_data:
        print("DEBUG: Tanda tangan data is None or empty")
        return None, "Tanda tangan kosong"

    try:
        # Bersihkan whitespace jika string
        if isinstance(tanda_tangan_data, str):
            tanda_tangan_data = tanda_tangan_data.strip()
            if not tanda_tangan_data:
                print("DEBUG: Tanda tangan data is empty after strip")
                return None, "Tanda tangan kosong setelah pembersihan"

        img_data = None

        # Debug: print tipe data
        print(f"DEBUG process_tanda_tangan: Data type: {type(tanda_tangan_data)}")
        if isinstance(tanda_tangan_data, str):
            data_preview = tanda_tangan_data[:100] + "..." if len(tanda_tangan_data) > 100 else tanda_tangan_data
            print(f"DEBUG process_tanda_tangan: Data preview: {data_preview}")
            print(f"DEBUG process_tanda_tangan: Data length: {len(tanda_tangan_data)}")
        else:
            print(f"DEBUG process_tanda_tangan: Data length: {len(str(tanda_tangan_data)) if hasattr(tanda_tangan_data, '__len__') else 'N/A'}")

        if isinstance(tanda_tangan_data, str):
            if tanda_tangan_data.startswith('data:image'):
                print("DEBUG process_tanda_tangan: Detected data:image format")
                try:
                    header, encoded = tanda_tangan_data.split(',', 1)
                    print(f"DEBUG process_tanda_tangan: Header: {header[:50]}")
                    print(f"DEBUG process_tanda_tangan: Encoded length: {len(encoded)}")
                    img_data = base64.b64decode(encoded, validate=True)
                    print(f"DEBUG process_tanda_tangan: Decoded base64, size: {len(img_data)} bytes")
                except Exception as e:
                    print(f"DEBUG process_tanda_tangan: Error decoding data:image: {e}")
                    return None, f"Gagal decode data:image: {str(e)}"
            elif 'uploads/' in tanda_tangan_data or tanda_tangan_data.startswith('static/'):
                print(f"DEBUG process_tanda_tangan: Detected file path: {tanda_tangan_data}")
                path = tanda_tangan_data
                if path.startswith('uploads/'):
                    path = os.path.join(BASE_DIR, 'static', path)
                elif path.startswith('static/'):
                    path = os.path.join(BASE_DIR, path)
                else:
                    # Jika hanya mengandung 'uploads/' di tengah string
                    path = os.path.join(BASE_DIR, 'static', path)
                print(f"DEBUG process_tanda_tangan: Full path: {path}")
                print(f"DEBUG process_tanda_tangan: Path exists: {os.path.exists(path)}")
                if os.path.exists(path):
                    print(f"DEBUG process_tanda_tangan: File exists, reading...")
                    with open(path, 'rb') as f:
                        img_data = f.read()
                    print(f"DEBUG process_tanda_tangan: Read file, size: {len(img_data)} bytes")
                else:
                    return None, f"File tidak ditemukan: {path}"
            else:
                print("DEBUG process_tanda_tangan: Trying direct base64 decode")
                try:
                    # Coba decode base64
                    img_data = base64.b64decode(tanda_tangan_data, validate=True)
                    print(f"DEBUG process_tanda_tangan: Direct base64 decode success, size: {len(img_data)} bytes")
                except Exception as e:
                    print(f"DEBUG process_tanda_tangan: Direct base64 decode failed: {e}")
                    return None, f"Gagal decode base64: {str(e)}"
        else:
            print(f"DEBUG process_tanda_tangan: Unsupported data type: {type(tanda_tangan_data)}")
            return None, f"Format data tidak dikenal: {type(tanda_tangan_data)}"

        if not img_data:
            print("DEBUG process_tanda_tangan: No image data extracted")
            return None, "Tidak ada data gambar"

        if len(img_data) < 100:
            print(f"DEBUG process_tanda_tangan: Image data too small: {len(img_data)} bytes")
            return None, f"Data gambar terlalu kecil: {len(img_data)} bytes"

        print(f"DEBUG process_tanda_tangan: Opening image from {len(img_data)} bytes")
        try:
            img = Image.open(io.BytesIO(img_data))
            print(f"DEBUG process_tanda_tangan: Image opened successfully, mode: {img.mode}, size: {img.size}")
        except Exception as e:
            print(f"DEBUG process_tanda_tangan: Error opening image: {e}")
            return None, f"Gagal membuka gambar: {str(e)}"

        # Convert ke RGBA jika belum untuk manipulasi alpha channel
        if img.mode != 'RGBA':
            print(f"DEBUG process_tanda_tangan: Converting from {img.mode} to RGBA")
            img = img.convert('RGBA')

        # Normalisasi: buat background putih dan goresan hitam
        print("DEBUG process_tanda_tangan: Normalizing signature to black on white using auto inversion + threshold...")
        img_gray = img.convert('L')
        median = ImageStat.Stat(img_gray).median[0]
        mean = ImageStat.Stat(img_gray).mean[0]
        print(f"DEBUG process_tanda_tangan: median={median}, mean={mean}")

        # Jika mayoritas gelap (background gelap), invert terlebih dahulu agar background jadi terang
        if median < 128:
            print("DEBUG process_tanda_tangan: Inverting grayscale because background seems dark")
            img_gray = ImageOps.invert(img_gray)

        # Tetapkan threshold adaptif: base dari median + offset, dibatasi range aman
        stroke_threshold = int(min(230, max(120, median + 40)))
        print(f"DEBUG process_tanda_tangan: stroke_threshold={stroke_threshold}")

        binary = img_gray.point(lambda p: 0 if p < stroke_threshold else 255)
        img = Image.merge('RGB', (binary, binary, binary))
        print(f"DEBUG process_tanda_tangan: Final image mode: {img.mode}, size: {img.size}")

        # Ukuran untuk dimasukkan ke tabel (dalam inch)
        max_img_width_inch = 3.0 * inch
        max_img_height_inch = 1.5 * inch

        # DPI untuk kualitas HD (300 DPI untuk kualitas tinggi)
        DPI = 300

        # Konversi ukuran maksimal ke pixel
        max_img_width_px = int(max_img_width_inch * DPI / 72.0)
        max_img_height_px = int(max_img_height_inch * DPI / 72.0)

        # Hitung ukuran baru dalam pixel dengan mempertahankan aspect ratio
        img_ratio = img.width / img.height
        original_width_px = img.width
        original_height_px = img.height

        if original_width_px > max_img_width_px:
            new_width_px = max_img_width_px
            new_height_px = int(new_width_px / img_ratio)
            if new_height_px > max_img_height_px:
                new_height_px = max_img_height_px
                new_width_px = int(new_height_px * img_ratio)
        elif original_height_px > max_img_height_px:
            new_height_px = max_img_height_px
            new_width_px = int(new_height_px * img_ratio)
        else:
            # Jika gambar terlalu kecil, perbesar minimal ke ukuran yang wajar
            min_width_px = int(1.5 * inch * DPI / 72.0)
            if original_width_px < min_width_px:
                new_width_px = min_width_px
                new_height_px = int(new_width_px / img_ratio)
                if new_height_px > max_img_height_px:
                    new_height_px = max_img_height_px
                    new_width_px = int(new_height_px * img_ratio)
            else:
                new_width_px = original_width_px
                new_height_px = original_height_px

        # Konversi kembali ke inch untuk RLImage
        new_width_inch = new_width_px * 72.0 / DPI
        new_height_inch = new_height_px * 72.0 / DPI

        print(f"DEBUG process_tanda_tangan: Original size: {original_width_px}x{original_height_px} px")
        print(f"DEBUG process_tanda_tangan: Resizing to {new_width_px}x{new_height_px} px ({new_width_inch:.2f}x{new_height_inch:.2f} inches)")

        # Resize dengan LANCZOS untuk kualitas tinggi
        img = img.resize((new_width_px, new_height_px), Image.Resampling.LANCZOS)

        # Simpan ke temporary file dengan kualitas tinggi
        print("DEBUG process_tanda_tangan: Saving to temporary file with high quality...")
        tanda_tangan_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.jpg')
        img.save(tanda_tangan_temp_file.name, format='JPEG', quality=95, optimize=True)
        tanda_tangan_temp_file.close()
        print(f"DEBUG process_tanda_tangan: Saved to: {tanda_tangan_temp_file.name}")

        # Simpan path untuk cleanup jika diperlukan
        if temp_files_list is not None:
            temp_files_list.append(tanda_tangan_temp_file.name)
            print(f"DEBUG process_tanda_tangan: Added to cleanup list")

        print("DEBUG process_tanda_tangan: Creating RLImage...")
        tanda_tangan_img = RLImage(tanda_tangan_temp_file.name, width=new_width_inch, height=new_height_inch)
        print(f"DEBUG process_tanda_tangan: RLImage created successfully, size: {new_width_inch:.2f}x{new_height_inch:.2f} inches")

        return tanda_tangan_img, None

    except Exception as e:
        error_msg = f"Error processing tanda tangan: {str(e)}"
        print(f"DEBUG process_tanda_tangan ERROR: {error_msg}")
        import traceback
        print("DEBUG process_tanda_tangan TRACEBACK:")
        traceback.print_exc()
        return None, error_msg

def normalize_buku_tabungan_path(path):
    """Normalisasi path buku tabungan ke format 'uploads/filename.jpg'"""
    if not path:
        return None
    if 'static/uploads/' in path:
        return path.replace('static/uploads/', 'uploads/')
    elif path.startswith('uploads/'):
        return path
    else:
        return f"uploads/{path.split('/')[-1]}"

def get_biodata_data(user_id, nik=None, nama_kegiatan=None):
    """Mengambil data biodata user dan normalisasi path
    Jika nik dan nama_kegiatan tidak diberikan, ambil data terakhir (untuk pre-fill form)
    Jika user_id None, gunakan nik untuk mencari data"""
    connection = get_db_connection()
    if not connection:
        return None

    try:
        cursor = connection.cursor()

        # Jika user_id None, gunakan nik untuk mencari
        if user_id is None and nik:
            if nama_kegiatan:
                # Ambil data untuk nik dan kegiatan tertentu
                cursor.execute("SELECT * FROM biodata_kegiatan WHERE nik = ? AND TRIM(nama_kegiatan) = TRIM(?) ORDER BY created_at DESC LIMIT 1", (nik, nama_kegiatan))
            else:
                # Ambil data terakhir untuk nik tersebut
                cursor.execute("SELECT * FROM biodata_kegiatan WHERE nik = ? ORDER BY created_at DESC LIMIT 1", (nik,))
        elif user_id is not None:
            if nik and nama_kegiatan:
                # Ambil data untuk nik dan kegiatan tertentu
                # Gunakan TRIM untuk memastikan perbandingan tanpa whitespace
                cursor.execute("SELECT * FROM biodata_kegiatan WHERE user_id = ? AND nik = ? AND TRIM(nama_kegiatan) = TRIM(?)", (user_id, nik, nama_kegiatan))
            elif nama_kegiatan:
                # Ambil data untuk kegiatan tertentu (data terakhir untuk kegiatan tersebut)
                cursor.execute("SELECT * FROM biodata_kegiatan WHERE user_id = ? AND TRIM(nama_kegiatan) = TRIM(?) ORDER BY created_at DESC LIMIT 1", (user_id, nama_kegiatan))
            elif nik:
                # Ambil data untuk nik tertentu (data terakhir untuk nik tersebut)
                cursor.execute("SELECT * FROM biodata_kegiatan WHERE user_id = ? AND nik = ? ORDER BY created_at DESC LIMIT 1", (user_id, nik))
            else:
                # Ambil data terakhir (untuk pre-fill form)
                cursor.execute("SELECT * FROM biodata_kegiatan WHERE user_id = ? ORDER BY created_at DESC LIMIT 1", (user_id,))
        else:
            # Tidak ada user_id dan tidak ada nik, return None
            return None

        biodata = cursor.fetchone()
        if biodata:
            biodata = row_to_dict(biodata)

        if biodata and biodata.get('buku_tabungan_path'):
            biodata['buku_tabungan_path'] = normalize_buku_tabungan_path(biodata['buku_tabungan_path'])

        return biodata
    except:
        return None
    finally:
        if connection:
            cursor.close()
            connection.close()

def get_form_data():
    """Mengambil semua data dari form request"""
    nama_kegiatan = request.form.get('nama_kegiatan', '').strip() if request.form.get('nama_kegiatan') else None
    return {
        'nik': request.form.get('NIK'),
        'nama_lengkap': request.form.get('nama_lengkap'),
        'nip_nippk': request.form.get('nip/nippk'),
        'tempat_lahir': request.form.get('tempat_lahir'),
        'tanggal_lahir': request.form.get('tanggal_lahir'),
        'jenis_kelamin': request.form.get('jenis_kelamin'),
        'agama': request.form.get('agama'),
        'pendidikan_terakhir': request.form.get('pendidikan_terakhir'),
        'jurusan': request.form.get('jurusan'),
        'alamat_domisili': request.form.get('alamat_domisili'),
        'alamat_email': request.form.get('alamat_email'),
        'no_hp': request.form.get('nohp'),
        'npwp': request.form.get('npwp'),
        'status_asn': request.form.get('status_asn'),
        'pangkat_golongan': request.form.get('Pangkat/Golongan'),
        'jabatan': request.form.get('jabatan'),
        'instansi': request.form.get('instansi'),
        'alamat_instansi': request.form.get('alamat_instansi'),
        'kabupaten_kota': request.form.get('kabupaten/kota'),
        'kabko_lainnya': request.form.get('kabko_lainnya') or None,
        'peran': request.form.get('peran'),
        'nama_kegiatan': nama_kegiatan,
        'waktu_pelaksanaan': request.form.get('waktu_pelaksanaan'),
        'tempat_pelaksanaan': request.form.get('tempat_pelaksanaan'),
        'nama_bank': request.form.get('nama_bank'),
        'nama_bank_lainnya': request.form.get('nama_bank_lainnya') or None,
        'no_rekening': request.form.get('no_rekening'),
        'nama_pemilik_rekening': request.form.get('nama_pemilik_rekening'),
        'tanda_tangan': request.form.get('ttd'),
        # Path file existing dari auto-fill (jika pengguna tidak upload ulang)
        'existing_buku_tabungan_path': request.form.get('existing_buku_tabungan_path') or None,
        # Data original untuk update mode
        'original_nama_kegiatan': request.form.get('original_nama_kegiatan') or None,
        'original_user_id': request.form.get('original_user_id') or None
    }

def validate_required_fields(form_data):
    """Validasi semua field wajib"""
    # Daftar field wajib dengan nama field untuk error message
    required_fields_map = {
        'nik': form_data.get('nik'),
        'nama_lengkap': form_data.get('nama_lengkap'),
        'nip_nippk': form_data.get('nip_nippk'),
        'tempat_lahir': form_data.get('tempat_lahir'),
        'tanggal_lahir': form_data.get('tanggal_lahir'),
        'jenis_kelamin': form_data.get('jenis_kelamin'),
        'agama': form_data.get('agama'),
        'pendidikan_terakhir': form_data.get('pendidikan_terakhir'),
        'jurusan': form_data.get('jurusan'),
        'alamat_domisili': form_data.get('alamat_domisili'),
        'alamat_email': form_data.get('alamat_email'),
        'no_hp': form_data.get('no_hp'),
        'npwp': form_data.get('npwp'),
        'status_asn': form_data.get('status_asn'),
        'pangkat_golongan': form_data.get('pangkat_golongan'),
        'jabatan': form_data.get('jabatan'),
        'instansi': form_data.get('instansi'),
        'alamat_instansi': form_data.get('alamat_instansi'),
        'kabupaten_kota': form_data.get('kabupaten_kota'),
        'peran': form_data.get('peran'),
        'nama_kegiatan': form_data.get('nama_kegiatan'),
        'waktu_pelaksanaan': form_data.get('waktu_pelaksanaan'),
        'tempat_pelaksanaan': form_data.get('tempat_pelaksanaan'),
        'nama_bank': form_data.get('nama_bank'),
        'no_rekening': form_data.get('no_rekening'),
        'nama_pemilik_rekening': form_data.get('nama_pemilik_rekening')
    }

    # Cek setiap field wajib
    missing_fields = []
    for field_name, field_value in required_fields_map.items():
        # Cek jika None, empty string, atau whitespace only
        if not field_value or (isinstance(field_value, str) and not field_value.strip()):
            missing_fields.append(field_name)

    if missing_fields:
        print(f"❌ Validasi gagal - Field yang kosong: {', '.join(missing_fields)}")
        return False

    return True

def get_biodata_values(form_data):
    """Mengembalikan tuple nilai untuk query INSERT/UPDATE (tanpa nik, user_id, buku_tabungan_path, tanda_tangan)"""
    return (
        form_data['nama_lengkap'], form_data['nip_nippk'],
        form_data['tempat_lahir'], form_data['tanggal_lahir'], form_data['jenis_kelamin'],
        form_data['agama'], form_data['pendidikan_terakhir'], form_data['jurusan'],
        form_data['alamat_domisili'], form_data['alamat_email'], form_data['no_hp'],
        form_data['npwp'], form_data['status_asn'], form_data['pangkat_golongan'],
        form_data['jabatan'], form_data['instansi'], form_data['alamat_instansi'],
        form_data['kabupaten_kota'], form_data['kabko_lainnya'], form_data['peran'],
        form_data['nama_kegiatan'], form_data['waktu_pelaksanaan'], form_data['tempat_pelaksanaan'],
        form_data['nama_bank'], form_data['nama_bank_lainnya'], form_data['no_rekening'],
        form_data['nama_pemilik_rekening']
    )

@app.route('/')
def index():
    """Halaman landing page"""
    return render_template('index.html')

# ============================================
# SESSION MANAGEMENT HELPERS
# ============================================

def is_logged_in():
    """Check apakah user sudah login"""
    return session.get('logged_in', False)

def get_user_id():
    """Get user ID dari session"""
    return session.get('user_id')

def get_user_role():
    """Get user role dari session"""
    return session.get('user_role', 'user')

def is_admin():
    """Check apakah user adalah admin atau operator"""
    return session.get('is_admin', False)

def get_username():
    """Get username dari session"""
    return session.get('username')

def get_user_display_name():
    """Helper function untuk mendapatkan nama yang akan ditampilkan (nama atau username)"""
    user_nama = session.get('user_nama')
    username = session.get('username')
    return user_nama if user_nama else (username if username else 'Admin')

def clear_session():
    """Clear semua session data"""
    session.clear()
    session.permanent = False

def set_session_data(user, user_role):
    """Set session data dengan konsisten"""
    session.permanent = True
    session['username'] = user.get('username')
    session['logged_in'] = True
    session['user_id'] = user.get('id')
    session['user_role'] = user_role
    session['user_nama'] = user.get('nama') or user.get('username')
    session['is_admin'] = (user_role == 'admin' or user_role == 'operator')

def validate_session():
    """Validate session dan return user info jika valid"""
    if not is_logged_in():
        return None

    user_id = get_user_id()
    if not user_id:
        clear_session()
        return None

    return {
        'user_id': user_id,
        'username': get_username(),
        'user_role': get_user_role(),
        'is_admin': is_admin()
    }

from functools import wraps

def login_required(f):
    """Decorator untuk route yang memerlukan login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not is_logged_in():
            flash('Silakan login terlebih dahulu!', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Decorator untuk route yang memerlukan akses admin"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not is_logged_in():
            flash('Silakan login terlebih dahulu!', 'error')
            return redirect(url_for('login'))
        if not is_admin():
            flash('Anda tidak memiliki akses ke halaman ini!', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

@app.before_request
def refresh_session():
    """Refresh session sebelum setiap request untuk mencegah masalah saat back button"""
    if session.get('logged_in'):
        # Mark session sebagai modified untuk memastikan di-save
        session.permanent = True
        session.modified = True

@app.after_request
def set_cache_headers(response):
    """Set header anti-cache untuk mencegah browser cache halaman yang memerlukan session"""
    # Hanya set anti-cache untuk halaman HTML, bukan untuk file static, PDF, dll
    if response.content_type and 'text/html' in response.content_type:
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        # Vary header untuk memastikan browser tidak cache berdasarkan cookie
        response.headers['Vary'] = 'Cookie'

    # Refresh session setelah request untuk memastikan tetap valid
    if session.get('logged_in'):
        session.permanent = True
        session.modified = True

    return response

@app.context_processor
def inject_user_role():
    """Inject user_role dan user_display_name ke semua template"""
    return dict(
        user_role=get_user_role(),
        user_display_name=get_user_display_name(),
        is_logged_in=is_logged_in(),
        is_admin=is_admin()
    )

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Halaman login"""
    # Jika user sudah login, redirect ke dashboard yang sesuai
    if is_logged_in():
        if is_admin():
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('dashboard'))

    if request.method == 'POST':
        login_type = request.form.get('login_type')
        username = request.form.get('username')
        password = request.form.get('password')

        # Validasi input
        if not login_type or not username or not password:
            flash('Semua field harus diisi!', 'error')
            return render_template('login.html')

        # Validasi login_type
        if login_type not in ['operator', 'admin']:
            flash('Tipe login tidak valid!', 'error')
            return render_template('login.html')

        # Koneksi ke database
        connection = get_db_connection()
        if connection is None:
            flash('Koneksi database gagal! Periksa file database dan konfigurasi.', 'error')
            return render_template('login.html')

        cursor = None
        try:
            cursor = connection.cursor()

            # Cek apakah tabel users ada, jika tidak, inisialisasi database
            try:
                table_exists_result = table_exists(connection, 'users')
                print(f"🔍 Debug login - Tabel users exists: {table_exists_result}")

                if not table_exists_result:
                    print("⚠️  Tabel 'users' tidak ditemukan! Mencoba inisialisasi database...")
                    cursor.close()
                    if connection:
                        connection.close()

                    # Coba inisialisasi database
                    print("🔄 Memanggil init_database()...")
                    init_result = init_database()
                    print(f"🔍 Debug login - Init database result: {init_result}")

                    if init_result:
                        flash('Database berhasil diinisialisasi! Silakan coba login lagi dengan username: admin, password: admin123', 'success')
                    else:
                        flash('Database belum diinisialisasi! Silakan kunjungi /init-db untuk inisialisasi manual.', 'error')
                    return render_template('login.html')

            except sqlite3.Error as e:
                print(f"⚠️  Error saat mengecek tabel: {e}")
                # Jika error saat cek tabel, mungkin tabel tidak ada
                if 'no such table' in str(e).lower() or 'doesn\'t exist' in str(e).lower():
                    print("⚠️  Tabel tidak ada! Mencoba inisialisasi...")
                    cursor.close()
                    if connection:
                        connection.close()
                    if init_database():
                        flash('Database berhasil diinisialisasi! Silakan coba login lagi dengan username: admin, password: admin123', 'success')
                    else:
                        flash('Database belum diinisialisasi! Silakan kunjungi /init-db untuk inisialisasi manual.', 'error')
                    return render_template('login.html')

                # Pastikan cursor masih valid
            if not cursor:
                cursor = connection.cursor()

            # Cari user berdasarkan username
            print(f"🔍 Debug login - Mencari user dengan username: {username}")
            try:
                cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
                user = cursor.fetchone()
                # Convert Row to dict for compatibility
                if user:
                    user = row_to_dict(user)
                print(f"🔍 Debug login - User ditemukan: {user is not None}")
            except sqlite3.Error as query_error:
                # Jika error karena tabel tidak ada, coba inisialisasi
                if 'no such table' in str(query_error).lower() or 'doesn\'t exist' in str(query_error).lower():
                    print(f"⚠️  Error: Tabel users tidak ada")
                    try:
                        cursor.close()
                    except:
                        pass
                    try:
                        if connection:
                            connection.close()
                    except:
                        pass
                    if init_database():
                        flash('Database berhasil diinisialisasi! Silakan coba login lagi dengan username: admin, password: admin123', 'success')
                    else:
                        flash('Database belum diinisialisasi! Silakan kunjungi /init-db untuk inisialisasi manual.', 'error')
                    return render_template('login.html')
                else:
                    # Error lain, re-raise
                    raise

            if user:
                # Verifikasi password (plain text comparison)
                stored_password = user['password']
                password_valid = (stored_password == password)

                print(f"🔍 Debug login - Username: {username}")
                print(f"🔍 Debug login - Password valid: {password_valid}")
                print(f"🔍 Debug login - Stored password: {stored_password[:3]}... (first 3 chars)")
                print(f"🔍 Debug login - Input password: {password[:3]}... (first 3 chars)")
                print(f"🔍 Debug login - Login type: {login_type}")
                print(f"🔍 Debug login - User role from DB: {user.get('role')}")

                if password_valid:
                    # Ambil role user dari database (normalisasi ke lowercase untuk konsistensi)
                    user_role = user.get('role', 'user')
                    if user_role:
                        user_role = user_role.lower().strip()

                    print(f"🔍 Debug login - User role (normalized): {user_role}")

                    # Validasi login_type sesuai dengan role user
                    role_mapping = {
                        'admin': 'admin',
                        'operator': 'operator'
                    }

                    expected_login_type = role_mapping.get(user_role)
                    print(f"🔍 Debug login - Expected login type: {expected_login_type}")

                    if not expected_login_type:
                        print(f"❌ Debug login - Role '{user_role}' tidak memiliki expected_login_type")
                        flash(f'Role "{user_role}" tidak dapat login melalui halaman ini! Hanya Admin dan Operator yang dapat login.', 'error')
                        return render_template('login.html')

                    if login_type != expected_login_type:
                        print(f"❌ Debug login - Login type mismatch: {login_type} != {expected_login_type}")
                        if user_role == 'admin':
                            flash('Admin hanya bisa login sebagai Admin!', 'error')
                        elif user_role == 'operator':
                            flash('Operator hanya bisa login sebagai Operator!', 'error')
                        else:
                            flash('User dengan role ini tidak dapat login melalui halaman ini!', 'error')
                        return render_template('login.html')

                    # Set session data dengan helper function
                    set_session_data(user, user_role)

                    # Redirect berdasarkan role
                    if user_role == 'admin' or user_role == 'operator':
                        if user_role == 'operator':
                            flash('Login sebagai Operator berhasil!', 'success')
                        else:
                            flash('Login sebagai Admin berhasil!', 'success')
                        return redirect(url_for('admin_dashboard'))
                    else:
                        flash('Login sebagai Peserta berhasil!', 'success')
                        return redirect(url_for('dashboard'))
                else:
                    flash('Password salah!', 'error')
            else:
                flash('Username tidak ditemukan!', 'error')

        except sqlite3.Error as e:
            print(f"❌ Error during login (SQLite Error): {e}")
            import traceback
            traceback.print_exc()

            # Jika error karena tabel tidak ada, coba inisialisasi
            if 'no such table' in str(e).lower() or 'doesn\'t exist' in str(e).lower():
                print("⚠️  Tabel tidak ada! Mencoba inisialisasi database...")
                try:
                    if init_database():
                        flash('Database berhasil diinisialisasi! Silakan coba login lagi dengan username: admin, password: admin123', 'success')
                    else:
                        flash('Database belum diinisialisasi! Silakan kunjungi /init-db untuk inisialisasi manual.', 'error')
                except Exception as init_error:
                    print(f"❌ Error saat inisialisasi: {init_error}")
                    flash('Gagal menginisialisasi database! Silakan kunjungi /init-db untuk inisialisasi manual.', 'error')
            else:
                flash(f'Terjadi kesalahan saat login: {str(e)}', 'error')
        except Exception as e:
            print(f"❌ Error during login (General Error): {e}")
            import traceback
            traceback.print_exc()
            flash(f'Terjadi kesalahan saat login: {str(e)}', 'error')
        finally:
            # Tutup cursor dan connection dengan aman
            if cursor:
                try:
                    cursor.close()
                except:
                    pass
            if connection:
                try:
                    if connection is not None:
                        connection.close()
                except:
                    pass

    # Ambil parameter type dari URL untuk pre-select login type
    login_type_param = request.args.get('type', '')
    return render_template('login.html', pre_select_type=login_type_param)

@app.route('/register', methods=['GET', 'POST'])
def register():
    """Halaman register - Fitur registrasi telah dinonaktifkan"""
    flash('Fitur registrasi tidak tersedia. Silakan hubungi admin untuk membuat akun.', 'info')
    return redirect(url_for('login'))

def user_has_biodata(user_id):
    """Cek apakah user sudah punya data biodata"""
    if not user_id:
        return False
    connection = get_db_connection()
    if not connection:
        return False
    try:
        cursor = connection.cursor()
        cursor.execute("SELECT COUNT(*) FROM biodata_kegiatan WHERE user_id = ?", (user_id,))
        return cursor.fetchone()[0] > 0
    except:
        return False
    finally:
        if connection:
            cursor.close()
            connection.close()

@app.route('/dashboard', methods=['GET', 'POST'])
@app.route('/user/dashboard', methods=['GET', 'POST'])
@login_required
def dashboard():
    """Halaman dashboard user setelah login"""
    # Jika admin mencoba mengakses dashboard user, arahkan ke halaman admin
    if is_admin():
        return redirect(url_for('admin_dashboard'))

    user_id = get_user_id()
    has_data = user_has_biodata(user_id)

    # Ambil data biodata jika ada
    biodata = None
    if has_data and user_id:
        biodata = get_biodata_data(user_id)

    # Ambil daftar kegiatan dari database untuk dropdown
    kegiatan_list = []
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor()
            cursor.execute("""
                SELECT DISTINCT nama_kegiatan, waktu_pelaksanaan, tempat_pelaksanaan
                FROM kegiatan_master
                WHERE nama_kegiatan IS NOT NULL AND TRIM(nama_kegiatan) != ''
                ORDER BY nama_kegiatan ASC
            """)
            kegiatan_list = cursor.fetchall()
            print(f"🔍 Debug dashboard - Jumlah kegiatan ditemukan: {len(kegiatan_list)}")
        except sqlite3.Error as e:
            print(f"❌ Error fetching kegiatan: {e}")
            import traceback
            traceback.print_exc()
        finally:
            if connection is not None:
                cursor.close()
                connection.close()

    # Handle POST request untuk update data dari dashboard
    if request.method == 'POST' and has_data:
        # Redirect ke tambah_data untuk handle update
        return redirect(url_for('tambah_data'))

    # Redirect ke daftar kegiatan jika user sudah punya data, atau tambah_data jika belum
    if has_data:
        return redirect(url_for('daftar_kegiatan'))
    else:
        return redirect(url_for('tambah_data'))

@app.route('/ubah-password', methods=['GET', 'POST'])
@login_required
def change_password():
    """Fitur ubah password untuk admin dan user"""
    user_id = get_user_id()
    is_admin_user = is_admin()

    if not user_id:
        return redirect(url_for('login'))

    if request.method == 'POST':
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        # Validasi input
        if not current_password or not new_password or not confirm_password:
            flash('Semua field harus diisi!', 'error')
            return render_template('user/change-password.html', is_admin=is_admin_user)

        if len(new_password) < 6:
            flash('Password baru minimal 6 karakter!', 'error')
            return render_template('user/change-password.html', is_admin=is_admin_user)

        if new_password != confirm_password:
            flash('Password baru dan konfirmasi password tidak cocok!', 'error')
            return render_template('user/change-password.html', is_admin=is_admin_user)

        connection = get_db_connection()
        if connection is None:
            flash('Koneksi database gagal!', 'error')
            return render_template('user/change-password.html', is_admin=is_admin_user)

        try:
            cursor = connection.cursor()
            cursor.execute("SELECT password FROM users WHERE id = ?", (user_id,))
            user = cursor.fetchone()
            if user:
                user = row_to_dict(user)

            if not user:
                flash('User tidak ditemukan!', 'error')
                return render_template('user/change-password.html', is_admin=is_admin_user)

            stored_password = user['password']

            # Verifikasi password lama (plain text comparison)
            if stored_password != current_password:
                flash('Password lama salah!', 'error')
                return render_template('user/change-password.html', is_admin=is_admin_user)

            # Simpan password baru sebagai plain text
            cursor.execute("UPDATE users SET password = ? WHERE id = ?", (new_password, user_id))
            connection.commit()

            flash('Password berhasil diubah!', 'success')
            return redirect(url_for('admin_dashboard' if is_admin_user else 'dashboard'))

        except sqlite3.Error as e:
            print(f"Error during change_password: {e}")
            flash(f'Terjadi kesalahan saat mengubah password: {str(e)}', 'error')
        finally:
            if connection is not None:
                cursor.close()
                connection.close()

    return render_template('user/change-password.html', is_admin=is_admin_user)

def get_or_create_user_by_nik(nik, nama_lengkap=None, email=None):
    """Mendapatkan atau membuat user berdasarkan NIK (untuk peserta tanpa login)"""
    connection = get_db_connection()
    if not connection:
        print("❌ Error: Tidak dapat membuat koneksi ke database")
        return None

    cursor = None
    try:
        cursor = connection.cursor()

        # Cek apakah sudah ada user dengan NIK ini di biodata_kegiatan
        cursor.execute("SELECT DISTINCT user_id FROM biodata_kegiatan WHERE nik = ? LIMIT 1", (nik,))
        existing = cursor.fetchone()
        if existing:
            existing = row_to_dict(existing)

        if existing and existing.get('user_id'):
            # User sudah ada, return user_id
            print(f"✅ User dengan NIK {nik} sudah ada, user_id: {existing['user_id']}")
            return existing['user_id']

        # Jika belum ada, buat user baru dengan username berdasarkan NIK
        username = f"peserta_{nik}"
        # Cek apakah username sudah ada
        cursor.execute("SELECT id FROM users WHERE username = ?", (username,))
        user_existing = cursor.fetchone()
        if user_existing:
            user_existing = row_to_dict(user_existing)

        if user_existing:
            print(f"✅ User dengan username {username} sudah ada, user_id: {user_existing['id']}")
            return user_existing['id']

        # Cek apakah kolom nama dan email ada di tabel users
        has_nama = column_exists(connection, 'users', 'nama')
        has_email = column_exists(connection, 'users', 'email')

        # Buat user baru dengan role 'user' (peserta)
        password = nik  # Password default adalah NIK (bisa diubah nanti jika perlu)

        # Build query berdasarkan kolom yang tersedia
        if has_nama and has_email:
            query = "INSERT INTO users (username, password, role, nama, email) VALUES (?, ?, 'user', ?, ?)"
            params = (username, password, nama_lengkap or username, email or '')
        elif has_nama:
            query = "INSERT INTO users (username, password, role, nama) VALUES (?, ?, 'user', ?)"
            params = (username, password, nama_lengkap or username)
        else:
            query = "INSERT INTO users (username, password, role) VALUES (?, ?, 'user')"
            params = (username, password)

        print(f"🔍 Debug get_or_create_user_by_nik - Creating user: username={username}, nama={nama_lengkap or username}, email={email or 'N/A'}")
        cursor.execute(query, params)
        connection.commit()

        # Ambil user_id yang baru dibuat
        user_id = cursor.lastrowid
        print(f"✅ User baru berhasil dibuat dengan user_id: {user_id}")
        return user_id

    except Exception as e:
        print(f"❌ Error in get_or_create_user_by_nik: {e}")
        import traceback
        traceback.print_exc()
        # Rollback jika ada error
        try:
            if connection:
                connection.rollback()
        except:
            pass
        return None
    finally:
        if cursor:
            try:
                cursor.close()
            except:
                pass
        if connection and connection is not None:
            connection.close()

def check_nik_exists(nik, exclude_user_id=None):
    """Cek apakah NIK sudah terdaftar"""
    connection = get_db_connection()
    if not connection:
        return False

    cursor = None
    try:
        cursor = connection.cursor()
        if exclude_user_id:
            cursor.execute("SELECT COUNT(*) FROM biodata_kegiatan WHERE nik = ? AND user_id != ?", (nik, exclude_user_id))
        else:
            cursor.execute("SELECT COUNT(*) FROM biodata_kegiatan WHERE nik = ?", (nik,))
        result = cursor.fetchone()
        return result[0] > 0 if result else False
    except Exception as e:
        print(f"Error in check_nik_exists: {e}")
        return False
    finally:
        if cursor:
            try:
                cursor.close()
            except:
                pass
        if connection and connection is not None:
            connection.close()

def insert_biodata_data(form_data, user_id, buku_tabungan_path=None):
    """Menyimpan data biodata baru (hanya INSERT, tidak UPDATE)
    User bisa memiliki banyak data untuk kegiatan yang berbeda"""
    connection = get_db_connection()
    if not connection:
        return False, 'Koneksi database gagal!'

    cursor = None
    try:
        cursor = connection.cursor()

        # Validasi buku_tabungan_path dan tanda_tangan
        # (Validasi sudah dilakukan di route, jadi di sini hanya double check)
        if not buku_tabungan_path:
            return False, 'File buku tabungan wajib diupload!'

        if not form_data.get('tanda_tangan'):
            return False, 'Tanda tangan wajib diisi!'

        # Cek apakah kombinasi user_id + nama_kegiatan sudah ada
        # User bisa punya banyak data, tapi tidak boleh duplikat untuk kegiatan yang sama
        # Gunakan TRIM untuk memastikan perbandingan tanpa whitespace
        print(f"🔍 Debug insert_biodata_data - Mengecek duplikat: user_id={user_id}, nama_kegiatan='{form_data['nama_kegiatan']}'")
        cursor.execute("""
            SELECT id, nik, nama_lengkap FROM biodata_kegiatan
            WHERE user_id = ? AND TRIM(nama_kegiatan) = TRIM(?)
            LIMIT 1
        """, (user_id, form_data['nama_kegiatan']))
        existing = cursor.fetchone()

        if existing:
            print(f"⚠️ Warning insert_biodata_data - Data duplikat ditemukan: id={existing[0]}, nik={existing[1]}, nama={existing[2]}")
            print(f"⚠️ User mencoba insert dengan: nik={form_data['nik']}, nama_kegiatan='{form_data['nama_kegiatan']}', user_id={user_id}")
            # Cek apakah NIK di existing sama dengan NIK yang diinput
            existing_nik = existing[1]
            if existing_nik != form_data['nik']:
                # NIK berbeda - user mengubah NIK, hapus data lama dan buat data baru
                print(f"🔄 NIK berbeda terdeteksi! Existing NIK: {existing_nik}, Input NIK: {form_data['nik']}")
                print(f"🔄 Menghapus data lama dan membuat data baru dengan NIK yang berbeda")
                # Hapus data lama dengan NIK dan nama_kegiatan yang sama
                cursor.execute("""
                    DELETE FROM biodata_kegiatan
                    WHERE user_id = ? AND TRIM(nama_kegiatan) = TRIM(?) AND nik = ?
                """, (user_id, form_data['nama_kegiatan'], existing_nik))
                deleted_rows = cursor.rowcount
                print(f"✅ Data lama dengan NIK {existing_nik} telah dihapus ({deleted_rows} row)")
                # Lanjutkan ke insert data baru
            else:
                # NIK sama - benar-benar duplikat
                return False, f'Anda sudah memiliki data untuk kegiatan "{form_data["nama_kegiatan"]}".'

        # Tidak perlu validasi NIK - 1 NIK bisa digunakan untuk beberapa kegiatan berbeda
        # Validasi utama adalah kombinasi user_id + nama_kegiatan (sudah dicek di atas)
        values = get_biodata_values(form_data)

        # Insert query
        query = """INSERT INTO biodata_kegiatan (
            nik, user_id, nama_lengkap, nip_nippk, tempat_lahir, tanggal_lahir,
            jenis_kelamin, agama, pendidikan_terakhir, jurusan,
            alamat_domisili, alamat_email, no_hp, npwp, status_asn,
            pangkat_golongan, jabatan, instansi, alamat_instansi,
            kabupaten_kota, kabko_lainnya, peran, nama_kegiatan,
            waktu_pelaksanaan, tempat_pelaksanaan, nama_bank,
            nama_bank_lainnya, no_rekening, nama_pemilik_rekening,
            buku_tabungan_path, tanda_tangan
        ) VALUES (
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
        )"""

        # Pastikan tanda tangan disimpan sebagai file, bukan base64
        tanda_tangan_value = form_data.get('tanda_tangan')
        print(f"🔍 Inserting data for user_id: {user_id}, kegiatan: {form_data['nama_kegiatan']}")
        print(f"🔍 NIK: {form_data['nik']}")
        print(f"🔍 buku_tabungan_path: {buku_tabungan_path}")
        print(f"🔍 tanda_tangan value: {tanda_tangan_value}")
        print(f"🔍 tanda_tangan type: {type(tanda_tangan_value)}")
        print(f"🔍 tanda_tangan length: {len(str(tanda_tangan_value)) if tanda_tangan_value else 0}")

        # Cek apakah masih base64 atau sudah berupa path file
        if tanda_tangan_value and not ('uploads/' in str(tanda_tangan_value) or str(tanda_tangan_value).startswith('static/')):
            # Masih base64, simpan sebagai file
            print(f"🔍 insert_biodata_data: Tanda tangan masih base64, menyimpan sebagai file...")
            tanda_tangan_path = save_tanda_tangan_file(tanda_tangan_value, form_data['nik'])
            if tanda_tangan_path:
                tanda_tangan_value = tanda_tangan_path
                print(f"🔍 insert_biodata_data: Tanda tangan disimpan sebagai file: {tanda_tangan_path}")
            else:
                print(f"🔍 insert_biodata_data: Gagal menyimpan tanda tangan sebagai file, menggunakan base64")
        elif tanda_tangan_value:
            # Sudah berupa path file, normalisasi saja
            tanda_tangan_value = normalize_buku_tabungan_path(tanda_tangan_value)
            print(f"🔍 insert_biodata_data: Tanda tangan sudah berupa path: {tanda_tangan_value}")

        cursor.execute(query, (form_data['nik'], user_id) + values + (buku_tabungan_path, tanda_tangan_value))
        connection.commit()
        print(f"✅ Data berhasil diinsert untuk user_id: {user_id}, kegiatan: {form_data['nama_kegiatan']}")
        return True, 'Data berhasil ditambahkan!'

    except sqlite3.Error as e:
        if connection:
            connection.rollback()
        print(f"❌ Error inserting data: {e}")
        import traceback
        traceback.print_exc()
        return False, f'Terjadi kesalahan saat menyimpan data: {str(e)}'
    except Exception as e:
        if connection:
            connection.rollback()
        print(f"❌ Unexpected error in insert_biodata_data: {e}")
        import traceback
        traceback.print_exc()
        return False, f'Terjadi kesalahan tidak terduga: {str(e)}'
    finally:
        if cursor:
            try:
                cursor.close()
            except:
                pass
        if connection and connection is not None:
            connection.close()

def save_biodata_data(form_data, user_id, buku_tabungan_path=None):
    """Menyimpan atau update data biodata"""
    connection = get_db_connection()
    if not connection:
        return False, 'Koneksi database gagal!'

    cursor = None
    try:
        cursor = connection.cursor()

        # Ambil nama_kegiatan dari form_data
        nama_kegiatan = form_data.get('nama_kegiatan', '').strip()
        if not nama_kegiatan:
            return False, 'Nama kegiatan wajib diisi!'

        # Untuk edit: gunakan old_nama_kegiatan jika ada (dari data existing)
        # Untuk insert: cek apakah sudah ada data untuk kegiatan ini
        old_nama_kegiatan = form_data.get('old_nama_kegiatan', '').strip()

        # Normalisasi untuk perbandingan (pastikan konsisten dengan TRIM di database)
        old_nama_kegiatan_normalized = old_nama_kegiatan.strip() if old_nama_kegiatan else ''
        nama_kegiatan_normalized = nama_kegiatan.strip()

        if old_nama_kegiatan_normalized:
            # Ini adalah edit - cari data berdasarkan old_nama_kegiatan
            cursor.execute("""
                SELECT nik, nama_kegiatan FROM biodata_kegiatan
                WHERE user_id = ? AND TRIM(nama_kegiatan) = TRIM(?)
                LIMIT 1
            """, (user_id, old_nama_kegiatan))
        else:
            # Ini mungkin insert atau edit tanpa old_nama_kegiatan
            # Cek apakah user sudah punya data untuk kegiatan ini
            cursor.execute("""
                SELECT nik, nama_kegiatan FROM biodata_kegiatan
                WHERE user_id = ? AND TRIM(nama_kegiatan) = TRIM(?)
                LIMIT 1
            """, (user_id, nama_kegiatan))

        existing = cursor.fetchone()

        values = get_biodata_values(form_data)

        # Jika ada old_nama_kegiatan dan nama kegiatan berbeda, hapus data lama dan buat data baru (replace)
        if old_nama_kegiatan_normalized and nama_kegiatan_normalized != old_nama_kegiatan_normalized:
            # User mengubah nama_kegiatan - hapus data lama, lalu buat data baru (replace)
            print(f"✅ Nama kegiatan diubah: '{old_nama_kegiatan_normalized}' -> '{nama_kegiatan_normalized}' - Menghapus data lama dan membuat data baru (replace)")

            # Cek apakah nama_kegiatan baru sudah dimiliki user
            cursor.execute("""
                SELECT id FROM biodata_kegiatan
                WHERE user_id = ? AND TRIM(nama_kegiatan) = TRIM(?)
                LIMIT 1
            """, (user_id, nama_kegiatan))
            duplicate_check = cursor.fetchone()

            if duplicate_check:
                return False, f'Anda sudah memiliki data untuk kegiatan "{nama_kegiatan}".  '

            # Hapus data lama dengan old_nama_kegiatan
            cursor.execute("""
                DELETE FROM biodata_kegiatan
                WHERE user_id = ? AND TRIM(nama_kegiatan) = TRIM(?)
            """, (user_id, old_nama_kegiatan_normalized))
            deleted_rows = cursor.rowcount
            print(f"✅ Data lama untuk kegiatan '{old_nama_kegiatan_normalized}' telah dihapus ({deleted_rows} row)")

            # Set existing = None agar masuk ke blok INSERT (bukan UPDATE)
            existing = None

        if existing:
            old_nik = existing[0]
            existing_nama_kegiatan = existing[1]

            # Jika old_nama_kegiatan tidak ada, set dari existing_nama_kegiatan untuk konsistensi
            if not old_nama_kegiatan_normalized:
                old_nama_kegiatan = existing_nama_kegiatan.strip()
                old_nama_kegiatan_normalized = old_nama_kegiatan
            else:
                # Nama kegiatan sama - UPDATE data yang ada
                print(f"✅ Nama kegiatan sama: '{nama_kegiatan_normalized}' - Mengupdate data yang ada")

                # Tidak perlu validasi NIK - 1 NIK bisa digunakan untuk beberapa kegiatan berbeda
                # Validasi utama adalah kombinasi user_id + nama_kegiatan (sudah dicek di atas)

            # Update query - update data untuk kegiatan spesifik ini
            # Gunakan old_nama_kegiatan sebagai identifier untuk WHERE clause
            identifier_nama_kegiatan = old_nama_kegiatan_normalized if old_nama_kegiatan_normalized else existing_nama_kegiatan.strip()

            # Ambil tanda_tangan yang sudah ada jika tidak ada yang baru
            if not form_data.get('tanda_tangan'):
                cursor.execute("""
                    SELECT tanda_tangan FROM biodata_kegiatan
                    WHERE user_id = ? AND TRIM(nama_kegiatan) = TRIM(?)
                    LIMIT 1
                """, (user_id, identifier_nama_kegiatan))
                existing_ttd = cursor.fetchone()
                if existing_ttd and existing_ttd[0]:
                    # Jika masih base64, simpan sebagai file
                    existing_ttd_value = existing_ttd[0]
                    if existing_ttd_value and not ('uploads/' in str(existing_ttd_value) or str(existing_ttd_value).startswith('static/')):
                        # Masih base64, simpan sebagai file
                        tanda_tangan_path = save_tanda_tangan_file(existing_ttd_value, form_data['nik'])
                        if tanda_tangan_path:
                            form_data['tanda_tangan'] = tanda_tangan_path
                        else:
                            form_data['tanda_tangan'] = existing_ttd_value
                    else:
                        form_data['tanda_tangan'] = normalize_buku_tabungan_path(existing_ttd_value) if existing_ttd_value else None

            if buku_tabungan_path:
                # Ada file baru, update termasuk buku_tabungan_path
                query = """UPDATE biodata_kegiatan SET
                    nik = ?, nama_lengkap = ?, nip_nippk = ?, tempat_lahir = ?, tanggal_lahir = ?,
                    jenis_kelamin = ?, agama = ?, pendidikan_terakhir = ?, jurusan = ?,
                    alamat_domisili = ?, alamat_email = ?, no_hp = ?, npwp = ?, status_asn = ?,
                    pangkat_golongan = ?, jabatan = ?, instansi = ?, alamat_instansi = ?,
                    kabupaten_kota = ?, kabko_lainnya = ?, peran = ?, nama_kegiatan = ?,
                    waktu_pelaksanaan = ?, tempat_pelaksanaan = ?, nama_bank = ?,
                    nama_bank_lainnya = ?, no_rekening = ?, nama_pemilik_rekening = ?,
                    buku_tabungan_path = ?, tanda_tangan = ?
                    WHERE user_id = ? AND TRIM(nama_kegiatan) = TRIM(?)"""
                tanda_tangan_update = form_data.get('tanda_tangan')
                print(f"🔍 Debug save_biodata_data UPDATE (dengan buku_tabungan) - tanda_tangan: {tanda_tangan_update}")
                cursor.execute(query, (form_data['nik'],) + values + (buku_tabungan_path, tanda_tangan_update, user_id, identifier_nama_kegiatan))
            else:
                # Tidak ada file baru, update tanpa mengubah buku_tabungan_path
                query = """UPDATE biodata_kegiatan SET
                    nik = ?, nama_lengkap = ?, nip_nippk = ?, tempat_lahir = ?, tanggal_lahir = ?,
                    jenis_kelamin = ?, agama = ?, pendidikan_terakhir = ?, jurusan = ?,
                    alamat_domisili = ?, alamat_email = ?, no_hp = ?, npwp = ?, status_asn = ?,
                    pangkat_golongan = ?, jabatan = ?, instansi = ?, alamat_instansi = ?,
                    kabupaten_kota = ?, kabko_lainnya = ?, peran = ?, nama_kegiatan = ?,
                    waktu_pelaksanaan = ?, tempat_pelaksanaan = ?, nama_bank = ?,
                    nama_bank_lainnya = ?, no_rekening = ?, nama_pemilik_rekening = ?,
                    tanda_tangan = ?
                    WHERE user_id = ? AND TRIM(nama_kegiatan) = TRIM(?)"""
                tanda_tangan_update = form_data.get('tanda_tangan')
                print(f"🔍 Debug save_biodata_data UPDATE (tanpa buku_tabungan) - tanda_tangan: {tanda_tangan_update}")
                cursor.execute(query, (form_data['nik'],) + values + (tanda_tangan_update, user_id, identifier_nama_kegiatan))

            # Cek apakah update berhasil (ada row yang terupdate)
            rows_affected = cursor.rowcount
            connection.commit()

            if rows_affected == 0:
                # Tidak ada row yang terupdate - mungkin data tidak ditemukan
                return False, 'Data tidak ditemukan atau tidak ada perubahan yang perlu disimpan.'

            return True, 'Data berhasil diperbarui!'
        else:
            # Insert baru - cek apakah user sudah punya data untuk kegiatan ini
            # Validasi: User tidak boleh memiliki data duplikat untuk kegiatan yang sama
            cursor.execute("""
                SELECT id FROM biodata_kegiatan
                WHERE user_id = ? AND TRIM(nama_kegiatan) = TRIM(?)
                LIMIT 1
            """, (user_id, nama_kegiatan))
            duplicate_check = cursor.fetchone()

            if duplicate_check:
                return False, f'Anda sudah memiliki data untuk kegiatan "{nama_kegiatan}".  '

            # Tidak perlu validasi NIK - 1 NIK bisa digunakan untuk beberapa kegiatan berbeda
            # Validasi utama adalah kombinasi user_id + nama_kegiatan (sudah dicek di atas)

            # Insert query
            query = """INSERT INTO biodata_kegiatan (
                nik, user_id, nama_lengkap, nip_nippk, tempat_lahir, tanggal_lahir,
                jenis_kelamin, agama, pendidikan_terakhir, jurusan,
                alamat_domisili, alamat_email, no_hp, npwp, status_asn,
                pangkat_golongan, jabatan, instansi, alamat_instansi,
                kabupaten_kota, kabko_lainnya, peran, nama_kegiatan,
                waktu_pelaksanaan, tempat_pelaksanaan, nama_bank,
                nama_bank_lainnya, no_rekening, nama_pemilik_rekening,
                buku_tabungan_path, tanda_tangan
            ) VALUES (
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
            )"""
            tanda_tangan_value = form_data.get('tanda_tangan')
            print(f"🔍 Debug save_biodata_data INSERT - tanda_tangan_value: {tanda_tangan_value}")
            print(f"🔍 Debug save_biodata_data INSERT - tanda_tangan type: {type(tanda_tangan_value)}")
            print(f"🔍 Debug save_biodata_data INSERT - buku_tabungan_path: {buku_tabungan_path}")
            print(f"🔍 Debug save_biodata_data INSERT - NIK: {form_data['nik']}, user_id: {user_id}")
            cursor.execute(query, (form_data['nik'], user_id) + values + (buku_tabungan_path, tanda_tangan_value))
            connection.commit()
            print(f"✅ save_biodata_data INSERT - Data berhasil disimpan dengan tanda_tangan: {tanda_tangan_value}")
            return True, 'Data berhasil ditambahkan!'

    except sqlite3.Error as e:
        if connection:
            connection.rollback()
        print(f"❌ Error saving data: {e}")
        return False, f'Terjadi kesalahan saat menyimpan data: {str(e)}'
    finally:
        if cursor:
            try:
                cursor.close()
            except:
                pass
        if connection and connection is not None:
            connection.close()

def admin_update_biodata(form_data, nik, nama_kegiatan, buku_tabungan_path=None):
    """Update biodata untuk admin berdasarkan NIK dan nama_kegiatan"""
    connection = get_db_connection()
    if not connection:
        return False, 'Koneksi database gagal!'

    cursor = None
    try:
        cursor = connection.cursor()

        # Ambil data existing untuk mendapatkan user_id dan old_nik
        cursor.execute("""
            SELECT user_id, nik FROM biodata_kegiatan
            WHERE nik = ? AND TRIM(nama_kegiatan) = TRIM(?)
            LIMIT 1
        """, (nik, nama_kegiatan))
        existing = cursor.fetchone()

        if not existing:
            return False, 'Data tidak ditemukan!'

        user_id = existing[0]
        old_nik = existing[1]

        # Ambil nama_kegiatan baru dari form_data
        new_nama_kegiatan = form_data.get('nama_kegiatan', '').strip()
        old_nama_kegiatan = nama_kegiatan.strip()

        # Validasi: Jika admin mengubah nama_kegiatan menjadi nama kegiatan yang sudah dimiliki user
        # (kecuali jika mengubah ke nama_kegiatan yang sama dengan yang sedang diedit)
        if new_nama_kegiatan and new_nama_kegiatan != old_nama_kegiatan:
            # Admin mengubah nama_kegiatan - cek apakah nama_kegiatan baru sudah dimiliki user
            cursor.execute("""
                SELECT id FROM biodata_kegiatan
                WHERE user_id = ? AND TRIM(nama_kegiatan) = TRIM(?) AND TRIM(nama_kegiatan) != TRIM(?)
                LIMIT 1
            """, (user_id, new_nama_kegiatan, old_nama_kegiatan))
            duplicate_check = cursor.fetchone()

            if duplicate_check:
                return False, f'Peserta sudah memiliki data untuk kegiatan "{new_nama_kegiatan}".  '

        # Tidak perlu validasi NIK - 1 NIK bisa digunakan untuk beberapa kegiatan berbeda
        # Validasi utama adalah kombinasi user_id + nama_kegiatan (sudah dicek di atas)

        values = get_biodata_values(form_data)

        # Pastikan tanda tangan disimpan sebagai file, bukan base64
        tanda_tangan_to_save = form_data.get('tanda_tangan', '')
        print(f"DEBUG admin_update_biodata: tanda_tangan to save - exists: {tanda_tangan_to_save is not None}, length: {len(tanda_tangan_to_save) if tanda_tangan_to_save else 0}")
        if tanda_tangan_to_save:
            print(f"DEBUG admin_update_biodata: tanda_tangan preview: {tanda_tangan_to_save[:50]}...")
            # Cek apakah masih base64 atau sudah berupa path file
            if not ('uploads/' in str(tanda_tangan_to_save) or str(tanda_tangan_to_save).startswith('static/')):
                # Masih base64, simpan sebagai file
                print(f"DEBUG admin_update_biodata: Tanda tangan masih base64, menyimpan sebagai file...")
                tanda_tangan_path = save_tanda_tangan_file(tanda_tangan_to_save, form_data['nik'])
                if tanda_tangan_path:
                    tanda_tangan_to_save = tanda_tangan_path
                    print(f"DEBUG admin_update_biodata: Tanda tangan disimpan sebagai file: {tanda_tangan_path}")
                else:
                    print(f"DEBUG admin_update_biodata: Gagal menyimpan tanda tangan sebagai file, menggunakan base64")
            else:
                # Sudah berupa path file, normalisasi saja
                tanda_tangan_to_save = normalize_buku_tabungan_path(tanda_tangan_to_save)
                print(f"DEBUG admin_update_biodata: Tanda tangan sudah berupa path: {tanda_tangan_to_save}")

        # Update query berdasarkan NIK dan nama_kegiatan (lama)
        if buku_tabungan_path:
            # Ada file baru, update termasuk buku_tabungan_path
            query = """UPDATE biodata_kegiatan SET
                nik = ?, nama_lengkap = ?, nip_nippk = ?, tempat_lahir = ?, tanggal_lahir = ?,
                jenis_kelamin = ?, agama = ?, pendidikan_terakhir = ?, jurusan = ?,
                alamat_domisili = ?, alamat_email = ?, no_hp = ?, npwp = ?, status_asn = ?,
                pangkat_golongan = ?, jabatan = ?, instansi = ?, alamat_instansi = ?,
                kabupaten_kota = ?, kabko_lainnya = ?, peran = ?, nama_kegiatan = ?,
                waktu_pelaksanaan = ?, tempat_pelaksanaan = ?, nama_bank = ?,
                nama_bank_lainnya = ?, no_rekening = ?, nama_pemilik_rekening = ?,
                buku_tabungan_path = ?, tanda_tangan = ?
                WHERE nik = ? AND TRIM(nama_kegiatan) = TRIM(?)"""
            cursor.execute(query, (form_data['nik'],) + values + (buku_tabungan_path, tanda_tangan_to_save, nik, nama_kegiatan))
        else:
            # Tidak ada file baru, update tanpa mengubah buku_tabungan_path
            query = """UPDATE biodata_kegiatan SET
                nik = ?, nama_lengkap = ?, nip_nippk = ?, tempat_lahir = ?, tanggal_lahir = ?,
                jenis_kelamin = ?, agama = ?, pendidikan_terakhir = ?, jurusan = ?,
                alamat_domisili = ?, alamat_email = ?, no_hp = ?, npwp = ?, status_asn = ?,
                pangkat_golongan = ?, jabatan = ?, instansi = ?, alamat_instansi = ?,
                kabupaten_kota = ?, kabko_lainnya = ?, peran = ?, nama_kegiatan = ?,
                waktu_pelaksanaan = ?, tempat_pelaksanaan = ?, nama_bank = ?,
                nama_bank_lainnya = ?, no_rekening = ?, nama_pemilik_rekening = ?,
                tanda_tangan = ?
                WHERE nik = ? AND TRIM(nama_kegiatan) = TRIM(?)"""
            cursor.execute(query, (form_data['nik'],) + values + (tanda_tangan_to_save, nik, nama_kegiatan))

        connection.commit()
        print(f"DEBUG admin_update_biodata: Update successful for NIK: {nik}, kegiatan: {nama_kegiatan}")
        return True, 'Data berhasil diperbarui!'

    except sqlite3.Error as e:
        if connection:
            connection.rollback()
        print(f"❌ Error updating biodata: {e}")
        return False, f'Terjadi kesalahan saat menyimpan data: {str(e)}'
    finally:
        if cursor:
            try:
                cursor.close()
            except:
                pass
        if connection and connection is not None:
            connection.close()

@app.route('/tambah-data', methods=['GET', 'POST'])
def tambah_data():
    """Halaman tambah data biodata kegiatan (bisa diakses tanpa login untuk peserta baru)"""
    # Halaman ini bisa diakses tanpa login
    # Jika user sudah login, gunakan user_id dari session
    # Jika belum login, akan dibuat user baru saat submit form
    user_id = get_user_id()
    print(f"🔍 Debug tambah_data (GET) - user_id dari session: {user_id}")

    # Untuk halaman tambah-data, form harus selalu kosong saat GET request
    # Karena ini halaman untuk menambah data baru, bukan edit
    biodata = None

    # Tahun saat ini menurut waktu Indonesia (WITA - Sulawesi Tengah / Asia/Makassar)
    current_year = datetime.now(ZoneInfo("Asia/Makassar")).year

    # Ambil daftar kegiatan dari database (semua kegiatan yang dibuat admin)
    kegiatan_list = []
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor()
            # Ambil semua kegiatan dari kegiatan_master
            # Gunakan query yang sama dengan halaman admin untuk konsistensi
            cursor.execute("""
                SELECT id, nama_kegiatan, waktu_pelaksanaan, tempat_pelaksanaan
                FROM kegiatan_master
                WHERE TRIM(nama_kegiatan) != ''
                    AND (is_hidden IS NULL OR is_hidden = 0)
                ORDER BY nama_kegiatan ASC
            """)
            rows = cursor.fetchall()
            # Convert Row objects to dicts
            kegiatan_list = [row_to_dict(row) for row in rows]
            print(f"🔍 Debug tambah_data - Jumlah kegiatan ditemukan: {len(kegiatan_list)}")
            if kegiatan_list:
                print(f"🔍 Debug tambah_data - Kegiatan pertama: {kegiatan_list[0]}")
                for idx, kg in enumerate(kegiatan_list[:3]):  # Print 3 pertama
                    print(f"  [{idx+1}] {kg.get('nama_kegiatan', 'N/A')}")
            else:
                print("⚠️  Warning: Tidak ada kegiatan ditemukan di database!")
        except sqlite3.Error as e:
            print(f"❌ Error fetching kegiatan: {e}")
            import traceback
            traceback.print_exc()
            flash(f'Terjadi kesalahan saat mengambil daftar kegiatan: {str(e)}', 'error')
        finally:
            if connection is not None:
                cursor.close()
                connection.close()
    else:
        print("❌ Error: Tidak dapat membuat koneksi ke database")
        flash('Tidak dapat terhubung ke database!', 'error')

    # Handle POST request
    if request.method == 'POST':
        try:
            form_data = get_form_data()

            # Validasi dasar
            if not form_data['nik']:
                flash('NIK wajib diisi!', 'error')
                return render_template('user/tambah-data.html', biodata=biodata, kegiatan_list=kegiatan_list, current_year=current_year, username=get_username())

            # Validasi NIK harus tepat 16 digit
            if not form_data['nik'].isdigit() or len(form_data['nik']) != 16:
                flash('NIK harus tepat 16 digit angka!', 'error')
                return render_template('user/tambah-data.html', biodata=biodata, kegiatan_list=kegiatan_list, current_year=current_year, username=get_username())

            # Validasi field wajib
            if not validate_required_fields(form_data):
                # get_form_data() sekarang sudah menampilkan field yang kosong di console
                flash('Semua field wajib harus diisi! Periksa kembali semua field wajib.', 'error')
                return render_template('user/tambah-data.html', biodata=biodata, kegiatan_list=kegiatan_list, current_year=current_year, username=get_username())

            # Validasi kegiatan tidak disembunyikan
            if form_data.get('nama_kegiatan'):
                connection_check = get_db_connection()
                if connection_check:
                    try:
                        cursor_check = connection_check.cursor()
                        cursor_check.execute("""
                            SELECT is_hidden FROM kegiatan_master
                            WHERE TRIM(nama_kegiatan) = TRIM(?)
                            LIMIT 1
                        """, (form_data['nama_kegiatan'],))
                        kegiatan_check = cursor_check.fetchone()
                        if kegiatan_check and kegiatan_check[0] == 1:
                            flash('Kegiatan yang dipilih tidak tersedia!', 'error')
                            return render_template('user/tambah-data.html', biodata=biodata, kegiatan_list=kegiatan_list, current_year=current_year, username=get_username())
                    except sqlite3.Error as e:
                        print(f"Error checking kegiatan hidden status: {e}")
                    finally:
                        if connection_check:
                            cursor_check.close()
                            connection_check.close()

            # Handle file upload
            # Jika user sudah punya data, file tidak wajib (gunakan file yang sudah ada)
            buku_tabungan_file = request.files.get('buku_tabungan')
            buku_tabungan_path = None
            if buku_tabungan_file and buku_tabungan_file.filename:
                # User upload file baru
                buku_tabungan_path = save_uploaded_file(buku_tabungan_file, form_data['nik'])
                if not buku_tabungan_path:
                    flash('Gagal mengupload file buku tabungan! Pastikan file adalah gambar (PNG, JPG, JPEG, GIF).', 'error')
                    return render_template('user/tambah-data.html', biodata=biodata, kegiatan_list=kegiatan_list, current_year=current_year, username=get_username())
            elif biodata and biodata.get('buku_tabungan_path'):
                # User tidak upload file baru, gunakan file yang sudah ada
                buku_tabungan_path = biodata.get('buku_tabungan_path')
            elif form_data.get('existing_buku_tabungan_path'):
                # Auto-fill dari NIK lookup, gunakan path existing yang sudah ditarik
                buku_tabungan_path = form_data.get('existing_buku_tabungan_path')
            else:
                # User belum punya data dan tidak upload file
                flash('File buku tabungan wajib diupload!', 'error')
                return render_template('user/tambah-data.html', biodata=biodata, kegiatan_list=kegiatan_list, current_year=current_year, username=get_username())

            # Handle tanda tangan - simpan sebagai file
            tanda_tangan_path = None
            tanda_tangan_base64 = form_data.get('tanda_tangan')
            print(f"🔍 Debug tambah_data - tanda_tangan_base64 dari form: {bool(tanda_tangan_base64)}")
            print(f"🔍 Debug tambah_data - tanda_tangan_base64 length: {len(str(tanda_tangan_base64)) if tanda_tangan_base64 else 0}")
            print(f"🔍 Debug tambah_data - tanda_tangan_base64 preview: {str(tanda_tangan_base64)[:100] if tanda_tangan_base64 else 'None'}")

            if tanda_tangan_base64:
                # User membuat tanda tangan baru, simpan sebagai file
                print(f"🔍 Debug tambah_data - Memanggil save_tanda_tangan_file...")
                tanda_tangan_path = save_tanda_tangan_file(tanda_tangan_base64, form_data['nik'])
                print(f"🔍 Debug tambah_data - Hasil save_tanda_tangan_file: {tanda_tangan_path}")
                if not tanda_tangan_path:
                    flash('Gagal menyimpan tanda tangan!', 'error')
                    return render_template('user/tambah-data.html', biodata=biodata, kegiatan_list=kegiatan_list, current_year=current_year, username=get_username())
            elif biodata and biodata.get('tanda_tangan'):
                # User tidak membuat tanda tangan baru, gunakan tanda tangan yang sudah ada
                # Cek apakah sudah berupa path atau masih base64
                existing_ttd = biodata.get('tanda_tangan')
                print(f"🔍 Debug tambah_data - Menggunakan existing tanda_tangan: {str(existing_ttd)[:100] if existing_ttd else 'None'}")
                if existing_ttd and ('uploads/' in str(existing_ttd) or str(existing_ttd).startswith('static/')):
                    # Sudah berupa path file
                    tanda_tangan_path = normalize_buku_tabungan_path(existing_ttd)
                    print(f"🔍 Debug tambah_data - Existing adalah path: {tanda_tangan_path}")
                elif existing_ttd:
                    # Masih base64, simpan sebagai file
                    print(f"🔍 Debug tambah_data - Existing adalah base64, menyimpan sebagai file...")
                    tanda_tangan_path = save_tanda_tangan_file(existing_ttd, form_data['nik'])
                    if not tanda_tangan_path:
                        flash('Gagal menyimpan tanda tangan yang sudah ada!', 'error')
                        return render_template('user/tambah-data.html', biodata=biodata, kegiatan_list=kegiatan_list, current_year=current_year, username=get_username())
            else:
                # User belum punya data dan tidak membuat tanda tangan
                flash('Tanda tangan wajib diisi! Silakan buat tanda tangan di canvas.', 'error')
                return render_template('user/tambah-data.html', biodata=biodata, kegiatan_list=kegiatan_list, current_year=current_year, username=get_username())

            # Update form_data dengan tanda_tangan_path (bukan base64)
            form_data['tanda_tangan'] = tanda_tangan_path
            print(f"🔍 Debug tambah_data - tanda_tangan_path final: {tanda_tangan_path}")
            print(f"🔍 Debug tambah_data - form_data['tanda_tangan'] final: {form_data.get('tanda_tangan')}")

            # Cek apakah NIK yang diinput berbeda dari NIK yang ada di biodata_kegiatan untuk user_id saat ini
            # Jika berbeda, kita perlu menggunakan user_id berdasarkan NIK yang diinput, bukan dari session
            session_user_id = get_user_id()
            session_nik = None

            if session_user_id:
                # Ambil NIK dari biodata_kegiatan (bukan dari users, karena users tidak punya kolom nik)
                nik_check_conn = get_db_connection()
                if nik_check_conn:
                    try:
                        nik_cursor = nik_check_conn.cursor()
                        # Ambil NIK terbaru dari biodata_kegiatan untuk user_id ini
                        nik_cursor.execute("SELECT nik FROM biodata_kegiatan WHERE user_id = ? ORDER BY created_at DESC LIMIT 1", (session_user_id,))
                        session_user_data = nik_cursor.fetchone()
                        if session_user_data:
                            session_nik = session_user_data.get('nik')
                        nik_cursor.close()
                    except Exception as e:
                        print(f"⚠️ Error saat cek NIK session: {e}")
                    finally:
                        if nik_check_conn:
                            nik_check_conn.close()

            # Jika NIK yang diinput berbeda dari NIK yang ada di biodata, reset user_id agar proses baru
            # Atau jika tidak ada data di biodata_kegiatan untuk user_id ini (data sudah dihapus),
            # reset user_id agar dibuat/dicari berdasarkan NIK baru
            input_nik = form_data['nik']
            if session_nik and session_nik != input_nik:
                print(f"🔄 NIK berbeda terdeteksi! Existing NIK: {session_nik}, Input NIK: {input_nik}")
                print(f"🔄 Akan menggunakan user berdasarkan NIK baru, bukan session")
                user_id = None  # Reset user_id agar dibuat/dicari berdasarkan NIK baru
            elif session_user_id and not session_nik:
                # Tidak ada data di biodata_kegiatan untuk user_id ini (mungkin sudah dihapus)
                # Reset user_id agar dibuat/dicari berdasarkan NIK baru
                print(f"🔄 Tidak ada data di biodata_kegiatan untuk user_id {session_user_id}, akan menggunakan user berdasarkan NIK baru")
                user_id = None

            # Cek dulu apakah NIK ini sudah pernah digunakan untuk kegiatan yang sama (duplikat)
            temp_connection = get_db_connection()
            if temp_connection:
                try:
                    temp_cursor = temp_connection.cursor()
                    temp_cursor.execute("""
                        SELECT user_id FROM biodata_kegiatan
                        WHERE nik = ? AND TRIM(nama_kegiatan) = TRIM(?)
                        LIMIT 1
                    """, (form_data['nik'], form_data['nama_kegiatan']))
                    nik_kegiatan_exists = temp_cursor.fetchone()
                    if nik_kegiatan_exists:
                        temp_cursor.close()
                        temp_connection.close()
                        print(f"⚠️ NIK {form_data['nik']} sudah pernah digunakan untuk kegiatan '{form_data['nama_kegiatan']}'")
                        flash(f'NIK {form_data["nik"]} sudah pernah digunakan untuk kegiatan "{form_data["nama_kegiatan"]}". Silakan gunakan kegiatan lain atau hubungi admin jika ini kesalahan.', 'error')
                        return render_template('user/tambah-data.html', biodata=biodata, kegiatan_list=kegiatan_list, current_year=current_year, username=get_username())
                    temp_cursor.close()
                except Exception as e:
                    print(f"⚠️ Error saat cek NIK untuk kegiatan: {e}")
                finally:
                    if temp_connection is not None:
                        temp_connection.close()

            # Jika user_id tidak ada (tidak login atau NIK berbeda), buat atau dapatkan user berdasarkan NIK
            if not user_id:
                print(f"🔍 Debug tambah_data - Akan membuat/dapatkan user berdasarkan NIK: {form_data['nik']}")

                user_id = get_or_create_user_by_nik(
                    form_data['nik'],
                    form_data.get('nama_lengkap'),
                    form_data.get('alamat_email')
                )
                print(f"🔍 Debug tambah_data - Hasil get_or_create_user_by_nik: user_id={user_id}")
                if not user_id:
                    flash('Gagal membuat user! Silakan coba lagi.', 'error')
                    return render_template('user/tambah-data.html', biodata=biodata, kegiatan_list=kegiatan_list, current_year=current_year, username=session.get('username'))

                # Setelah user dibuat, ambil data user dan buat session agar user otomatis login
                connection = get_db_connection()
                if connection:
                    try:
                        cursor = connection.cursor()
                        cursor.execute("SELECT * FROM users WHERE id = ?", (user_id,))
                        new_user = cursor.fetchone()
                        if new_user:
                            # Set session untuk user baru
                            user_role = new_user.get('role', 'user')
                            if user_role:
                                user_role = user_role.lower().strip()
                            set_session_data(new_user, user_role)
                            print(f"✅ Session dibuat untuk user dengan ID: {user_id}")
                        cursor.close()
                    except Exception as e:
                        print(f"⚠️ Error saat membuat session untuk user: {e}")
                        import traceback
                        traceback.print_exc()
                    finally:
                        if connection is not None:
                            connection.close()

            # Cek apakah ini mode update (tombol Edit Data diklik)
            action = request.form.get('action', 'save')
            is_update_mode = (action == 'update' and form_data.get('original_nama_kegiatan'))

            if is_update_mode:
                # Mode UPDATE: gunakan save_biodata_data dengan old_nama_kegiatan
                print(f"✅ Mode UPDATE: Mengupdate data untuk kegiatan '{form_data.get('original_nama_kegiatan')}'")
                # Set old_nama_kegiatan untuk identifikasi data yang akan diupdate
                form_data['old_nama_kegiatan'] = form_data.get('original_nama_kegiatan')
                success, message = save_biodata_data(form_data, user_id, buku_tabungan_path)
            else:
                # Mode INSERT: simpan data baru
                print(f"✅ Mode INSERT: Menyimpan data baru untuk kegiatan '{form_data.get('nama_kegiatan')}'")
                print(f"🔍 Debug tambah_data (INSERT) - user_id yang akan digunakan: {user_id}, NIK: {form_data.get('nik')}")
                success, message = insert_biodata_data(form_data, user_id, buku_tabungan_path)

            if success:
                flash(f'Biodata kegiatan untuk "{form_data["nama_lengkap"]}" {message}', 'success')
                # Redirect untuk refresh halaman setelah berhasil
                return redirect(url_for('tambah_data'))
            else:
                flash(message, 'error')
                # Jika error, tetap tampilkan form dengan data yang sudah diisi (jika ada)
                return render_template('user/tambah-data.html', biodata=biodata, kegiatan_list=kegiatan_list, current_year=current_year, username=get_username())

        except Exception as e:
            print(f"Unexpected error in tambah_data: {e}")
            import traceback
            traceback.print_exc()
            flash(f'Terjadi kesalahan tidak terduga: {str(e)}', 'error')
            return render_template('user/tambah-data.html', biodata=biodata, kegiatan_list=kegiatan_list, current_year=current_year, username=get_username())

    # Handle GET request - tampilkan form dengan data existing jika ada
    return render_template('user/tambah-data.html', biodata=biodata, kegiatan_list=kegiatan_list, current_year=current_year, username=get_username())


@app.route('/user/lihat-data', methods=['GET'])
@login_required
def lihat_data():
    """Halaman untuk melihat detail data biodata kegiatan - redirect ke daftar kegiatan"""
    # Redirect ke daftar kegiatan karena halaman lihat-data tidak digunakan
    return redirect(url_for('daftar_kegiatan'))

@app.route('/user/daftar-kegiatan', methods=['GET', 'POST'])
@login_required
def daftar_kegiatan():
    """Halaman daftar kegiatan yang diikuti user"""

    # Jika admin mencoba mengakses, arahkan ke halaman admin
    if is_admin():
        return redirect(url_for('admin_kegiatan'))

    user_id = get_user_id()
    if not user_id:
        flash('Session tidak valid! Silakan login kembali.', 'error')
        return redirect(url_for('login'))

    # Ambil semua kegiatan yang diikuti user (yang memiliki biodata)
    kegiatan_user_list = []
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor()
            # Ambil semua biodata user dengan informasi kegiatannya
            cursor.execute("""
                SELECT DISTINCT
                    bk.nama_kegiatan,
                    COALESCE(km.waktu_pelaksanaan, bk.waktu_pelaksanaan, '') as waktu_pelaksanaan,
                    COALESCE(km.tempat_pelaksanaan, bk.tempat_pelaksanaan, '') as tempat_pelaksanaan,
                    bk.nik,
                    bk.nama_lengkap
                FROM biodata_kegiatan bk
                LEFT JOIN kegiatan_master km ON TRIM(km.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                WHERE bk.user_id = ?
                    AND TRIM(bk.nama_kegiatan) != ''
                    AND bk.nama_kegiatan IS NOT NULL
                ORDER BY bk.created_at DESC
            """, (user_id,))
            kegiatan_user_list = cursor.fetchall()
        except sqlite3.Error as e:
            print(f"Error fetching kegiatan user: {e}")
        finally:
            if connection is not None:
                cursor.close()
                connection.close()

    return render_template('user/daftar-kegiatan.html',
                         username=get_username(),
                         kegiatan_user_list=kegiatan_user_list)

@app.route('/admin', methods=['GET', 'POST'])
@app.route('/admin/dashboard', methods=['GET'])
@admin_required
def admin_dashboard():
    """Halaman dashboard admin/operator dengan statistik dan overview. Operator hanya melihat data kegiatan yang dipegang."""

    user_role = session.get('user_role', 'admin')
    user_id = get_user_id()

    # Ambil statistik dari database
    stats = {
        'total_biodata': 0,
        'total_kegiatan': 0,
        'total_users': 0,
        'total_kabupaten': 0
    }

    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor()

            if user_role == 'operator' and user_id:
                # Operator: hanya data kegiatan yang dipegang
                # Total biodata (biodata dari kegiatan operator)
                cursor.execute("""
                    SELECT COUNT(*) as count
                    FROM biodata_kegiatan bk
                    INNER JOIN kegiatan_master k ON TRIM(k.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                    INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                    WHERE ok.user_id = ?
                """, (user_id,))
                result = cursor.fetchone()
                stats['total_biodata'] = result['count'] if result else 0

                # Total kegiatan (kegiatan yang dipegang operator)
                cursor.execute("""
                    SELECT COUNT(*) as count
                    FROM kegiatan_master k
                    INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                    WHERE ok.user_id = ? AND TRIM(k.nama_kegiatan) != ''
                """, (user_id,))
                result = cursor.fetchone()
                stats['total_kegiatan'] = result['count'] if result else 0

                # Total operator (tetap hitung semua operator)
                cursor.execute("SELECT COUNT(*) as count FROM users WHERE role = 'operator'")
                result = cursor.fetchone()
                stats['total_users'] = result['count'] if result else 0

                # Total kabupaten dari biodata kegiatan operator
                cursor.execute("""
                    SELECT COUNT(DISTINCT bk.kabupaten_kota) as count
                    FROM biodata_kegiatan bk
                    INNER JOIN kegiatan_master k ON TRIM(k.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                    INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                    WHERE ok.user_id = ? AND TRIM(bk.kabupaten_kota) != '' AND bk.kabupaten_kota IS NOT NULL
                """, (user_id,))
                result = cursor.fetchone()
                stats['total_kabupaten'] = result['count'] if result else 0
            else:
                # Admin: semua data
                cursor.execute("SELECT COUNT(*) as count FROM biodata_kegiatan")
                result = cursor.fetchone()
                stats['total_biodata'] = result['count'] if result else 0

                cursor.execute("SELECT COUNT(*) as count FROM kegiatan_master WHERE TRIM(nama_kegiatan) != ''")
                result = cursor.fetchone()
                stats['total_kegiatan'] = result['count'] if result else 0

                cursor.execute("SELECT COUNT(*) as count FROM users WHERE role = 'operator'")
                result = cursor.fetchone()
                stats['total_users'] = result['count'] if result else 0

                cursor.execute("""
                    SELECT COUNT(DISTINCT kabupaten_kota) as count
                    FROM biodata_kegiatan
                    WHERE TRIM(kabupaten_kota) != '' AND kabupaten_kota IS NOT NULL
                """)
                result = cursor.fetchone()
                stats['total_kabupaten'] = result['count'] if result else 0

        except sqlite3.Error as e:
            print(f"Error fetching stats: {e}")
        finally:
            if connection is not None:
                cursor.close()
                connection.close()

    # Ambil data kabupaten summary untuk popup dan grafik
    kabupaten_summary = []
    connection2 = get_db_connection()
    if connection2:
        try:
            all_kabupaten_list = [
                'BANGGAI', 'BANGGAI KEPULAUAN', 'BANGGAI LAUT', 'BUOL', 'DONGGALA',
                'MOROWALI', 'MOROWALI UTARA', 'PALU', 'PARIGI MOUTONG', 'POSO',
                'SIGI', 'TOJO UNA-UNA', 'TOLI-TOLI'
            ]
            all_kabupaten_upper = {k.upper().strip() for k in all_kabupaten_list}

            cursor2 = connection2.cursor()
            if user_role == 'operator' and user_id:
                cursor2.execute("""
                    SELECT bk.kabupaten_kota, COUNT(*) as jumlah_peserta
                    FROM biodata_kegiatan bk
                    INNER JOIN kegiatan_master k ON TRIM(k.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                    INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                    WHERE ok.user_id = ? AND TRIM(bk.kabupaten_kota) != '' AND bk.kabupaten_kota IS NOT NULL
                    GROUP BY bk.kabupaten_kota
                """, (user_id,))
            else:
                cursor2.execute("""
                    SELECT kabupaten_kota, COUNT(*) as jumlah_peserta
                    FROM biodata_kegiatan
                    WHERE TRIM(kabupaten_kota) != '' AND kabupaten_kota IS NOT NULL
                    GROUP BY kabupaten_kota
                """)
            kabupaten_counts = {row['kabupaten_kota']: row['jumlah_peserta'] for row in cursor2.fetchall()}

            lainnya_count = 0
            for kab, count in kabupaten_counts.items():
                if (kab or '').strip().upper() not in all_kabupaten_upper:
                    lainnya_count += count

            for kabupaten in all_kabupaten_list:
                kabupaten_summary.append({
                    'nama': kabupaten,
                    'jumlah_peserta': kabupaten_counts.get(kabupaten, 0)
                })
            kabupaten_summary.append({'nama': 'LAINNYA', 'jumlah_peserta': lainnya_count})
            # Pastikan "LAINNYA" selalu paling terakhir
            kabupaten_summary.sort(key=lambda x: ((x.get('nama') or '').strip().upper() == 'LAINNYA', x.get('nama') or ''))
        except sqlite3.Error as e:
            print(f"Error fetching kabupaten summary: {e}")
        finally:
            if connection2:
                cursor2.close()
                connection2.close()

    # Tentukan role untuk template
    dashboard_title = 'Dashboard Operator' if user_role == 'operator' else 'Dashboard Admin'

    return render_template(
        'admin/admin-dashboard.html',
        username=get_username(),
        stats=stats,
        user_role=user_role,
        dashboard_title=dashboard_title,
        kabupaten_summary=kabupaten_summary
    )

@app.route('/admin/export-rekap-kabupaten-pdf/<path:kabupaten>')
@admin_required
def export_rekap_kabupaten_pdf(kabupaten):
    """Export rekap per kabupaten ke PDF - mengikuti style rekap tahunan"""
    from io import BytesIO
    from urllib.parse import unquote
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors

    # Decode URL encoding
    kabupaten = unquote(kabupaten)
    
    # Ambil user role dan user_id untuk filter operator
    user_role = get_user_role()
    user_id = get_user_id()

    connection = get_db_connection()
    if not connection:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_rekap_filter'))

    try:
        cursor = connection.cursor()
        # Ambil semua biodata untuk kabupaten ini
        # Jika operator, batasi hanya pada kegiatan yang ia pegang
        if user_role == 'operator' and user_id:
            cursor.execute("""
                SELECT
                    bk.*,
                    COALESCE(km.nama_kegiatan, bk.nama_kegiatan, '') AS nama_kegiatan,
                    COALESCE(km.tempat_pelaksanaan, bk.tempat_pelaksanaan, '') AS tempat_pelaksanaan,
                    COALESCE(km.waktu_pelaksanaan, bk.waktu_pelaksanaan, '') AS waktu_pelaksanaan
                FROM biodata_kegiatan bk
                LEFT JOIN kegiatan_master km ON TRIM(km.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                WHERE TRIM(bk.kabupaten_kota) = TRIM(?)
                  AND EXISTS (
                      SELECT 1
                      FROM kegiatan_master k
                      INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                      WHERE ok.user_id = ?
                        AND TRIM(k.nama_kegiatan) = TRIM(COALESCE(km.nama_kegiatan, bk.nama_kegiatan, ''))
                  )
                ORDER BY bk.nama_lengkap ASC
            """, (kabupaten, user_id))
        else:
            cursor.execute("""
                SELECT
                    bk.*,
                    COALESCE(km.nama_kegiatan, bk.nama_kegiatan, '') AS nama_kegiatan,
                    COALESCE(km.tempat_pelaksanaan, bk.tempat_pelaksanaan, '') AS tempat_pelaksanaan,
                    COALESCE(km.waktu_pelaksanaan, bk.waktu_pelaksanaan, '') AS waktu_pelaksanaan
                FROM biodata_kegiatan bk
                LEFT JOIN kegiatan_master km ON TRIM(km.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                WHERE TRIM(bk.kabupaten_kota) = TRIM(?)
                ORDER BY bk.nama_lengkap ASC
            """, (kabupaten,))
        rows = cursor.fetchall()
        all_biodata = [row_to_dict(row) for row in rows]

        if not all_biodata:
            flash(f'Tidak ada data untuk kabupaten {kabupaten}!', 'error')
            return redirect(url_for('admin_rekap_filter'))

        # Normalisasi path buku tabungan
        for biodata in all_biodata:
            if biodata.get('buku_tabungan_path'):
                biodata['buku_tabungan_path'] = normalize_buku_tabungan_path(biodata['buku_tabungan_path'])

    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan saat mengambil data: {str(e)}', 'error')
        return redirect(url_for('admin_rekap_filter'))
    finally:
        if connection:
            cursor.close()
            connection.close()

    # Load logo untuk header
    logo_path = os.path.join(BASE_DIR, 'static', 'Logo_BGTK.png')
    logo_img = None
    logo_width = 0
    logo_height = 0
    if os.path.exists(logo_path):
        try:
            logo_pil = Image.open(logo_path)
            max_logo_height = 0.6 * inch
            logo_ratio = logo_pil.width / logo_pil.height
            logo_width = max_logo_height * logo_ratio
            logo_height = max_logo_height

            logo_buffer = io.BytesIO()
            logo_pil.save(logo_buffer, format='PNG')
            logo_buffer.seek(0)
            logo_img = RLImage(logo_buffer, width=logo_width, height=logo_height)
        except Exception as e:
            print(f"Error loading logo: {e}")
            logo_img = None

    # Load logo Pendidikan Bermutu dengan background transparan
    pendidikan_bermutu_path = os.path.join(BASE_DIR, 'static', 'Pendidikan Bermutu untuk Semua.png')
    pendidikan_bermutu_processed_path = None
    pendidikan_bermutu_width = 0
    pendidikan_bermutu_height = 0
    pendidikan_bermutu_temp_file = None
    if os.path.exists(pendidikan_bermutu_path):
        try:
            pendidikan_bermutu_pil = Image.open(pendidikan_bermutu_path)
            if pendidikan_bermutu_pil.mode != 'RGBA':
                pendidikan_bermutu_pil = pendidikan_bermutu_pil.convert('RGBA')

            pixels = pendidikan_bermutu_pil.load()
            width, height = pendidikan_bermutu_pil.size
            black_threshold = 30

            for x in range(width):
                for y in range(height):
                    r, g, b, a = pixels[x, y]
                    if r < black_threshold and g < black_threshold and b < black_threshold:
                        pixels[x, y] = (r, g, b, 0)

            max_logo_height = 0.5 * inch
            pendidikan_bermutu_ratio = pendidikan_bermutu_pil.width / pendidikan_bermutu_pil.height
            pendidikan_bermutu_width = max_logo_height * pendidikan_bermutu_ratio
            pendidikan_bermutu_height = max_logo_height

            pendidikan_bermutu_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            pendidikan_bermutu_pil.save(pendidikan_bermutu_temp_file.name, format='PNG', optimize=False)
            pendidikan_bermutu_processed_path = pendidikan_bermutu_temp_file.name
        except Exception as e:
            print(f"Error loading logo Pendidikan Bermutu: {e}")

    # Load logo Ramah dengan background transparan
    ramah_path = os.path.join(BASE_DIR, 'static', 'Ramah.png')
    ramah_processed_path = None
    ramah_width = 0
    ramah_height = 0
    ramah_temp_file = None
    if os.path.exists(ramah_path):
        try:
            ramah_pil = Image.open(ramah_path)
            if ramah_pil.mode != 'RGBA':
                ramah_pil = ramah_pil.convert('RGBA')

            pixels = ramah_pil.load()
            width, height = ramah_pil.size
            black_threshold = 30

            for x in range(width):
                for y in range(height):
                    r, g, b, a = pixels[x, y]
                    if r < black_threshold and g < black_threshold and b < black_threshold:
                        pixels[x, y] = (r, g, b, 0)

            max_logo_height = 0.5 * inch
            ramah_ratio = ramah_pil.width / ramah_pil.height
            ramah_width = max_logo_height * ramah_ratio
            ramah_height = max_logo_height

            ramah_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            ramah_pil.save(ramah_temp_file.name, format='PNG', optimize=False)
            ramah_processed_path = ramah_temp_file.name
        except Exception as e:
            print(f"Error loading logo Ramah: {e}")

    # Fungsi untuk header dengan logo dan footer
    def add_header_footer(canvas, doc):
        canvas.saveState()

        max_logo_h = logo_height if logo_height > 0 else 0

        # Header - Logo BGTK di kiri
        if logo_img and logo_height > 0:
            try:
                logo_x = 25
                logo_y = F4_SIZE[1] - 25 - max_logo_h
                canvas.drawImage(logo_path, logo_x, logo_y, width=logo_width, height=logo_height, preserveAspectRatio=True)
            except Exception as e:
                print(f"Error drawing logo: {e}")

        # Garis header
        canvas.setStrokeColor(colors.HexColor('#067ac1'))
        canvas.setLineWidth(1.5)
        max_header_height = logo_height if logo_height > 0 else 0
        header_line_y = F4_SIZE[1] - 25 - max_header_height - 5
        canvas.line(25, header_line_y, F4_SIZE[0] - 25, header_line_y)

        # Footer - Logo Pendidikan Bermutu dan Ramah di kanan bawah
        max_footer_logo_height = max(pendidikan_bermutu_height if pendidikan_bermutu_height > 0 else 0,
                                    ramah_height if ramah_height > 0 else 0)
        footer_logo_y = 45

        if max_footer_logo_height > 0:
            total_footer_logo_width = (pendidikan_bermutu_width if pendidikan_bermutu_height > 0 else 0) + \
                                     (ramah_width if ramah_height > 0 else 0) + 10
            footer_logo_start_x = F4_SIZE[0] - 25 - total_footer_logo_width

            if pendidikan_bermutu_processed_path and os.path.exists(pendidikan_bermutu_processed_path) and pendidikan_bermutu_height > 0:
                try:
                    pendidikan_bermutu_footer_x = footer_logo_start_x
                    pendidikan_bermutu_footer_y = footer_logo_y
                    canvas.drawImage(pendidikan_bermutu_processed_path, pendidikan_bermutu_footer_x, pendidikan_bermutu_footer_y,
                                   width=pendidikan_bermutu_width, height=pendidikan_bermutu_height, preserveAspectRatio=True, mask='auto')
                except Exception as e:
                    print(f"Error drawing logo Pendidikan Bermutu di footer: {e}")

            if ramah_processed_path and os.path.exists(ramah_processed_path) and ramah_height > 0:
                try:
                    ramah_footer_x = footer_logo_start_x + (pendidikan_bermutu_width if pendidikan_bermutu_height > 0 else 0) + 10
                    ramah_footer_y = footer_logo_y
                    canvas.drawImage(ramah_processed_path, ramah_footer_x, ramah_footer_y,
                                   width=ramah_width, height=ramah_height, preserveAspectRatio=True, mask='auto')
                except Exception as e:
                    print(f"Error drawing logo Ramah di footer: {e}")

        # Garis footer
        canvas.setStrokeColor(colors.HexColor('#067ac1'))
        canvas.setLineWidth(1)
        footer_line_y = footer_logo_y - 10
        canvas.line(25, footer_line_y, F4_SIZE[0] - 25, footer_line_y)

        # Footer - Tanggal dan waktu export WITA
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.black)
        wita_time = datetime.utcnow() + timedelta(hours=8)
        footer_text = f"Dicetak pada: {wita_time.strftime('%d/%m/%Y %H:%M')} WITA"
        footer_width = canvas.stringWidth(footer_text, 'Helvetica', 8)
        footer_x = 25
        footer_y = footer_line_y - 15
        canvas.drawString(footer_x, footer_y, footer_text)

        canvas.restoreState()

    # Gunakan kertas F4 dan beri ruang header
    buffer = io.BytesIO()
    max_header_logo_height = logo_height if logo_height > 0 else 0
    top_margin_with_logo = 25 + max_header_logo_height + 15
    doc = SimpleDocTemplate(
        buffer,
        pagesize=F4_SIZE,
        rightMargin=25,
        leftMargin=25,
        topMargin=top_margin_with_logo,
        bottomMargin=40
    )

    elements = []
    styles = getSampleStyleSheet()

    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.black,
        spaceAfter=12,
        spaceBefore=0,
        alignment=1,  # Center
        fontName='Helvetica-Bold',
        leading=24
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.black,
        alignment=0  # Left
    )
    value_style = ParagraphStyle(
        'ValueStyle',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.black,
        alignment=0,  # Left
        leftIndent=0,
        rightIndent=0
    )

    # Field mapping untuk label yang lebih readable dan profesional
    field_labels = {
        'nik': 'NIK',
        'nama_lengkap': 'Nama Lengkap',
        'nip_nippk': 'NIP/NIPPK',
        'tempat_lahir': 'Tempat Lahir',
        'tanggal_lahir': 'Tanggal Lahir',
        'jenis_kelamin': 'Jenis Kelamin',
        'agama': 'Agama',
        'pendidikan_terakhir': 'Pendidikan Terakhir',
        'jurusan': 'Jurusan',
        'status_asn': 'Status Kepegawaian',
        'pangkat_golongan': 'Pangkat / Golongan',
        'jabatan': 'Jabatan',
        'instansi': 'Nama Instansi',
        'alamat_instansi': 'Alamat Instansi',
        'alamat_domisili': 'Alamat Domisili',
        'kabupaten_kota': 'Kabupaten/Kota',
        'kabko_lainnya': 'Kabupaten/Kota Lainnya',
        'peran': 'Peran dalam Kegiatan',
        'no_hp': 'Nomor HP',
        'alamat_email': 'Alamat Email',
        'npwp': 'NPWP',
        'nama_bank': 'Nama Bank',
        'nama_bank_lainnya': 'Nama Bank Lainnya',
        'no_rekening': 'Nomor Rekening',
        'nama_pemilik_rekening': 'Nama Pemilik Rekening'
    }

    # Urutan field yang diinginkan
    field_order = [
        'nik', 'nama_lengkap', 'nip_nippk', 'tempat_lahir', 'tanggal_lahir',
        'jenis_kelamin', 'agama', 'pendidikan_terakhir', 'jurusan',
        'status_asn', 'pangkat_golongan', 'jabatan', 'instansi',
        'alamat_instansi', 'alamat_domisili', 'kabupaten_kota', 'kabko_lainnya',
        'peran',
        'no_hp', 'alamat_email', 'npwp',
        'nama_bank', 'nama_bank_lainnya', 'no_rekening', 'nama_pemilik_rekening'
    ]

    # Exclude fields yang tidak perlu ditampilkan
    exclude_fields = ['id', 'user_id', 'buku_tabungan_path', 'tanda_tangan', 'created_at', 'updated_at', 'nama_kegiatan', 'waktu_pelaksanaan', 'tempat_pelaksanaan']

    # List untuk menyimpan temporary files tanda tangan untuk cleanup
    tanda_tangan_temp_files = []


    # Loop untuk setiap user - buat 1 halaman per user
    for user_idx, biodata in enumerate(all_biodata):
        # Tambahkan page break kecuali untuk user pertama
        if user_idx > 0:
            elements.append(PageBreak())

        # Title
        title = Paragraph("BIODATA KEGIATAN", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.15*inch))

        # Ambil data kegiatan untuk header tabel
        biodata_nama_kegiatan = biodata.get('nama_kegiatan', '-')
        waktu_pelaksanaan = biodata.get('waktu_pelaksanaan', '-')
        tempat_pelaksanaan = biodata.get('tempat_pelaksanaan', '-')

        # Buat header informasi kegiatan di atas tabel dengan titik dua yang sejajar
        if biodata_nama_kegiatan or waktu_pelaksanaan or tempat_pelaksanaan:
            info_style = ParagraphStyle(
                'InfoStyle',
                parent=styles['Normal'],
                fontSize=10,
                leading=14,
                textColor=colors.black,
                alignment=0,  # Left align
                fontName='Helvetica-Bold'
            )

            info_table_data = []
            if biodata_nama_kegiatan and str(biodata_nama_kegiatan).strip() and str(biodata_nama_kegiatan).strip() != '-':
                info_table_data.append([
                    Paragraph("<b>Nama Kegiatan</b>", info_style),
                    Paragraph(":", info_style),
                    Paragraph(str(biodata_nama_kegiatan), info_style)
                ])
            if waktu_pelaksanaan and str(waktu_pelaksanaan).strip() and str(waktu_pelaksanaan).strip() != '-':
                info_table_data.append([
                    Paragraph("<b>Waktu Pelaksanaan</b>", info_style),
                    Paragraph(":", info_style),
                    Paragraph(str(waktu_pelaksanaan), info_style)
                ])
            if tempat_pelaksanaan and str(tempat_pelaksanaan).strip() and str(tempat_pelaksanaan).strip() != '-':
                info_table_data.append([
                    Paragraph("<b>Tempat Pelaksanaan</b>", info_style),
                    Paragraph(":", info_style),
                    Paragraph(str(tempat_pelaksanaan), info_style)
                ])

            if info_table_data:
                available_width = (8.27 * inch) - (25 * 2)
                info_table = Table(info_table_data, colWidths=[2.2*inch, 0.15*inch, available_width - 2.2*inch - 0.15*inch])
                info_table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('ALIGN', (2, 0), (2, -1), 'LEFT'),
                    ('LEFTPADDING', (0, 0), (0, -1), 0),
                    ('LEFTPADDING', (1, 0), (1, -1), 0),
                    ('LEFTPADDING', (2, 0), (2, -1), 0),
                    ('RIGHTPADDING', (0, 0), (0, -1), 4),
                    ('RIGHTPADDING', (1, 0), (1, -1), 0),
                    ('RIGHTPADDING', (2, 0), (2, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ]))
                elements.append(info_table)
                elements.append(Spacer(1, 0.1*inch))

        # Semua data dalam 1 tabel
        all_data = []

        # Collect all data in order
        for key in field_order:
            if key in biodata and key not in exclude_fields:
                label = field_labels.get(key, key.replace('_', ' ').title())
                value = biodata[key]
                display_value = str(value) if value and str(value).strip() else '-'
                all_data.append([label, display_value])

        # Add any remaining fields not in field_order
        for key, value in biodata.items():
            if key not in exclude_fields and key not in field_order:
                label = field_labels.get(key, key.replace('_', ' ').title())
                display_value = str(value) if value and str(value).strip() else '-'
                all_data.append([label, display_value])

        # Process tanda tangan menggunakan helper function
        tanda_tangan_raw = biodata.get('tanda_tangan')
        tanda_tangan_img, error_msg = process_tanda_tangan_for_pdf(tanda_tangan_raw, tanda_tangan_temp_files)

        # Buat 1 tabel untuk semua data
        if all_data:
            table_data = []
            for row in all_data:
                field_label = Paragraph(f"<b>{row[0]}</b>", value_style)
                value_text = str(row[1]) if row[1] else '-'
                value_text = value_text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                if len(value_text) > 50:
                    words = value_text.split()
                    lines = []
                    current_line = []
                    current_length = 0
                    for word in words:
                        if current_length + len(word) + 1 > 50:
                            if current_line:
                                lines.append(' '.join(current_line))
                            current_line = [word]
                            current_length = len(word)
                        else:
                            current_line.append(word)
                            current_length += len(word) + 1
                    if current_line:
                        lines.append(' '.join(current_line))
                    value_text = '<br/>'.join(lines)
                value_para = Paragraph(value_text, value_style)
                table_data.append([field_label, value_para])

            # Tambahkan tanda tangan ke tabel jika ada
            if tanda_tangan_img:
                field_label = Paragraph("<b>Tanda Tangan</b>", value_style)
                table_data.append([field_label, tanda_tangan_img])
            elif biodata.get('tanda_tangan'):
                field_label = Paragraph("<b>Tanda Tangan</b>", value_style)
                error_msg_para = Paragraph("<i>Tanda tangan tidak tersedia</i>", value_style)
                table_data.append([field_label, error_msg_para])

            available_width = (8.27 * inch) - (25 * 2)
            table = Table(table_data, colWidths=[2.2*inch, available_width - 2.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (0, -1), 4),
                ('LEFTPADDING', (1, 0), (1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
                ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#067ac1')),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.08*inch))

    # Build PDF
    try:
        doc.build(elements, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
        buffer.seek(0)

        # Refresh session sebelum return response untuk mencegah logout
        session.permanent = True
        session.modified = True
    finally:
        # Cleanup temporary files
        try:
            if pendidikan_bermutu_temp_file and os.path.exists(pendidikan_bermutu_processed_path):
                os.unlink(pendidikan_bermutu_processed_path)
        except Exception as e:
            print(f"Error cleaning up Pendidikan Bermutu temp file: {e}")

        try:
            if ramah_temp_file and os.path.exists(ramah_processed_path):
                os.unlink(ramah_processed_path)
        except Exception as e:
            print(f"Error cleaning up Ramah temp file: {e}")

        # Cleanup temporary files tanda tangan
        for temp_file_path in tanda_tangan_temp_files:
            try:
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)
            except Exception as e:
                print(f"Error cleaning up tanda tangan temp file {temp_file_path}: {e}")

    # Generate filename
    filename = f"Rekap_Kabupaten_{kabupaten.replace(' ', '_')}.pdf"

    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
    )

@app.route('/admin/export-rekap-kabupaten-excel/<path:kabupaten>')
@admin_required
def export_rekap_kabupaten_excel(kabupaten):
    """Export rekap per kabupaten ke Excel - mengikuti style rekap tahunan"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from urllib.parse import unquote

    # Decode URL encoding
    kabupaten = unquote(kabupaten)
    
    # Ambil user role dan user_id untuk filter operator
    user_role = get_user_role()
    user_id = get_user_id()

    connection = get_db_connection()
    if not connection:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_rekap_filter'))

    try:
        cursor = connection.cursor()
        # Ambil semua biodata untuk kabupaten ini
        # Jika operator, batasi hanya pada kegiatan yang ia pegang
        if user_role == 'operator' and user_id:
            cursor.execute("""
                SELECT
                    bk.*,
                    COALESCE(km.nama_kegiatan, bk.nama_kegiatan, '') AS nama_kegiatan,
                    COALESCE(km.tempat_pelaksanaan, bk.tempat_pelaksanaan, '') AS tempat_pelaksanaan,
                    COALESCE(km.waktu_pelaksanaan, bk.waktu_pelaksanaan, '') AS waktu_pelaksanaan
                FROM biodata_kegiatan bk
                LEFT JOIN kegiatan_master km ON TRIM(km.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                WHERE TRIM(bk.kabupaten_kota) = TRIM(?)
                  AND EXISTS (
                      SELECT 1
                      FROM kegiatan_master k
                      INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                      WHERE ok.user_id = ?
                        AND TRIM(k.nama_kegiatan) = TRIM(COALESCE(km.nama_kegiatan, bk.nama_kegiatan, ''))
                  )
                ORDER BY bk.nama_lengkap ASC
            """, (kabupaten, user_id))
        else:
            cursor.execute("""
                SELECT
                    bk.*,
                    COALESCE(km.nama_kegiatan, bk.nama_kegiatan, '') AS nama_kegiatan,
                    COALESCE(km.tempat_pelaksanaan, bk.tempat_pelaksanaan, '') AS tempat_pelaksanaan,
                    COALESCE(km.waktu_pelaksanaan, bk.waktu_pelaksanaan, '') AS waktu_pelaksanaan
                FROM biodata_kegiatan bk
                LEFT JOIN kegiatan_master km ON TRIM(km.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                WHERE TRIM(bk.kabupaten_kota) = TRIM(?)
                ORDER BY bk.nama_lengkap ASC
            """, (kabupaten,))
        rows = cursor.fetchall()
        all_biodata = [row_to_dict(row) for row in rows]

        if not all_biodata:
            flash(f'Tidak ada data untuk kabupaten {kabupaten}!', 'error')
            return redirect(url_for('admin_rekap_filter'))

        # Normalisasi path buku tabungan
        for biodata in all_biodata:
            if biodata.get('buku_tabungan_path'):
                biodata['buku_tabungan_path'] = normalize_buku_tabungan_path(biodata['buku_tabungan_path'])

    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan saat mengambil data: {str(e)}', 'error')
        return redirect(url_for('admin_rekap_filter'))
    finally:
        if connection:
            cursor.close()
            connection.close()

    # Field mapping untuk label yang lebih readable dan profesional
    field_labels = {
        'nik': 'NIK',
        'nama_lengkap': 'Nama Lengkap',
        'nip_nippk': 'NIP/NIPPK',
        'tempat_lahir': 'Tempat Lahir',
        'tanggal_lahir': 'Tanggal Lahir',
        'jenis_kelamin': 'Jenis Kelamin',
        'agama': 'Agama',
        'pendidikan_terakhir': 'Pendidikan Terakhir',
        'jurusan': 'Jurusan',
        'status_asn': 'Status Kepegawaian',
        'pangkat_golongan': 'Pangkat / Golongan',
        'jabatan': 'Jabatan',
        'instansi': 'Nama Instansi',
        'alamat_instansi': 'Alamat Instansi',
        'alamat_domisili': 'Alamat Domisili',
        'kabupaten_kota': 'Kabupaten/Kota',
        'kabko_lainnya': 'Kabupaten/Kota Lainnya',
        'peran': 'Peran dalam Kegiatan',
        'no_hp': 'Nomor HP',
        'alamat_email': 'Alamat Email',
        'npwp': 'NPWP',
        'nama_bank': 'Nama Bank',
        'nama_bank_lainnya': 'Nama Bank Lainnya',
        'no_rekening': 'Nomor Rekening',
        'nama_pemilik_rekening': 'Nama Pemilik Rekening'
    }

    # Urutan field yang diinginkan
    field_order = [
        'nik', 'nama_lengkap', 'nip_nippk', 'tempat_lahir', 'tanggal_lahir',
        'jenis_kelamin', 'agama', 'pendidikan_terakhir', 'jurusan',
        'status_asn', 'pangkat_golongan', 'jabatan', 'instansi',
        'alamat_instansi', 'alamat_domisili', 'kabupaten_kota', 'kabko_lainnya',
        'peran',
        'no_hp', 'alamat_email', 'npwp',
        'nama_bank', 'nama_bank_lainnya', 'no_rekening', 'nama_pemilik_rekening'
    ]

    # Exclude fields yang tidak perlu ditampilkan
    exclude_fields = ['id', 'user_id', 'buku_tabungan_path', 'tanda_tangan', 'created_at', 'updated_at', 'waktu_pelaksanaan', 'tempat_pelaksanaan']

    # Buat workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Rekap {kabupaten}"

    # Styles
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_font = Font(bold=True, color="000000", size=11)
    title_font = Font(bold=True, size=12, color="000000")
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Buat list semua field yang akan ditampilkan (dalam urutan yang diinginkan)
    all_fields = []
    # Tambahkan nama_kegiatan di awal
    all_fields.append({'key': 'nama_kegiatan', 'label': 'Nama Kegiatan'})
    for key in field_order:
        if key not in exclude_fields:
            label = field_labels.get(key, key.replace('_', ' ').title())
            all_fields.append({'key': key, 'label': label})

    # Tambahkan field tanda tangan
    all_fields.append({'key': 'tanda_tangan', 'label': 'Tanda Tangan'})

    # Header kegiatan
    num_cols = len(all_fields)
    last_col_letter = get_column_letter(num_cols)

    # Title
    title_text = f"REKAP DATA PESERTA - {kabupaten.upper()}"

    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = title_text
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    # Tambahkan informasi tanggal export
    export_date = datetime.now().strftime("%d %B %Y, %H:%M:%S")
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = f"Tanggal Export: {export_date}"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = left_align
    ws.row_dimensions[2].height = 18

    # Buat header kolom (semua field sebagai kolom)
    for col_idx, field_info in enumerate(all_fields, 1):
        cell = ws.cell(row=3, column=col_idx, value=field_info['label'])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_style
    ws.row_dimensions[3].height = 22

    # Set lebar kolom pertama (Nama Kegiatan) lebih lebar
    ws.column_dimensions['A'].width = 35
    # Set lebar semua kolom lainnya
    for col_idx in range(2, len(all_fields) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 25

    # Data untuk setiap user (setiap user = 1 baris)
    current_row = 4
    for user_idx, biodata in enumerate(all_biodata):
        for col_idx, field_info in enumerate(all_fields, 1):
            field_key = field_info['key']

            # Ambil nilai field
            if field_key == 'tanda_tangan':
                display_value = "[Tersedia]" if biodata.get('tanda_tangan') else "-"
            else:
                value = biodata.get(field_key, '')
                display_value = str(value) if value and str(value).strip() else '-'

            # Tulis ke cell
            cell = ws.cell(row=current_row, column=col_idx, value=display_value)
            cell.border = border_style

            # Alternating row colors untuk readability
            if user_idx % 2 == 0:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # Alignment dengan wrap text untuk semua kolom
            if field_key in ['nik', 'nip_nippk', 'no_hp', 'npwp', 'no_rekening']:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            elif field_key in ['tanggal_lahir']:
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Auto adjust row height
        max_height = 18
        for col_idx, field_info in enumerate(all_fields, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            if cell.value:
                col_letter = get_column_letter(col_idx)
                col_width = ws.column_dimensions[col_letter].width or 25
                text_length = len(str(cell.value))
                estimated_lines = max(1, (text_length / max(col_width * 0.8, 1)))
                estimated_height = estimated_lines * 15
                max_height = max(max_height, min(estimated_height, 60))

        ws.row_dimensions[current_row].height = max_height
        current_row += 1

    # Auto-adjust column widths berdasarkan konten aktual
    for col_idx, field_info in enumerate(all_fields, 1):
        col_letter = get_column_letter(col_idx)

        if col_letter == 'A':
            continue  # Skip kolom A (Nama Kegiatan) - sudah di-set

        max_length = 0
        header_cell = ws.cell(row=3, column=col_idx)
        if header_cell.value:
            max_length = len(str(header_cell.value))

        for row in range(4, current_row):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value:
                text = str(cell.value)
                text_length = len(text)
                if text_length > 50:
                    lines = text.split('\n') if '\n' in text else [text]
                    max_line_length = max(len(line) for line in lines)
                    max_length = max(max_length, min(max_line_length, 40))
                else:
                    max_length = max(max_length, text_length)

        optimal_width = max(12, min(max_length + 2, 40))
        if 'Alamat' in field_info['label']:
            optimal_width = max(optimal_width, min(max_length + 2, 45))
        ws.column_dimensions[col_letter].width = optimal_width

    # Freeze panes untuk memudahkan scrolling (freeze header)
    ws.freeze_panes = 'A4'

    # Enable auto filter untuk header
    ws.auto_filter.ref = f'A3:{last_col_letter}{current_row - 1}'

    # Set print settings
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # Set row heights untuk header
    ws.row_dimensions[1].height = 25

    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Refresh session
    session.permanent = True
    session.modified = True

    # Generate filename
    filename = f"Rekap_Kabupaten_{kabupaten.replace(' ', '_')}.xlsx"

    return Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
    )

@app.route('/admin/rekap-filter', methods=['GET'])
@admin_required
def admin_rekap_filter():
    """Halaman rekap biodata dengan filter tahun/kabupaten/kegiatan (DataTables)."""
    user_role = get_user_role()
    user_id = get_user_id()

    selected_year = request.args.get('tahun', '').strip()
    selected_kabupaten = request.args.get('kabupaten_kota', '').strip()
    selected_kegiatan = request.args.get('nama_kegiatan', '').strip()

    connection = get_db_connection()
    if not connection:
        flash('Koneksi database gagal!', 'error')
        return render_template(
            'admin/admin-rekap-filter.html',
            biodata_list=[],
            tahun_list=[],
            kabupaten_list=[],
            kegiatan_list=[],
            selected_year=selected_year,
            selected_kabupaten=selected_kabupaten,
            selected_kegiatan=selected_kegiatan,
            username=get_username(),
            user_role=user_role
        )

    biodata_list = []
    tahun_list = []
    kabupaten_list = []
    kegiatan_list = []

    try:
        cursor = connection.cursor()

        # Dropdown Tahun (berdasarkan created_at biodata_kegiatan)
        tahun_query = """
            SELECT DISTINCT CAST(strftime('%Y', created_at) AS INTEGER) AS tahun
            FROM biodata_kegiatan
            WHERE created_at IS NOT NULL
            ORDER BY tahun DESC
        """
        cursor.execute(tahun_query)
        tahun_list = [row[0] for row in cursor.fetchall() if row and row[0]]

        # Dropdown Kabupaten/Kota
        kab_query = """
            SELECT DISTINCT kabupaten_kota
            FROM biodata_kegiatan
            WHERE TRIM(kabupaten_kota) != '' AND kabupaten_kota IS NOT NULL
            ORDER BY kabupaten_kota ASC
        """
        cursor.execute(kab_query)
        kabupaten_list = [row[0] for row in cursor.fetchall() if row and row[0]]

        # Dropdown Nama Kegiatan
        keg_query = """
            SELECT DISTINCT nama_kegiatan
            FROM biodata_kegiatan
            WHERE TRIM(nama_kegiatan) != '' AND nama_kegiatan IS NOT NULL
            ORDER BY nama_kegiatan ASC
        """
        cursor.execute(keg_query)
        kegiatan_list = [row[0] for row in cursor.fetchall() if row and row[0]]

        where_conditions = ["1=1"]
        params = []

        # Filter tahun
        if selected_year:
            try:
                year_int = int(selected_year)
                where_conditions.append("CAST(strftime('%Y', bk.created_at) AS INTEGER) = ?")
                params.append(year_int)
            except ValueError:
                pass

        # Filter kabupaten/kota
        if selected_kabupaten:
            where_conditions.append("TRIM(bk.kabupaten_kota) = TRIM(?)")
            params.append(selected_kabupaten)

        # Filter nama kegiatan
        if selected_kegiatan:
            where_conditions.append("TRIM(bk.nama_kegiatan) = TRIM(?)")
            params.append(selected_kegiatan)

        # Jika operator, batasi hanya kegiatan yang dia pegang
        if user_role == 'operator' and user_id:
            where_conditions.append("""
                EXISTS (
                    SELECT 1
                    FROM kegiatan_master k
                    INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                    WHERE ok.user_id = ?
                      AND TRIM(k.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                )
            """)
            params.append(user_id)

        where_clause = " AND ".join(where_conditions)

        biodata_query = f"""
            SELECT bk.*
            FROM biodata_kegiatan bk
            WHERE {where_clause}
            ORDER BY bk.created_at DESC, bk.id DESC
        """
        cursor.execute(biodata_query, tuple(params))
        rows = cursor.fetchall()
        biodata_list = [row_to_dict(r) for r in rows]

        # Normalisasi path buku tabungan agar bisa ditampilkan/diakses konsisten
        for b in biodata_list:
            if b.get('buku_tabungan_path'):
                b['buku_tabungan_path'] = normalize_buku_tabungan_path(b['buku_tabungan_path'])

    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan saat mengambil data: {str(e)}', 'error')
        biodata_list = []
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        connection.close()

    return render_template(
        'admin/admin-rekap-filter.html',
        biodata_list=biodata_list,
        tahun_list=tahun_list,
        kabupaten_list=kabupaten_list,
        kegiatan_list=kegiatan_list,
        selected_year=selected_year,
        selected_kabupaten=selected_kabupaten,
        selected_kegiatan=selected_kegiatan,
        username=get_username(),
        user_role=user_role
    )

@app.route('/admin/export-rekap-filter-pdf')
@admin_required
def export_rekap_filter_pdf():
    """Export Rekap ke PDF (format biodata lengkap seperti rekap tahunan) sesuai filter."""
    from io import BytesIO
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import inch

    user_role = get_user_role()
    user_id = get_user_id()

    selected_year = request.args.get('tahun', '').strip()
    selected_kabupaten = request.args.get('kabupaten_kota', '').strip()
    selected_kegiatan = request.args.get('nama_kegiatan', '').strip()

    connection = get_db_connection()
    if not connection:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_rekap_filter'))

    try:
        cursor = connection.cursor()
        where_conditions = ["1=1"]
        params = []

        if selected_year:
            try:
                year_int = int(selected_year)
                where_conditions.append("CAST(strftime('%Y', bk.created_at) AS INTEGER) = ?")
                params.append(year_int)
            except ValueError:
                pass

        if selected_kabupaten:
            where_conditions.append("TRIM(bk.kabupaten_kota) = TRIM(?)")
            params.append(selected_kabupaten)

        if selected_kegiatan:
            where_conditions.append("TRIM(bk.nama_kegiatan) = TRIM(?)")
            params.append(selected_kegiatan)

        if user_role == 'operator' and user_id:
            where_conditions.append("""
                EXISTS (
                    SELECT 1
                    FROM kegiatan_master k
                    INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                    WHERE ok.user_id = ?
                      AND TRIM(k.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                )
            """)
            params.append(user_id)

        where_clause = " AND ".join(where_conditions)
        cursor.execute(f"""
            SELECT bk.*
            FROM biodata_kegiatan bk
            WHERE {where_clause}
            ORDER BY bk.nama_kegiatan ASC, bk.kabupaten_kota ASC, bk.nama_lengkap ASC, bk.id DESC
        """, tuple(params))
        rows = cursor.fetchall()
        all_biodata = [row_to_dict(r) for r in rows]

        # Normalisasi path buku tabungan
        for biodata in all_biodata:
            if biodata.get('buku_tabungan_path'):
                biodata['buku_tabungan_path'] = normalize_buku_tabungan_path(biodata['buku_tabungan_path'])

        if not all_biodata:
            flash('Tidak ada data untuk diekspor!', 'error')
            return redirect(url_for('admin_rekap_filter', tahun=selected_year, kabupaten_kota=selected_kabupaten, nama_kegiatan=selected_kegiatan))

        MAX_EXPORT_ROWS = 2000
        if len(all_biodata) > MAX_EXPORT_ROWS:
            flash(f'Data terlalu besar ({len(all_biodata)} rows). Maksimal {MAX_EXPORT_ROWS} peserta untuk export PDF. Silakan gunakan filter yang lebih spesifik.', 'error')
            return redirect(url_for('admin_rekap_filter', tahun=selected_year, kabupaten_kota=selected_kabupaten, nama_kegiatan=selected_kegiatan))

        # ==== Mulai: blok gaya PDF sama dengan Rekap Tahunan ====
        # Load logo untuk header
        logo_path = os.path.join(BASE_DIR, 'static', 'Logo_BGTK.png')
        logo_img = None
        logo_width = 0
        logo_height = 0
        if os.path.exists(logo_path):
            try:
                logo_pil = Image.open(logo_path)
                max_logo_height = 0.6 * inch
                logo_ratio = logo_pil.width / logo_pil.height
                logo_width = max_logo_height * logo_ratio
                logo_height = max_logo_height

                logo_buffer = io.BytesIO()
                logo_pil.save(logo_buffer, format='PNG')
                logo_buffer.seek(0)
                logo_img = RLImage(logo_buffer, width=logo_width, height=logo_height)
            except Exception as e:
                print(f"Error loading logo: {e}")
                logo_img = None

        # Load logo Pendidikan Bermutu
        pendidikan_bermutu_path = os.path.join(BASE_DIR, 'static', 'Pendidikan Bermutu untuk Semua.png')
        pendidikan_bermutu_processed_path = None
        pendidikan_bermutu_width = 0
        pendidikan_bermutu_height = 0
        pendidikan_bermutu_temp_file = None
        if os.path.exists(pendidikan_bermutu_path):
            try:
                pendidikan_bermutu_pil = Image.open(pendidikan_bermutu_path)
                if pendidikan_bermutu_pil.mode != 'RGBA':
                    pendidikan_bermutu_pil = pendidikan_bermutu_pil.convert('RGBA')

                pixels = pendidikan_bermutu_pil.load()
                width, height = pendidikan_bermutu_pil.size
                black_threshold = 30
                for x in range(width):
                    for y in range(height):
                        r, g, b, a = pixels[x, y]
                        if r < black_threshold and g < black_threshold and b < black_threshold:
                            pixels[x, y] = (r, g, b, 0)

                max_logo_height = 0.5 * inch
                pendidikan_bermutu_ratio = pendidikan_bermutu_pil.width / pendidikan_bermutu_pil.height
                pendidikan_bermutu_width = max_logo_height * pendidikan_bermutu_ratio
                pendidikan_bermutu_height = max_logo_height

                pendidikan_bermutu_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
                pendidikan_bermutu_pil.save(pendidikan_bermutu_temp_file.name, format='PNG', optimize=False)
                pendidikan_bermutu_processed_path = pendidikan_bermutu_temp_file.name
            except Exception as e:
                print(f"Error loading logo Pendidikan Bermutu: {e}")

        # Load logo Ramah
        ramah_path = os.path.join(BASE_DIR, 'static', 'Ramah.png')
        ramah_processed_path = None
        ramah_width = 0
        ramah_height = 0
        ramah_temp_file = None
        if os.path.exists(ramah_path):
            try:
                ramah_pil = Image.open(ramah_path)
                if ramah_pil.mode != 'RGBA':
                    ramah_pil = ramah_pil.convert('RGBA')

                pixels = ramah_pil.load()
                width, height = ramah_pil.size
                black_threshold = 30
                for x in range(width):
                    for y in range(height):
                        r, g, b, a = pixels[x, y]
                        if r < black_threshold and g < black_threshold and b < black_threshold:
                            pixels[x, y] = (r, g, b, 0)

                max_logo_height = 0.5 * inch
                ramah_ratio = ramah_pil.width / ramah_pil.height
                ramah_width = max_logo_height * ramah_ratio
                ramah_height = max_logo_height

                ramah_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
                ramah_pil.save(ramah_temp_file.name, format='PNG', optimize=False)
                ramah_processed_path = ramah_temp_file.name
            except Exception as e:
                print(f"Error loading logo Ramah: {e}")

        def add_header_footer(canvas, doc):
            canvas.saveState()

            max_logo_h = logo_height if logo_height > 0 else 0

            if logo_img and logo_height > 0:
                try:
                    logo_x = 25
                    logo_y = F4_SIZE[1] - 25 - max_logo_h
                    canvas.drawImage(logo_path, logo_x, logo_y, width=logo_width, height=logo_height, preserveAspectRatio=True)
                except Exception as e:
                    print(f"Error drawing logo: {e}")

            canvas.setStrokeColor(colors.HexColor('#067ac1'))
            canvas.setLineWidth(1.5)
            max_header_height = logo_height if logo_height > 0 else 0
            header_line_y = F4_SIZE[1] - 25 - max_header_height - 5
            canvas.line(25, header_line_y, F4_SIZE[0] - 25, header_line_y)

            max_footer_logo_height = max(
                pendidikan_bermutu_height if pendidikan_bermutu_height > 0 else 0,
                ramah_height if ramah_height > 0 else 0
            )
            footer_logo_y = 45

            if max_footer_logo_height > 0:
                total_footer_logo_width = (
                    (pendidikan_bermutu_width if pendidikan_bermutu_height > 0 else 0) +
                    (ramah_width if ramah_height > 0 else 0) + 10
                )
                footer_logo_start_x = F4_SIZE[0] - 25 - total_footer_logo_width

                if pendidikan_bermutu_processed_path and os.path.exists(pendidikan_bermutu_processed_path) and pendidikan_bermutu_height > 0:
                    try:
                        pendidikan_bermutu_footer_x = footer_logo_start_x
                        pendidikan_bermutu_footer_y = footer_logo_y
                        canvas.drawImage(
                            pendidikan_bermutu_processed_path,
                            pendidikan_bermutu_footer_x,
                            pendidikan_bermutu_footer_y,
                            width=pendidikan_bermutu_width,
                            height=pendidikan_bermutu_height,
                            preserveAspectRatio=True,
                            mask='auto'
                        )
                    except Exception as e:
                        print(f"Error drawing logo Pendidikan Bermutu di footer: {e}")

                if ramah_processed_path and os.path.exists(ramah_processed_path) and ramah_height > 0:
                    try:
                        ramah_footer_x = footer_logo_start_x + (pendidikan_bermutu_width if pendidikan_bermutu_height > 0 else 0) + 10
                        ramah_footer_y = footer_logo_y
                        canvas.drawImage(
                            ramah_processed_path,
                            ramah_footer_x,
                            ramah_footer_y,
                            width=ramah_width,
                            height=ramah_height,
                            preserveAspectRatio=True,
                            mask='auto'
                        )
                    except Exception as e:
                        print(f"Error drawing logo Ramah di footer: {e}")

            canvas.setStrokeColor(colors.HexColor('#067ac1'))
            canvas.setLineWidth(1)
            footer_line_y = footer_logo_y - 10
            canvas.line(25, footer_line_y, F4_SIZE[0] - 25, footer_line_y)

            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(colors.black)
            wita_time = datetime.utcnow() + timedelta(hours=8)
            footer_text = f"Dicetak pada: {wita_time.strftime('%d/%m/%Y %H:%M')} WITA"
            footer_x = 25
            footer_y = footer_line_y - 15
            canvas.drawString(footer_x, footer_y, footer_text)

            canvas.restoreState()

        buffer = io.BytesIO()
        max_header_logo_height = logo_height if logo_height > 0 else 0
        top_margin_with_logo = 25 + max_header_logo_height + 15
        doc = SimpleDocTemplate(
            buffer,
            pagesize=F4_SIZE,
            rightMargin=25,
            leftMargin=25,
            topMargin=top_margin_with_logo,
            bottomMargin=40
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.black,
            spaceAfter=12,
            spaceBefore=0,
            alignment=1,
            fontName='Helvetica-Bold',
            leading=24
        )
        value_style = ParagraphStyle(
            'ValueStyle',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            textColor=colors.black,
            alignment=0,
            leftIndent=0,
            rightIndent=0
        )

        field_labels = {
            'nik': 'NIK',
            'nama_lengkap': 'Nama Lengkap',
            'nip_nippk': 'NIP/NIPPK',
            'tempat_lahir': 'Tempat Lahir',
            'tanggal_lahir': 'Tanggal Lahir',
            'jenis_kelamin': 'Jenis Kelamin',
            'agama': 'Agama',
            'pendidikan_terakhir': 'Pendidikan Terakhir',
            'jurusan': 'Jurusan',
            'status_asn': 'Status Kepegawaian',
            'pangkat_golongan': 'Pangkat / Golongan',
            'jabatan': 'Jabatan',
            'instansi': 'Nama Instansi',
            'alamat_instansi': 'Alamat Instansi',
            'alamat_domisili': 'Alamat Domisili',
            'kabupaten_kota': 'Kabupaten/Kota',
            'kabko_lainnya': 'Kabupaten/Kota Lainnya',
            'peran': 'Peran dalam Kegiatan',
            'no_hp': 'Nomor HP',
            'alamat_email': 'Alamat Email',
            'npwp': 'NPWP',
            'nama_bank': 'Nama Bank',
            'nama_bank_lainnya': 'Nama Bank Lainnya',
            'no_rekening': 'Nomor Rekening',
            'nama_pemilik_rekening': 'Nama Pemilik Rekening'
        }
        field_order = [
            'nik', 'nama_lengkap', 'nip_nippk', 'tempat_lahir', 'tanggal_lahir',
            'jenis_kelamin', 'agama', 'pendidikan_terakhir', 'jurusan',
            'status_asn', 'pangkat_golongan', 'jabatan', 'instansi',
            'alamat_instansi', 'alamat_domisili', 'kabupaten_kota', 'kabko_lainnya',
            'peran',
            'no_hp', 'alamat_email', 'npwp',
            'nama_bank', 'nama_bank_lainnya', 'no_rekening', 'nama_pemilik_rekening'
        ]
        exclude_fields = ['id', 'user_id', 'buku_tabungan_path', 'tanda_tangan', 'created_at', 'updated_at', 'nama_kegiatan', 'waktu_pelaksanaan', 'tempat_pelaksanaan']

        elements = []
        tanda_tangan_temp_files = []

        for user_idx, biodata in enumerate(all_biodata):
            if user_idx > 0:
                elements.append(PageBreak())

            title = Paragraph("BIODATA KEGIATAN", title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.15 * inch))

            biodata_nama_kegiatan = biodata.get('nama_kegiatan', '-')
            waktu_pelaksanaan = biodata.get('waktu_pelaksanaan', '-')
            tempat_pelaksanaan = biodata.get('tempat_pelaksanaan', '-')

            info_table_data = []
            if biodata_nama_kegiatan and str(biodata_nama_kegiatan).strip() and str(biodata_nama_kegiatan).strip() != '-':
                info_table_data.append([
                    Paragraph("<b>Nama Kegiatan</b>", value_style),
                    Paragraph(":", value_style),
                    Paragraph(str(biodata_nama_kegiatan), value_style)
                ])
            if waktu_pelaksanaan and str(waktu_pelaksanaan).strip() and str(waktu_pelaksanaan).strip() != '-':
                info_table_data.append([
                    Paragraph("<b>Waktu Pelaksanaan</b>", value_style),
                    Paragraph(":", value_style),
                    Paragraph(str(waktu_pelaksanaan), value_style)
                ])
            if tempat_pelaksanaan and str(tempat_pelaksanaan).strip() and str(tempat_pelaksanaan).strip() != '-':
                info_table_data.append([
                    Paragraph("<b>Tempat Pelaksanaan</b>", value_style),
                    Paragraph(":", value_style),
                    Paragraph(str(tempat_pelaksanaan), value_style)
                ])

            if info_table_data:
                available_width = (8.27 * inch) - (25 * 2)
                info_table = Table(info_table_data, colWidths=[2.2 * inch, 0.15 * inch, available_width - 2.2 * inch - 0.15 * inch])
                info_table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('ALIGN', (2, 0), (2, -1), 'LEFT'),
                    ('LEFTPADDING', (0, 0), (0, -1), 0),
                    ('LEFTPADDING', (1, 0), (1, -1), 0),
                    ('LEFTPADDING', (2, 0), (2, -1), 0),
                    ('RIGHTPADDING', (0, 0), (0, -1), 4),
                    ('RIGHTPADDING', (1, 0), (1, -1), 0),
                    ('RIGHTPADDING', (2, 0), (2, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ]))
                elements.append(info_table)
                elements.append(Spacer(1, 0.1 * inch))

            all_data = []
            for key in field_order:
                if key in biodata and key not in exclude_fields:
                    label = field_labels.get(key, key.replace('_', ' ').title())
                    value = biodata[key]
                    display_value = str(value) if value and str(value).strip() else '-'
                    all_data.append([label, display_value])

            for key, value in biodata.items():
                if key not in exclude_fields and key not in field_order:
                    label = field_labels.get(key, key.replace('_', ' ').title())
                    display_value = str(value) if value and str(value).strip() else '-'
                    all_data.append([label, display_value])

            tanda_tangan_raw = biodata.get('tanda_tangan')
            tanda_tangan_img, error_msg = process_tanda_tangan_for_pdf(tanda_tangan_raw, tanda_tangan_temp_files)

            if all_data:
                table_data = []
                for row in all_data:
                    field_label = Paragraph(f"<b>{row[0]}</b>", value_style)
                    value_text = str(row[1]) if row[1] else '-'
                    value_text = value_text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    if len(value_text) > 50:
                        words = value_text.split()
                        lines = []
                        current_line = []
                        current_length = 0
                        for word in words:
                            if current_length + len(word) + 1 > 50:
                                if current_line:
                                    lines.append(' '.join(current_line))
                                current_line = [word]
                                current_length = len(word)
                            else:
                                current_line.append(word)
                                current_length += len(word) + 1
                        if current_line:
                            lines.append(' '.join(current_line))
                        value_text = '<br/>'.join(lines)
                    value_para = Paragraph(value_text, value_style)
                    table_data.append([field_label, value_para])

                if tanda_tangan_img:
                    field_label = Paragraph("<b>Tanda Tangan</b>", value_style)
                    table_data.append([field_label, tanda_tangan_img])
                elif biodata.get('tanda_tangan'):
                    field_label = Paragraph("<b>Tanda Tangan</b>", value_style)
                    error_msg_para = Paragraph("<i>Tanda tangan tidak tersedia</i>", value_style)
                    table_data.append([field_label, error_msg_para])

                available_width = (8.27 * inch) - (25 * 2)
                table = Table(table_data, colWidths=[2.2 * inch, available_width - 2.2 * inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (0, -1), 4),
                    ('LEFTPADDING', (1, 0), (1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
                    ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#067ac1')),
                ]))
                elements.append(table)
                elements.append(Spacer(1, 0.08 * inch))

        try:
            doc.build(elements, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
            buffer.seek(0)
        finally:
            try:
                if pendidikan_bermutu_temp_file and os.path.exists(pendidikan_bermutu_processed_path):
                    os.unlink(pendidikan_bermutu_processed_path)
            except Exception as e:
                print(f"Error cleaning up Pendidikan Bermutu temp file: {e}")
            try:
                if ramah_temp_file and os.path.exists(ramah_processed_path):
                    os.unlink(ramah_processed_path)
            except Exception as e:
                print(f"Error cleaning up Ramah temp file: {e}")
            for temp_file_path in tanda_tangan_temp_files:
                try:
                    if os.path.exists(temp_file_path):
                        os.unlink(temp_file_path)
                except Exception as e:
                    print(f"Error cleaning up tanda tangan temp file {temp_file_path}: {e}")

        filename = "Rekap_Filter.pdf"
        return Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename=\"{filename}\"'}
        )
    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan saat export: {str(e)}', 'error')
        return redirect(url_for('admin_rekap_filter'))
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        connection.close()

@app.route('/admin/export-rekap-filter-excel')
@admin_required
def export_rekap_filter_excel():
    """Export Rekap ke Excel (format biodata lengkap seperti rekap tahunan) sesuai filter."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    user_role = get_user_role()
    user_id = get_user_id()

    selected_year = request.args.get('tahun', '').strip()
    selected_kabupaten = request.args.get('kabupaten_kota', '').strip()
    selected_kegiatan = request.args.get('nama_kegiatan', '').strip()

    connection = get_db_connection()
    if not connection:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_rekap_filter'))

    try:
        cursor = connection.cursor()
        where_conditions = ["1=1"]
        params = []

        if selected_year:
            try:
                year_int = int(selected_year)
                where_conditions.append("CAST(strftime('%Y', bk.created_at) AS INTEGER) = ?")
                params.append(year_int)
            except ValueError:
                pass

        if selected_kabupaten:
            where_conditions.append("TRIM(bk.kabupaten_kota) = TRIM(?)")
            params.append(selected_kabupaten)

        if selected_kegiatan:
            where_conditions.append("TRIM(bk.nama_kegiatan) = TRIM(?)")
            params.append(selected_kegiatan)

        if user_role == 'operator' and user_id:
            where_conditions.append("""
                EXISTS (
                    SELECT 1
                    FROM kegiatan_master k
                    INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                    WHERE ok.user_id = ?
                      AND TRIM(k.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                )
            """)
            params.append(user_id)

        where_clause = " AND ".join(where_conditions)
        cursor.execute(f"""
            SELECT bk.*
            FROM biodata_kegiatan bk
            WHERE {where_clause}
            ORDER BY bk.nama_kegiatan ASC, bk.kabupaten_kota ASC, bk.nama_lengkap ASC, bk.id DESC
        """, tuple(params))
        rows = cursor.fetchall()
        all_biodata = [row_to_dict(r) for r in rows]

        if not all_biodata:
            flash('Tidak ada data untuk diekspor!', 'error')
            return redirect(url_for('admin_rekap_filter', tahun=selected_year, kabupaten_kota=selected_kabupaten, nama_kegiatan=selected_kegiatan))

        MAX_EXPORT_ROWS = 50000
        if len(all_biodata) > MAX_EXPORT_ROWS:
            flash(f'Data terlalu besar ({len(all_biodata)} rows). Maksimal {MAX_EXPORT_ROWS} rows untuk export Excel. Silakan gunakan filter yang lebih spesifik.', 'error')
            return redirect(url_for('admin_rekap_filter', tahun=selected_year, kabupaten_kota=selected_kabupaten, nama_kegiatan=selected_kegiatan))

        # Field mapping (samakan dengan rekap tahunan excel)
        field_labels = {
            'nik': 'NIK',
            'nama_lengkap': 'Nama Lengkap',
            'nip_nippk': 'NIP/NIPPK',
            'tempat_lahir': 'Tempat Lahir',
            'tanggal_lahir': 'Tanggal Lahir',
            'jenis_kelamin': 'Jenis Kelamin',
            'agama': 'Agama',
            'pendidikan_terakhir': 'Pendidikan Terakhir',
            'jurusan': 'Jurusan',
            'status_asn': 'Status Kepegawaian',
            'pangkat_golongan': 'Pangkat / Golongan',
            'jabatan': 'Jabatan',
            'instansi': 'Nama Instansi',
            'alamat_instansi': 'Alamat Instansi',
            'alamat_domisili': 'Alamat Domisili',
            'kabupaten_kota': 'Kabupaten/Kota',
            'kabko_lainnya': 'Kabupaten/Kota Lainnya',
            'peran': 'Peran dalam Kegiatan',
            'no_hp': 'Nomor HP',
            'alamat_email': 'Alamat Email',
            'npwp': 'NPWP',
            'nama_bank': 'Nama Bank',
            'nama_bank_lainnya': 'Nama Bank Lainnya',
            'no_rekening': 'Nomor Rekening',
            'nama_pemilik_rekening': 'Nama Pemilik Rekening'
        }
        field_order = [
            'nik', 'nama_lengkap', 'nip_nippk', 'tempat_lahir', 'tanggal_lahir',
            'jenis_kelamin', 'agama', 'pendidikan_terakhir', 'jurusan',
            'status_asn', 'pangkat_golongan', 'jabatan', 'instansi',
            'alamat_instansi', 'alamat_domisili', 'kabupaten_kota', 'kabko_lainnya',
            'peran',
            'no_hp', 'alamat_email', 'npwp',
            'nama_bank', 'nama_bank_lainnya', 'no_rekening', 'nama_pemilik_rekening'
        ]
        exclude_fields = ['id', 'user_id', 'buku_tabungan_path', 'tanda_tangan', 'created_at', 'updated_at', 'waktu_pelaksanaan', 'tempat_pelaksanaan']

        # === Format Excel disamakan dengan export Rekap Tahunan ===
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Biodata"

        # Styles (samakan dengan rekap tahunan excel)
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        header_font = Font(bold=True, color="000000", size=11)
        title_font = Font(bold=True, size=12, color="000000")
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Field list (nama_kegiatan + biodata fields + tanda tangan)
        all_fields = []
        all_fields.append({'key': 'nama_kegiatan', 'label': 'Nama Kegiatan'})
        for key in field_order:
            if key not in exclude_fields:
                label = field_labels.get(key, key.replace('_', ' ').title())
                all_fields.append({'key': key, 'label': label})
        all_fields.append({'key': 'tanda_tangan', 'label': 'Tanda Tangan'})

        num_cols = len(all_fields)
        last_col_letter = get_column_letter(num_cols)

        # Title (row 1) + info export (row 2) seperti rekap tahunan
        title_text = "Rekap"
        filter_parts = []
        if selected_year:
            filter_parts.append(f"Tahun {selected_year}")
        if selected_kabupaten:
            filter_parts.append(f"Kab/Kota: {selected_kabupaten}")
        if selected_kegiatan:
            filter_parts.append(f"Kegiatan: {selected_kegiatan}")
        if filter_parts:
            title_text += " - " + " | ".join(filter_parts)

        ws.merge_cells(f'A1:{last_col_letter}1')
        ws['A1'] = title_text
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align

        export_date = datetime.now().strftime("%d %B %Y, %H:%M:%S")
        ws.merge_cells(f'A2:{last_col_letter}2')
        ws['A2'] = f"Tanggal Export: {export_date}"
        ws['A2'].font = Font(size=10, italic=True)
        ws['A2'].alignment = left_align
        ws.row_dimensions[2].height = 18

        # Header kolom (row 3)
        for col_idx, field_info in enumerate(all_fields, 1):
            cell = ws.cell(row=3, column=col_idx, value=field_info['label'])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border_style
        ws.row_dimensions[3].height = 22

        # Lebar kolom default (samakan)
        ws.column_dimensions['A'].width = 35  # Nama Kegiatan
        for col_idx in range(2, len(all_fields) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 25

        # Data rows (mulai row 4)
        current_row = 4
        for user_idx, biodata in enumerate(all_biodata):
            for col_idx, field_info in enumerate(all_fields, 1):
                field_key = field_info['key']

                if field_key == 'tanda_tangan':
                    display_value = "[Tersedia]" if biodata.get('tanda_tangan') else "-"
                else:
                    value = biodata.get(field_key, '')
                    display_value = str(value) if value and str(value).strip() else '-'

                cell = ws.cell(row=current_row, column=col_idx, value=display_value)
                cell.border = border_style

                # Alternating row colors
                if user_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

                # Alignment
                if field_key in ['nik', 'nip_nippk', 'no_hp', 'npwp', 'no_rekening']:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                elif field_key in ['tanggal_lahir']:
                    cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # Auto adjust row height
            max_height = 18
            for col_idx, _field_info in enumerate(all_fields, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                if cell.value:
                    col_letter = get_column_letter(col_idx)
                    col_width = ws.column_dimensions[col_letter].width or 25
                    text_length = len(str(cell.value))
                    estimated_lines = max(1, (text_length / max(col_width * 0.8, 1)))
                    estimated_height = estimated_lines * 15
                    max_height = max(max_height, min(estimated_height, 60))
            ws.row_dimensions[current_row].height = max_height
            current_row += 1

        # Auto-adjust column widths berdasarkan konten (samakan)
        for col_idx, field_info in enumerate(all_fields, 1):
            col_letter = get_column_letter(col_idx)
            if col_letter == 'A':
                continue

            max_length = 0
            header_cell = ws.cell(row=3, column=col_idx)
            if header_cell.value:
                max_length = len(str(header_cell.value))

            for row in range(4, current_row):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value:
                    text = str(cell.value)
                    text_length = len(text)
                    if text_length > 50:
                        lines = text.split('\n') if '\n' in text else [text]
                        max_line_length = max(len(line) for line in lines)
                        max_length = max(max_length, min(max_line_length, 40))
                    else:
                        max_length = max(max_length, text_length)

            optimal_width = max(12, min(max_length + 2, 40))
            if 'Alamat' in field_info['label']:
                optimal_width = max(optimal_width, min(max_length + 2, 45))
            ws.column_dimensions[col_letter].width = optimal_width

        # Freeze panes untuk header
        ws.freeze_panes = 'A4'

        # Auto filter untuk header
        ws.auto_filter.ref = f'A3:{last_col_letter}{current_row - 1}'

        # Print settings
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # Header row height
        ws.row_dimensions[1].height = 25

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = "Rekap_Filter.xlsx"
        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=\"{filename}\"'}
        )
    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan saat export: {str(e)}', 'error')
        return redirect(url_for('admin_rekap_filter'))
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        connection.close()

@app.route('/admin/rekap-tahunan')
@admin_required
def admin_rekap_tahunan():
    """Halaman rekap kegiatan per tahun dengan filter bulan"""

    user_role = get_user_role()
    user_id = get_user_id()

    # Ambil parameter tahun dari URL (default: tahun saat ini)
    selected_year = request.args.get('tahun', '').strip()
    if selected_year:
        try:
            selected_year = int(selected_year)
        except ValueError:
            selected_year = datetime.now().year
    else:
        selected_year = datetime.now().year

    # Ambil parameter bulan awal dan bulan akhir
    bulan_awal = request.args.get('bulan_awal', '').strip()
    bulan_akhir = request.args.get('bulan_akhir', '').strip()

    if bulan_awal:
        try:
            bulan_awal = int(bulan_awal)
            if bulan_awal < 1 or bulan_awal > 12:
                bulan_awal = None
        except ValueError:
            bulan_awal = None
    else:
        bulan_awal = None

    if bulan_akhir:
        try:
            bulan_akhir = int(bulan_akhir)
            if bulan_akhir < 1 or bulan_akhir > 12:
                bulan_akhir = None
        except ValueError:
            bulan_akhir = None
    else:
        bulan_akhir = None

    # Jika hanya satu bulan yang dipilih, set keduanya sama
    if bulan_awal and not bulan_akhir:
        bulan_akhir = bulan_awal
    elif bulan_akhir and not bulan_awal:
        bulan_awal = bulan_akhir

    # Pastikan bulan_awal <= bulan_akhir
    if bulan_awal and bulan_akhir and bulan_awal > bulan_akhir:
        bulan_awal, bulan_akhir = bulan_akhir, bulan_awal

    # Inisialisasi data
    tahun_list = []
    tahun_stats = {
        'total_kegiatan': 0,
        'total_peserta': 0,
        'rata_peserta': 0,
        'jumlah_kabupaten': 0
    }
    kegiatan_data = []

    connection = get_db_connection()
    if connection is None:
        flash('Koneksi database gagal!', 'error')
        return render_template(
            'admin/admin-rekap-tahunan.html',
            tahun_list=tahun_list,
            selected_year=selected_year,
            bulan_awal=bulan_awal,
            bulan_akhir=bulan_akhir,
            tahun_stats=tahun_stats,
            kegiatan_data=kegiatan_data,
            username=get_username()
        )

    try:
        cursor = connection.cursor()

        # Ambil daftar tahun yang tersedia dari created_at
        if user_role == 'operator' and user_id:
            # Hanya tahun dari kegiatan yang dipegang operator
            cursor.execute("""
                SELECT DISTINCT CAST(strftime('%Y', bk.created_at) AS INTEGER) as tahun
                FROM biodata_kegiatan bk
                INNER JOIN kegiatan_master k
                    ON TRIM(k.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                INNER JOIN operator_kegiatan ok
                    ON k.id = ok.kegiatan_id
                WHERE bk.created_at IS NOT NULL
                  AND ok.user_id = ?
                ORDER BY tahun DESC
            """, (user_id,))
        else:
            cursor.execute("""
                SELECT DISTINCT strftime('%Y', created_at) as tahun
                FROM biodata_kegiatan
                WHERE created_at IS NOT NULL
                ORDER BY tahun DESC
            """)
        rows = cursor.fetchall()
        tahun_list = []
        for row in rows:
            row_dict = row_to_dict(row)
            tahun_val = row_dict.get('tahun')
            # Convert ke int jika masih string
            if isinstance(tahun_val, str):
                try:
                    tahun_val = int(tahun_val)
                except (ValueError, TypeError):
                    continue
            if tahun_val:
                tahun_list.append(tahun_val)

        # Jika tidak ada data, set tahun_list ke tahun saat ini
        if not tahun_list:
            tahun_list = [datetime.now().year]

        # Jika selected_year tidak ada di daftar, set ke tahun terbaru
        if selected_year not in tahun_list and tahun_list:
            selected_year = int(tahun_list[0]) if isinstance(tahun_list[0], str) else tahun_list[0]

        # Buat kondisi WHERE untuk filter bulan (SQLite: gunakan strftime untuk ekstrak tahun/bulan)
        selected_year_str = str(selected_year)
        where_conditions = ["CAST(strftime('%Y', bk.created_at) AS INTEGER) = ?"]
        params = [selected_year_str]

        if bulan_awal and bulan_akhir:
            # SQLite: bandingkan bulan sebagai integer dari strftime('%m')
            where_conditions.append("CAST(strftime('%m', bk.created_at) AS INTEGER) BETWEEN ? AND ?")
            params.extend([bulan_awal, bulan_akhir])

        # Jika operator, batasi hanya pada kegiatan yang ia pegang
        if user_role == 'operator' and user_id:
            where_conditions.append("""
                EXISTS (
                    SELECT 1
                    FROM kegiatan_master k
                    INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                    WHERE ok.user_id = ?
                      AND TRIM(k.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                )
            """)
            params.append(user_id)

        where_clause = " AND ".join(where_conditions)

        # Ambil statistik untuk tahun yang dipilih (dengan filter bulan jika ada)
        stats_query = f"""
            SELECT
                COUNT(DISTINCT bk.nama_kegiatan) as total_kegiatan,
                COUNT(DISTINCT bk.id) as total_peserta,
                COUNT(DISTINCT bk.kabupaten_kota) as jumlah_kabupaten
            FROM biodata_kegiatan bk
            WHERE {where_clause}
                AND TRIM(bk.kabupaten_kota) != ''
                AND bk.kabupaten_kota IS NOT NULL
        """

        cursor.execute(stats_query, tuple(params))

        stats_row = cursor.fetchone()
        if stats_row:
            stats_result = row_to_dict(stats_row)
            tahun_stats['total_kegiatan'] = stats_result.get('total_kegiatan', 0) or 0
            tahun_stats['total_peserta'] = stats_result.get('total_peserta', 0) or 0
            tahun_stats['jumlah_kabupaten'] = stats_result.get('jumlah_kabupaten', 0) or 0
            if tahun_stats['total_kegiatan'] > 0:
                tahun_stats['rata_peserta'] = round(tahun_stats['total_peserta'] / tahun_stats['total_kegiatan'], 2)

        # Ambil daftar kegiatan per tahun dengan detail (dengan filter bulan jika ada)
        kegiatan_query = f"""
            SELECT
                bk.nama_kegiatan,
                COUNT(DISTINCT bk.id) as jumlah_peserta,
                COUNT(DISTINCT bk.kabupaten_kota) as jumlah_kabupaten,
                MIN(bk.created_at) as tanggal_awal,
                MAX(bk.created_at) as tanggal_akhir
            FROM biodata_kegiatan bk
            WHERE {where_clause}
                AND TRIM(bk.nama_kegiatan) != ''
            GROUP BY bk.nama_kegiatan
            ORDER BY jumlah_peserta DESC, bk.nama_kegiatan ASC
        """

        cursor.execute(kegiatan_query, tuple(params))

        rows = cursor.fetchall()
        # Convert Row to dict dan ambil detail kabupaten per kegiatan
        kegiatan_data = []
        for row in rows:
            row_dict = row_to_dict(row)
            nama_kegiatan = row_dict.get('nama_kegiatan')

            # Ambil detail per kabupaten untuk kegiatan ini
            kabupaten_detail_query = f"""
                SELECT
                    bk.kabupaten_kota,
                    COUNT(DISTINCT bk.id) as jumlah_peserta,
                    MIN(bk.created_at) as tanggal_awal,
                    MAX(bk.created_at) as tanggal_akhir
                FROM biodata_kegiatan bk
                WHERE {where_clause}
                    AND TRIM(bk.nama_kegiatan) = TRIM(?)
                    AND TRIM(bk.kabupaten_kota) != ''
                    AND bk.kabupaten_kota IS NOT NULL
                GROUP BY bk.kabupaten_kota
                ORDER BY bk.kabupaten_kota ASC
            """
            kabupaten_params = list(params) + [nama_kegiatan]
            cursor.execute(kabupaten_detail_query, tuple(kabupaten_params))
            kabupaten_rows = cursor.fetchall()

            # Convert kabupaten rows to dict dan format
            kabupaten_list = []
            for kab_row in kabupaten_rows:
                kab_dict = row_to_dict(kab_row)
                # Convert tanggal string ke datetime object
                if kab_dict.get('tanggal_awal'):
                    try:
                        tanggal_str = kab_dict['tanggal_awal']
                        if isinstance(tanggal_str, str):
                            if ' ' in tanggal_str:
                                kab_dict['tanggal_awal'] = datetime.strptime(tanggal_str, '%Y-%m-%d %H:%M:%S')
                            else:
                                kab_dict['tanggal_awal'] = datetime.strptime(tanggal_str, '%Y-%m-%d')
                    except (ValueError, TypeError):
                        kab_dict['tanggal_awal'] = None
                if kab_dict.get('tanggal_akhir'):
                    try:
                        tanggal_str = kab_dict['tanggal_akhir']
                        if isinstance(tanggal_str, str):
                            if ' ' in tanggal_str:
                                kab_dict['tanggal_akhir'] = datetime.strptime(tanggal_str, '%Y-%m-%d %H:%M:%S')
                            else:
                                kab_dict['tanggal_akhir'] = datetime.strptime(tanggal_str, '%Y-%m-%d')
                    except (ValueError, TypeError):
                        kab_dict['tanggal_akhir'] = None
                kabupaten_list.append(kab_dict)

            row_dict['kabupaten_detail'] = kabupaten_list
            # Convert tanggal kegiatan juga
            if row_dict.get('tanggal_awal'):
                try:
                    tanggal_str = row_dict['tanggal_awal']
                    if isinstance(tanggal_str, str):
                        if ' ' in tanggal_str:
                            row_dict['tanggal_awal'] = datetime.strptime(tanggal_str, '%Y-%m-%d %H:%M:%S')
                        else:
                            row_dict['tanggal_awal'] = datetime.strptime(tanggal_str, '%Y-%m-%d')
                except (ValueError, TypeError):
                    row_dict['tanggal_awal'] = None
            if row_dict.get('tanggal_akhir'):
                try:
                    tanggal_str = row_dict['tanggal_akhir']
                    if isinstance(tanggal_str, str):
                        if ' ' in tanggal_str:
                            row_dict['tanggal_akhir'] = datetime.strptime(tanggal_str, '%Y-%m-%d %H:%M:%S')
                        else:
                            row_dict['tanggal_akhir'] = datetime.strptime(tanggal_str, '%Y-%m-%d')
                except (ValueError, TypeError):
                    row_dict['tanggal_akhir'] = None
            # Convert tanggal string ke datetime object untuk template
            if row_dict.get('tanggal_awal'):
                try:
                    # SQLite datetime format: 'YYYY-MM-DD HH:MM:SS' atau 'YYYY-MM-DD'
                    tanggal_str = row_dict['tanggal_awal']
                    if isinstance(tanggal_str, str):
                        if ' ' in tanggal_str:
                            row_dict['tanggal_awal'] = datetime.strptime(tanggal_str, '%Y-%m-%d %H:%M:%S')
                        else:
                            row_dict['tanggal_awal'] = datetime.strptime(tanggal_str, '%Y-%m-%d')
                except (ValueError, TypeError):
                    row_dict['tanggal_awal'] = None
            if row_dict.get('tanggal_akhir'):
                try:
                    tanggal_str = row_dict['tanggal_akhir']
                    if isinstance(tanggal_str, str):
                        if ' ' in tanggal_str:
                            row_dict['tanggal_akhir'] = datetime.strptime(tanggal_str, '%Y-%m-%d %H:%M:%S')
                        else:
                            row_dict['tanggal_akhir'] = datetime.strptime(tanggal_str, '%Y-%m-%d')
                except (ValueError, TypeError):
                    row_dict['tanggal_akhir'] = None
            kegiatan_data.append(row_dict)

    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan saat mengambil data: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    except Exception as e:
        flash(f'Terjadi kesalahan tidak terduga: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    finally:
        if connection:
            try:
                cursor.close()
            except:
                pass
            connection.close()

    # Daftar nama bulan untuk dropdown
    nama_bulan = [
        ('', 'Semua Bulan'),
        (1, 'Januari'), (2, 'Februari'), (3, 'Maret'), (4, 'April'),
        (5, 'Mei'), (6, 'Juni'), (7, 'Juli'), (8, 'Agustus'),
        (9, 'September'), (10, 'Oktober'), (11, 'November'), (12, 'Desember')
    ]

    return render_template(
        'admin/admin-rekap-tahunan.html',
        tahun_list=tahun_list,
        selected_year=selected_year,
        bulan_awal=bulan_awal,
        bulan_akhir=bulan_akhir,
        nama_bulan=nama_bulan,
        tahun_stats=tahun_stats,
        kegiatan_data=kegiatan_data,
        username=get_username(),
        user_role=user_role
    )

@app.route('/admin/export-rekap-tahunan-pdf')
@admin_required
def export_rekap_tahunan_pdf():
    """Export rekap tahunan ke PDF - semua kegiatan dengan semua biodata"""
    from io import BytesIO
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors

    user_role = get_user_role()
    user_id = get_user_id()

    # Ambil parameter tahun dari URL
    selected_year = request.args.get('tahun', '').strip()
    if selected_year:
        try:
            selected_year = int(selected_year)
        except ValueError:
            selected_year = datetime.now().year
    else:
        selected_year = datetime.now().year

    # Ambil parameter bulan awal dan bulan akhir
    bulan_awal = request.args.get('bulan_awal', '').strip()
    bulan_akhir = request.args.get('bulan_akhir', '').strip()

    if bulan_awal:
        try:
            bulan_awal = int(bulan_awal)
            if bulan_awal < 1 or bulan_awal > 12:
                bulan_awal = None
        except ValueError:
            bulan_awal = None
    else:
        bulan_awal = None

    if bulan_akhir:
        try:
            bulan_akhir = int(bulan_akhir)
            if bulan_akhir < 1 or bulan_akhir > 12:
                bulan_akhir = None
        except ValueError:
            bulan_akhir = None

    # Jika hanya satu bulan yang dipilih, set keduanya sama
    if bulan_awal and not bulan_akhir:
        bulan_akhir = bulan_awal
    elif bulan_akhir and not bulan_awal:
        bulan_awal = bulan_akhir

    # Pastikan bulan_awal <= bulan_akhir
    if bulan_awal and bulan_akhir and bulan_awal > bulan_akhir:
        bulan_awal, bulan_akhir = bulan_akhir, bulan_awal

    # Nama bulan
    bulan_names = {
        1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
        5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
        9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
    }

    # Ambil data dari database
    connection = get_db_connection()
    if not connection:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_rekap_tahunan'))

    # Ambil daftar kegiatan per tahun
    kegiatan_list = []
    try:
        cursor = connection.cursor()

        # Buat kondisi WHERE untuk filter bulan (SQLite)
        selected_year_str = str(selected_year)
        where_conditions = ["CAST(strftime('%Y', bk.created_at) AS INTEGER) = ?"]
        params = [selected_year_str]

        if bulan_awal and bulan_akhir:
            where_conditions.append("CAST(strftime('%m', bk.created_at) AS INTEGER) BETWEEN ? AND ?")
            params.extend([bulan_awal, bulan_akhir])

        # Jika operator, batasi hanya pada kegiatan yang ia pegang
        if user_role == 'operator' and user_id:
            where_conditions.append("""
                EXISTS (
                    SELECT 1
                    FROM kegiatan_master k
                    INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                    WHERE ok.user_id = ?
                      AND TRIM(k.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                )
            """)
            params.append(user_id)

        where_clause = " AND ".join(where_conditions)

        # Ambil daftar kegiatan per tahun dengan detail
        kegiatan_query = f"""
            SELECT
                bk.nama_kegiatan,
                COUNT(DISTINCT bk.id) as jumlah_peserta,
                COUNT(DISTINCT bk.kabupaten_kota) as jumlah_kabupaten,
                MIN(bk.created_at) as tanggal_awal,
                MAX(bk.created_at) as tanggal_akhir
            FROM biodata_kegiatan bk
            WHERE {where_clause}
                AND TRIM(bk.nama_kegiatan) != ''
            GROUP BY bk.nama_kegiatan
            ORDER BY jumlah_peserta DESC, bk.nama_kegiatan ASC
        """

        cursor.execute(kegiatan_query, tuple(params))
        rows = cursor.fetchall()
        kegiatan_list = [row_to_dict(row) for row in rows]

        if not kegiatan_list:
            flash('Tidak ada kegiatan untuk tahun yang dipilih!', 'error')
            return redirect(url_for('admin_rekap_tahunan'))

        # Ambil detail per kabupaten untuk setiap kegiatan dan semua biodata
        kegiatan_data = []
        all_biodata = []

        for kegiatan in kegiatan_list:
            nama_kegiatan = kegiatan.get('nama_kegiatan')

            # Ambil detail per kabupaten untuk kegiatan ini
            kabupaten_detail_query = f"""
                SELECT
                    bk.kabupaten_kota,
                    COUNT(DISTINCT bk.id) as jumlah_peserta,
                    MIN(bk.created_at) as tanggal_awal,
                    MAX(bk.created_at) as tanggal_akhir
                FROM biodata_kegiatan bk
                WHERE {where_clause}
                    AND TRIM(bk.nama_kegiatan) = TRIM(?)
                    AND TRIM(bk.kabupaten_kota) != ''
                    AND bk.kabupaten_kota IS NOT NULL
                GROUP BY bk.kabupaten_kota
                ORDER BY bk.kabupaten_kota ASC
            """
            kabupaten_params = list(params) + [nama_kegiatan]
            cursor.execute(kabupaten_detail_query, tuple(kabupaten_params))
            kabupaten_rows = cursor.fetchall()

            # Convert kabupaten rows to dict
            kabupaten_list = []
            for kab_row in kabupaten_rows:
                kab_dict = row_to_dict(kab_row)
                kabupaten_list.append(kab_dict)

            kegiatan['kabupaten_detail'] = kabupaten_list
            kegiatan_data.append(kegiatan)

            # Ambil semua biodata untuk kegiatan ini
            biodata_query = f"""
                SELECT * FROM biodata_kegiatan
                WHERE TRIM(nama_kegiatan) = TRIM(?)
                    AND CAST(strftime('%Y', created_at) AS INTEGER) = ?
            """
            biodata_params = [nama_kegiatan, selected_year_str]

            if bulan_awal and bulan_akhir:
                biodata_query += " AND CAST(strftime('%m', created_at) AS INTEGER) BETWEEN ? AND ?"
                biodata_params.extend([bulan_awal, bulan_akhir])

            biodata_query += " ORDER BY kabupaten_kota ASC, nama_lengkap ASC"

            cursor.execute(biodata_query, tuple(biodata_params))
            biodata_rows = cursor.fetchall()
            biodata_list = [row_to_dict(row) for row in biodata_rows]

            # Normalisasi path buku tabungan
            for biodata in biodata_list:
                if biodata.get('buku_tabungan_path'):
                    biodata['buku_tabungan_path'] = normalize_buku_tabungan_path(biodata['buku_tabungan_path'])

            all_biodata.extend(biodata_list)

        # Group biodata by kegiatan dan kabupaten untuk PDF export
        biodata_by_kegiatan = {}
        for biodata in all_biodata:
            nama_kegiatan = biodata.get('nama_kegiatan', 'Unknown')
            kabupaten = biodata.get('kabupaten_kota', 'Unknown')

            if nama_kegiatan not in biodata_by_kegiatan:
                biodata_by_kegiatan[nama_kegiatan] = {}
            if kabupaten not in biodata_by_kegiatan[nama_kegiatan]:
                biodata_by_kegiatan[nama_kegiatan][kabupaten] = []

            biodata_by_kegiatan[nama_kegiatan][kabupaten].append(biodata)

    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan saat mengambil data: {str(e)}', 'error')
        return redirect(url_for('admin_rekap_tahunan'))
    finally:
        if connection:
            cursor.close()
            connection.close()

    if not all_biodata:
        flash('Tidak ada data untuk diekspor!', 'error')
        return redirect(url_for('admin_rekap_tahunan'))

    # Limit maksimal untuk export (5000 rows)
    MAX_EXPORT_ROWS = 5000
    if len(all_biodata) > MAX_EXPORT_ROWS:
        flash(f'Data terlalu besar ({len(all_biodata)} rows). Maksimal {MAX_EXPORT_ROWS} rows untuk export. Silakan gunakan filter yang lebih spesifik.', 'error')
        return redirect(url_for('admin_rekap_tahunan'))

    # Warning jika data > 1000 rows
    if len(all_biodata) > 1000:
        print(f"WARNING: Export PDF dengan {len(all_biodata)} rows - mungkin memakan waktu lama")

    # Load logo untuk header
    logo_path = os.path.join(BASE_DIR, 'static', 'Logo_BGTK.png')
    logo_img = None
    logo_width = 0
    logo_height = 0
    if os.path.exists(logo_path):
        try:
            logo_pil = Image.open(logo_path)
            # Resize logo untuk header (max height 0.6 inch)
            max_logo_height = 0.6 * inch
            logo_ratio = logo_pil.width / logo_pil.height
            logo_width = max_logo_height * logo_ratio
            logo_height = max_logo_height

            # Convert PIL Image to reportlab Image
            logo_buffer = io.BytesIO()
            logo_pil.save(logo_buffer, format='PNG')
            logo_buffer.seek(0)
            logo_img = RLImage(logo_buffer, width=logo_width, height=logo_height)
        except Exception as e:
            print(f"Error loading logo: {e}")
            logo_img = None

    # Load logo Pendidikan Bermutu dengan background transparan
    pendidikan_bermutu_path = os.path.join(BASE_DIR, 'static', 'Pendidikan Bermutu untuk Semua.png')
    pendidikan_bermutu_processed_path = None
    pendidikan_bermutu_width = 0
    pendidikan_bermutu_height = 0
    pendidikan_bermutu_temp_file = None
    if os.path.exists(pendidikan_bermutu_path):
        try:
            pendidikan_bermutu_pil = Image.open(pendidikan_bermutu_path)
            if pendidikan_bermutu_pil.mode != 'RGBA':
                pendidikan_bermutu_pil = pendidikan_bermutu_pil.convert('RGBA')

            # Hapus background hitam (buat transparan)
            pixels = pendidikan_bermutu_pil.load()
            width, height = pendidikan_bermutu_pil.size
            black_threshold = 30

            for x in range(width):
                for y in range(height):
                    r, g, b, a = pixels[x, y]
                    if r < black_threshold and g < black_threshold and b < black_threshold:
                        pixels[x, y] = (r, g, b, 0)

            max_logo_height = 0.5 * inch
            pendidikan_bermutu_ratio = pendidikan_bermutu_pil.width / pendidikan_bermutu_pil.height
            pendidikan_bermutu_width = max_logo_height * pendidikan_bermutu_ratio
            pendidikan_bermutu_height = max_logo_height

            pendidikan_bermutu_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            pendidikan_bermutu_pil.save(pendidikan_bermutu_temp_file.name, format='PNG', optimize=False)
            pendidikan_bermutu_processed_path = pendidikan_bermutu_temp_file.name
        except Exception as e:
            print(f"Error loading logo Pendidikan Bermutu: {e}")

    # Load logo Ramah dengan background transparan
    ramah_path = os.path.join(BASE_DIR, 'static', 'Ramah.png')
    ramah_processed_path = None
    ramah_width = 0
    ramah_height = 0
    ramah_temp_file = None
    if os.path.exists(ramah_path):
        try:
            ramah_pil = Image.open(ramah_path)
            if ramah_pil.mode != 'RGBA':
                ramah_pil = ramah_pil.convert('RGBA')

            # Hapus background hitam (buat transparan)
            pixels = ramah_pil.load()
            width, height = ramah_pil.size
            black_threshold = 30

            for x in range(width):
                for y in range(height):
                    r, g, b, a = pixels[x, y]
                    if r < black_threshold and g < black_threshold and b < black_threshold:
                        pixels[x, y] = (r, g, b, 0)

            max_logo_height = 0.5 * inch
            ramah_ratio = ramah_pil.width / ramah_pil.height
            ramah_width = max_logo_height * ramah_ratio
            ramah_height = max_logo_height

            ramah_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            ramah_pil.save(ramah_temp_file.name, format='PNG', optimize=False)
            ramah_processed_path = ramah_temp_file.name
        except Exception as e:
            print(f"Error loading logo Ramah: {e}")

    # Fungsi untuk header dengan logo dan footer
    def add_header_footer(canvas, doc):
        canvas.saveState()

        # Hitung tinggi logo terbesar untuk alignment vertikal (hanya logo BGTK di header)
        max_logo_h = logo_height if logo_height > 0 else 0

        # Header - Logo BGTK di kiri
        if logo_img and logo_height > 0:
            try:
                logo_x = 25
                logo_y = F4_SIZE[1] - 25 - max_logo_h
                canvas.drawImage(logo_path, logo_x, logo_y, width=logo_width, height=logo_height, preserveAspectRatio=True)
            except Exception as e:
                print(f"Error drawing logo: {e}")

        # Garis header
        canvas.setStrokeColor(colors.HexColor('#067ac1'))
        canvas.setLineWidth(1.5)
        max_header_height = logo_height if logo_height > 0 else 0
        header_line_y = F4_SIZE[1] - 25 - max_header_height - 5
        canvas.line(25, header_line_y, F4_SIZE[0] - 25, header_line_y)

        # Footer - Logo Pendidikan Bermutu dan Ramah di kanan bawah
        max_footer_logo_height = max(pendidikan_bermutu_height if pendidikan_bermutu_height > 0 else 0,
                                    ramah_height if ramah_height > 0 else 0)
        footer_logo_y = 45  # Posisi Y untuk logo footer

        if max_footer_logo_height > 0:
            # Hitung total lebar kedua logo + spacing
            total_footer_logo_width = (pendidikan_bermutu_width if pendidikan_bermutu_height > 0 else 0) + \
                                     (ramah_width if ramah_height > 0 else 0) + 10  # 10 pt spacing antara logo
            # Posisi X untuk logo di kanan (rata kanan dengan margin 25)
            footer_logo_start_x = F4_SIZE[0] - 25 - total_footer_logo_width

            # Logo Pendidikan Bermutu di kiri (dalam footer)
            if pendidikan_bermutu_processed_path and os.path.exists(pendidikan_bermutu_processed_path) and pendidikan_bermutu_height > 0:
                try:
                    pendidikan_bermutu_footer_x = footer_logo_start_x
                    pendidikan_bermutu_footer_y = footer_logo_y
                    canvas.drawImage(pendidikan_bermutu_processed_path, pendidikan_bermutu_footer_x, pendidikan_bermutu_footer_y,
                                   width=pendidikan_bermutu_width, height=pendidikan_bermutu_height, preserveAspectRatio=True, mask='auto')
                except Exception as e:
                    print(f"Error drawing logo Pendidikan Bermutu di footer: {e}")

            # Logo Ramah di kanan (dalam footer)
            if ramah_processed_path and os.path.exists(ramah_processed_path) and ramah_height > 0:
                try:
                    ramah_footer_x = footer_logo_start_x + (pendidikan_bermutu_width if pendidikan_bermutu_height > 0 else 0) + 10
                    ramah_footer_y = footer_logo_y
                    canvas.drawImage(ramah_processed_path, ramah_footer_x, ramah_footer_y,
                                   width=ramah_width, height=ramah_height, preserveAspectRatio=True, mask='auto')
                except Exception as e:
                    print(f"Error drawing logo Ramah di footer: {e}")

        # Garis footer (di bawah logo)
        canvas.setStrokeColor(colors.HexColor('#067ac1'))
        canvas.setLineWidth(1)
        footer_line_y = footer_logo_y - 10  # Garis di bawah logo dengan spacing 10
        canvas.line(25, footer_line_y, F4_SIZE[0] - 25, footer_line_y)

        # Footer - Tanggal dan waktu export WITA (di bawah garis)
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.black)
        wita_time = datetime.utcnow() + timedelta(hours=8)
        footer_text = f"Dicetak pada: {wita_time.strftime('%d/%m/%Y %H:%M')} WITA"
        footer_width = canvas.stringWidth(footer_text, 'Helvetica', 8)
        footer_x = 25  # Rata kiri dengan margin 25 pt
        footer_y = footer_line_y - 15  # Tanggal di bawah garis
        canvas.drawString(footer_x, footer_y, footer_text)

        canvas.restoreState()

    # Gunakan kertas F4 (8.27 x 13 inch) dan beri ruang header
    buffer = io.BytesIO()
    max_header_logo_height = logo_height if logo_height > 0 else 0  # Hanya logo BGTK di header
    top_margin_with_logo = 25 + max_header_logo_height + 15
    doc = SimpleDocTemplate(
        buffer,
        pagesize=F4_SIZE,
        rightMargin=25,
        leftMargin=25,
        topMargin=top_margin_with_logo,
        bottomMargin=40
    )

    elements = []
    styles = getSampleStyleSheet()

    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.black,
        spaceAfter=12,
        spaceBefore=0,
        alignment=1,  # Center
        fontName='Helvetica-Bold',
        leading=24
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.black,
        alignment=0  # Left
    )
    value_style = ParagraphStyle(
        'ValueStyle',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.black,
        alignment=0,  # Left
        leftIndent=0,
        rightIndent=0
    )

    # Field mapping untuk label yang lebih readable dan profesional
    field_labels = {
        'nik': 'NIK',
        'nama_lengkap': 'Nama Lengkap',
        'nip_nippk': 'NIP/NIPPK',
        'tempat_lahir': 'Tempat Lahir',
        'tanggal_lahir': 'Tanggal Lahir',
        'jenis_kelamin': 'Jenis Kelamin',
        'agama': 'Agama',
        'pendidikan_terakhir': 'Pendidikan Terakhir',
        'jurusan': 'Jurusan',
        'status_asn': 'Status Kepegawaian',
        'pangkat_golongan': 'Pangkat / Golongan',
        'jabatan': 'Jabatan',
        'instansi': 'Nama Instansi',
        'alamat_instansi': 'Alamat Instansi',
        'alamat_domisili': 'Alamat Domisili',
        'kabupaten_kota': 'Kabupaten/Kota',
        'kabko_lainnya': 'Kabupaten/Kota Lainnya',
        'peran': 'Peran dalam Kegiatan',
        'no_hp': 'Nomor HP',
        'alamat_email': 'Alamat Email',
        'npwp': 'NPWP',
        'nama_bank': 'Nama Bank',
        'nama_bank_lainnya': 'Nama Bank Lainnya',
        'no_rekening': 'Nomor Rekening',
        'nama_pemilik_rekening': 'Nama Pemilik Rekening'
    }

    # Urutan field yang diinginkan
    field_order = [
        'nik', 'nama_lengkap', 'nip_nippk', 'tempat_lahir', 'tanggal_lahir',
        'jenis_kelamin', 'agama', 'pendidikan_terakhir', 'jurusan',
        'status_asn', 'pangkat_golongan', 'jabatan', 'instansi',
        'alamat_instansi', 'alamat_domisili', 'kabupaten_kota', 'kabko_lainnya',
        'peran',
        'no_hp', 'alamat_email', 'npwp',
        'nama_bank', 'nama_bank_lainnya', 'no_rekening', 'nama_pemilik_rekening'
    ]

    # Exclude fields yang tidak perlu ditampilkan
    exclude_fields = ['id', 'user_id', 'buku_tabungan_path', 'tanda_tangan', 'created_at', 'updated_at', 'nama_kegiatan', 'waktu_pelaksanaan', 'tempat_pelaksanaan']

    # List untuk menyimpan temporary files tanda tangan untuk cleanup
    tanda_tangan_temp_files = []

    # Loop untuk setiap user - buat 1 halaman per user
    for user_idx, biodata in enumerate(all_biodata):
        # Tambahkan page break kecuali untuk user pertama
        if user_idx > 0:
            elements.append(PageBreak())

        # Title
        title = Paragraph("BIODATA KEGIATAN", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.15*inch))

        # Ambil data kegiatan untuk header tabel
        biodata_nama_kegiatan = biodata.get('nama_kegiatan', '-')
        waktu_pelaksanaan = biodata.get('waktu_pelaksanaan', '-')
        tempat_pelaksanaan = biodata.get('tempat_pelaksanaan', '-')

        # Buat header informasi kegiatan di atas tabel dengan titik dua yang sejajar
        if biodata_nama_kegiatan or waktu_pelaksanaan or tempat_pelaksanaan:
            info_style = ParagraphStyle(
                'InfoStyle',
                parent=styles['Normal'],
                fontSize=10,
                leading=14,
                textColor=colors.black,
                alignment=0,  # Left align
                fontName='Helvetica-Bold'
            )

            # Buat tabel kecil untuk info kegiatan agar titik dua sejajar
            info_table_data = []
            if biodata_nama_kegiatan and str(biodata_nama_kegiatan).strip() and str(biodata_nama_kegiatan).strip() != '-':
                info_table_data.append([
                    Paragraph("<b>Nama Kegiatan</b>", info_style),
                    Paragraph(":", info_style),
                    Paragraph(str(biodata_nama_kegiatan), info_style)
                ])
            if waktu_pelaksanaan and str(waktu_pelaksanaan).strip() and str(waktu_pelaksanaan).strip() != '-':
                info_table_data.append([
                    Paragraph("<b>Waktu Pelaksanaan</b>", info_style),
                    Paragraph(":", info_style),
                    Paragraph(str(waktu_pelaksanaan), info_style)
                ])
            if tempat_pelaksanaan and str(tempat_pelaksanaan).strip() and str(tempat_pelaksanaan).strip() != '-':
                info_table_data.append([
                    Paragraph("<b>Tempat Pelaksanaan</b>", info_style),
                    Paragraph(":", info_style),
                    Paragraph(str(tempat_pelaksanaan), info_style)
                ])

            if info_table_data:
                available_width = (8.27 * inch) - (25 * 2)
                info_table = Table(info_table_data, colWidths=[2.2*inch, 0.15*inch, available_width - 2.2*inch - 0.15*inch])
                info_table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('ALIGN', (2, 0), (2, -1), 'LEFT'),
                    ('LEFTPADDING', (0, 0), (0, -1), 0),
                    ('LEFTPADDING', (1, 0), (1, -1), 0),
                    ('LEFTPADDING', (2, 0), (2, -1), 0),
                    ('RIGHTPADDING', (0, 0), (0, -1), 4),
                    ('RIGHTPADDING', (1, 0), (1, -1), 0),
                    ('RIGHTPADDING', (2, 0), (2, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ]))
                elements.append(info_table)
                elements.append(Spacer(1, 0.1*inch))

        # Semua data dalam 1 tabel
        all_data = []

        # Collect all data in order
        for key in field_order:
            if key in biodata and key not in exclude_fields:
                label = field_labels.get(key, key.replace('_', ' ').title())
                value = biodata[key]
                display_value = str(value) if value and str(value).strip() else '-'
                all_data.append([label, display_value])

        # Add any remaining fields not in field_order
        for key, value in biodata.items():
            if key not in exclude_fields and key not in field_order:
                label = field_labels.get(key, key.replace('_', ' ').title())
                display_value = str(value) if value and str(value).strip() else '-'
                all_data.append([label, display_value])

        # Process tanda tangan menggunakan helper function
        tanda_tangan_raw = biodata.get('tanda_tangan')
        tanda_tangan_img, error_msg = process_tanda_tangan_for_pdf(tanda_tangan_raw, tanda_tangan_temp_files)

        # Buat 1 tabel untuk semua data
        if all_data:
            table_data = []
            for row in all_data:
                field_label = Paragraph(f"<b>{row[0]}</b>", value_style)
                value_text = str(row[1]) if row[1] else '-'
                value_text = value_text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                if len(value_text) > 50:
                    words = value_text.split()
                    lines = []
                    current_line = []
                    current_length = 0
                    for word in words:
                        if current_length + len(word) + 1 > 50:
                            if current_line:
                                lines.append(' '.join(current_line))
                            current_line = [word]
                            current_length = len(word)
                        else:
                            current_line.append(word)
                            current_length += len(word) + 1
                    if current_line:
                        lines.append(' '.join(current_line))
                    value_text = '<br/>'.join(lines)
                value_para = Paragraph(value_text, value_style)
                table_data.append([field_label, value_para])

            # Tambahkan tanda tangan ke tabel jika ada
            if tanda_tangan_img:
                field_label = Paragraph("<b>Tanda Tangan</b>", value_style)
                table_data.append([field_label, tanda_tangan_img])
            elif biodata.get('tanda_tangan'):
                field_label = Paragraph("<b>Tanda Tangan</b>", value_style)
                error_msg_para = Paragraph("<i>Tanda tangan tidak tersedia</i>", value_style)
                table_data.append([field_label, error_msg_para])

            available_width = (8.27 * inch) - (25 * 2)
            table = Table(table_data, colWidths=[2.2*inch, available_width - 2.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (0, -1), 4),
                ('LEFTPADDING', (1, 0), (1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
                ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#067ac1')),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.08*inch))

    # Build PDF
    try:
        doc.build(elements, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
        buffer.seek(0)

        # Refresh session sebelum return response untuk mencegah logout
        session.permanent = True
        session.modified = True
    finally:
        # Cleanup temporary files
        try:
            if pendidikan_bermutu_temp_file and os.path.exists(pendidikan_bermutu_processed_path):
                os.unlink(pendidikan_bermutu_processed_path)
        except Exception as e:
            print(f"Error cleaning up Pendidikan Bermutu temp file: {e}")

        try:
            if ramah_temp_file and os.path.exists(ramah_processed_path):
                os.unlink(ramah_processed_path)
        except Exception as e:
            print(f"Error cleaning up Ramah temp file: {e}")

        # Cleanup temporary files tanda tangan
        for temp_file_path in tanda_tangan_temp_files:
            try:
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)
            except Exception as e:
                print(f"Error cleaning up tanda tangan temp file {temp_file_path}: {e}")

    # Generate filename
    filename = f"Rekap_Tahun_{selected_year}"
    if bulan_awal and bulan_akhir:
        if bulan_awal == bulan_akhir:
            filename += f"_{bulan_names[bulan_awal]}"
        else:
            filename += f"_{bulan_names[bulan_awal]}_{bulan_names[bulan_akhir]}"
    filename += ".pdf"

    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
    )

@app.route('/admin/export-rekap-tahunan-excel')
@admin_required
def export_rekap_tahunan_excel():
    """Export rekap tahunan ke Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    user_role = get_user_role()
    user_id = get_user_id()

    # Ambil parameter tahun dari URL
    selected_year = request.args.get('tahun', '').strip()
    if selected_year:
        try:
            selected_year = int(selected_year)
        except ValueError:
            selected_year = datetime.now().year
    else:
        selected_year = datetime.now().year

    # Ambil parameter bulan awal dan bulan akhir
    bulan_awal = request.args.get('bulan_awal', '').strip()
    bulan_akhir = request.args.get('bulan_akhir', '').strip()

    if bulan_awal:
        try:
            bulan_awal = int(bulan_awal)
            if bulan_awal < 1 or bulan_awal > 12:
                bulan_awal = None
        except ValueError:
            bulan_awal = None
    else:
        bulan_awal = None

    if bulan_akhir:
        try:
            bulan_akhir = int(bulan_akhir)
            if bulan_akhir < 1 or bulan_akhir > 12:
                bulan_akhir = None
        except ValueError:
            bulan_akhir = None

    # Jika hanya satu bulan yang dipilih, set keduanya sama
    if bulan_awal and not bulan_akhir:
        bulan_akhir = bulan_awal
    elif bulan_akhir and not bulan_awal:
        bulan_awal = bulan_akhir

    # Pastikan bulan_awal <= bulan_akhir
    if bulan_awal and bulan_akhir and bulan_awal > bulan_akhir:
        bulan_awal, bulan_akhir = bulan_akhir, bulan_awal

    # Nama bulan
    bulan_names = {
        1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
        5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
        9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
    }

    # Ambil data dari database
    connection = get_db_connection()
    if not connection:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_rekap_tahunan'))

    # Ambil daftar kegiatan per tahun
    kegiatan_list = []
    try:
        cursor = connection.cursor()

        # Buat kondisi WHERE untuk filter bulan (SQLite)
        selected_year_str = str(selected_year)
        where_conditions = ["CAST(strftime('%Y', bk.created_at) AS INTEGER) = ?"]
        params = [selected_year_str]

        if bulan_awal and bulan_akhir:
            where_conditions.append("CAST(strftime('%m', bk.created_at) AS INTEGER) BETWEEN ? AND ?")
            params.extend([bulan_awal, bulan_akhir])

        # Jika operator, batasi hanya pada kegiatan yang ia pegang
        if user_role == 'operator' and user_id:
            where_conditions.append("""
                EXISTS (
                    SELECT 1
                    FROM kegiatan_master k
                    INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                    WHERE ok.user_id = ?
                      AND TRIM(k.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                )
            """)
            params.append(user_id)

        where_clause = " AND ".join(where_conditions)

        # Ambil daftar kegiatan per tahun dengan detail
        kegiatan_query = f"""
            SELECT
                bk.nama_kegiatan,
                COUNT(DISTINCT bk.id) as jumlah_peserta,
                COUNT(DISTINCT bk.kabupaten_kota) as jumlah_kabupaten,
                MIN(bk.created_at) as tanggal_awal,
                MAX(bk.created_at) as tanggal_akhir
            FROM biodata_kegiatan bk
            WHERE {where_clause}
                AND TRIM(bk.nama_kegiatan) != ''
            GROUP BY bk.nama_kegiatan
            ORDER BY jumlah_peserta DESC, bk.nama_kegiatan ASC
        """

        cursor.execute(kegiatan_query, tuple(params))
        rows = cursor.fetchall()
        kegiatan_list = [row_to_dict(row) for row in rows]

        # Ambil detail per kabupaten untuk setiap kegiatan dan semua biodata
        kegiatan_data = []
        all_biodata = []

        for kegiatan in kegiatan_list:
            nama_kegiatan = kegiatan.get('nama_kegiatan')

            # Ambil detail per kabupaten untuk kegiatan ini
            kabupaten_detail_query = f"""
                SELECT
                    bk.kabupaten_kota,
                    COUNT(DISTINCT bk.id) as jumlah_peserta,
                    MIN(bk.created_at) as tanggal_awal,
                    MAX(bk.created_at) as tanggal_akhir
                FROM biodata_kegiatan bk
                WHERE {where_clause}
                    AND TRIM(bk.nama_kegiatan) = TRIM(?)
                    AND TRIM(bk.kabupaten_kota) != ''
                    AND bk.kabupaten_kota IS NOT NULL
                GROUP BY bk.kabupaten_kota
                ORDER BY bk.kabupaten_kota ASC
            """
            kabupaten_params = list(params) + [nama_kegiatan]
            cursor.execute(kabupaten_detail_query, tuple(kabupaten_params))
            kabupaten_rows = cursor.fetchall()

            # Convert kabupaten rows to dict
            kabupaten_list = []
            for kab_row in kabupaten_rows:
                kab_dict = row_to_dict(kab_row)
                kabupaten_list.append(kab_dict)

            kegiatan['kabupaten_detail'] = kabupaten_list

            # Ambil semua biodata untuk kegiatan ini
            biodata_query = f"""
                SELECT * FROM biodata_kegiatan
                WHERE TRIM(nama_kegiatan) = TRIM(?)
                    AND CAST(strftime('%Y', created_at) AS INTEGER) = ?
            """
            biodata_params = [nama_kegiatan, selected_year_str]

            if bulan_awal and bulan_akhir:
                biodata_query += " AND CAST(strftime('%m', created_at) AS INTEGER) BETWEEN ? AND ?"
                biodata_params.extend([bulan_awal, bulan_akhir])

            biodata_query += " ORDER BY kabupaten_kota ASC, nama_lengkap ASC"

            cursor.execute(biodata_query, tuple(biodata_params))
            biodata_rows = cursor.fetchall()
            biodata_list = [row_to_dict(row) for row in biodata_rows]

            # Normalisasi path buku tabungan
            for biodata in biodata_list:
                if biodata.get('buku_tabungan_path'):
                    biodata['buku_tabungan_path'] = normalize_buku_tabungan_path(biodata['buku_tabungan_path'])

            all_biodata.extend(biodata_list)

        # Sort by nama_kegiatan, then by kabupaten_kota, then by nama_lengkap
        all_biodata.sort(key=lambda x: (
            x.get('nama_kegiatan', ''),
            x.get('kabupaten_kota', ''),
            x.get('nama_lengkap', '')
        ))

    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan saat mengambil data: {str(e)}', 'error')
        return redirect(url_for('admin_rekap_tahunan'))
    finally:
        if connection:
            cursor.close()
            connection.close()

    if not all_biodata:
        flash('Tidak ada data untuk diekspor!', 'error')
        return redirect(url_for('admin_rekap_tahunan'))

    # Field mapping untuk label yang lebih readable dan profesional
    field_labels = {
        'nik': 'NIK',
        'nama_lengkap': 'Nama Lengkap',
        'nip_nippk': 'NIP/NIPPK',
        'tempat_lahir': 'Tempat Lahir',
        'tanggal_lahir': 'Tanggal Lahir',
        'jenis_kelamin': 'Jenis Kelamin',
        'agama': 'Agama',
        'pendidikan_terakhir': 'Pendidikan Terakhir',
        'jurusan': 'Jurusan',
        'status_asn': 'Status Kepegawaian',
        'pangkat_golongan': 'Pangkat / Golongan',
        'jabatan': 'Jabatan',
        'instansi': 'Nama Instansi',
        'alamat_instansi': 'Alamat Instansi',
        'alamat_domisili': 'Alamat Domisili',
        'kabupaten_kota': 'Kabupaten/Kota',
        'kabko_lainnya': 'Kabupaten/Kota Lainnya',
        'peran': 'Peran dalam Kegiatan',
        'no_hp': 'Nomor HP',
        'alamat_email': 'Alamat Email',
        'npwp': 'NPWP',
        'nama_bank': 'Nama Bank',
        'nama_bank_lainnya': 'Nama Bank Lainnya',
        'no_rekening': 'Nomor Rekening',
        'nama_pemilik_rekening': 'Nama Pemilik Rekening'
    }

    # Urutan field yang diinginkan
    field_order = [
        'nik', 'nama_lengkap', 'nip_nippk', 'tempat_lahir', 'tanggal_lahir',
        'jenis_kelamin', 'agama', 'pendidikan_terakhir', 'jurusan',
        'status_asn', 'pangkat_golongan', 'jabatan', 'instansi',
        'alamat_instansi', 'alamat_domisili', 'kabupaten_kota', 'kabko_lainnya',
        'peran',
        'no_hp', 'alamat_email', 'npwp',
        'nama_bank', 'nama_bank_lainnya', 'no_rekening', 'nama_pemilik_rekening'
    ]

    # Exclude fields yang tidak perlu ditampilkan
    exclude_fields = ['id', 'user_id', 'buku_tabungan_path', 'tanda_tangan', 'created_at', 'updated_at', 'waktu_pelaksanaan', 'tempat_pelaksanaan']

    # Buat workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Biodata"

    # Styles
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_font = Font(bold=True, color="000000", size=11)
    title_font = Font(bold=True, size=12, color="000000")
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Buat list semua field yang akan ditampilkan (dalam urutan yang diinginkan)
    all_fields = []
    # Tambahkan nama_kegiatan di awal
    all_fields.append({'key': 'nama_kegiatan', 'label': 'Nama Kegiatan'})
    for key in field_order:
        if key not in exclude_fields:
            label = field_labels.get(key, key.replace('_', ' ').title())
            all_fields.append({'key': key, 'label': label})

    # Tambahkan field tanda tangan
    all_fields.append({'key': 'tanda_tangan', 'label': 'Tanda Tangan'})

    # Header kegiatan
    num_cols = len(all_fields)
    last_col_letter = get_column_letter(num_cols)

    # Title
    title_text = f"Rekap Kegiatan Tahun {selected_year}"
    if bulan_awal and bulan_akhir:
        if bulan_awal == bulan_akhir:
            title_text += f" - {bulan_names[bulan_awal]}"
        else:
            title_text += f" - {bulan_names[bulan_awal]} s/d {bulan_names[bulan_akhir]}"

    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = title_text
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    # Tambahkan informasi tanggal export
    export_date = datetime.now().strftime("%d %B %Y, %H:%M:%S")
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = f"Tanggal Export: {export_date}"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = left_align
    ws.row_dimensions[2].height = 18

    # Buat header kolom (semua field sebagai kolom)
    for col_idx, field_info in enumerate(all_fields, 1):
        cell = ws.cell(row=3, column=col_idx, value=field_info['label'])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_style
    ws.row_dimensions[3].height = 22

    # Set lebar kolom pertama (Nama Kegiatan) lebih lebar
    ws.column_dimensions['A'].width = 35
    # Set lebar semua kolom lainnya
    for col_idx in range(2, len(all_fields) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 25

    # Data untuk setiap user (setiap user = 1 baris)
    current_row = 4
    for user_idx, biodata in enumerate(all_biodata):
        for col_idx, field_info in enumerate(all_fields, 1):
            field_key = field_info['key']

            # Ambil nilai field
            if field_key == 'tanda_tangan':
                display_value = "[Tersedia]" if biodata.get('tanda_tangan') else "-"
            else:
                value = biodata.get(field_key, '')
                display_value = str(value) if value and str(value).strip() else '-'

            # Tulis ke cell
            cell = ws.cell(row=current_row, column=col_idx, value=display_value)
            cell.border = border_style

            # Alternating row colors untuk readability
            if user_idx % 2 == 0:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # Alignment dengan wrap text untuk semua kolom
            if field_key in ['nik', 'nip_nippk', 'no_hp', 'npwp', 'no_rekening']:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            elif field_key in ['tanggal_lahir']:
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Auto adjust row height
        max_height = 18
        for col_idx, field_info in enumerate(all_fields, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            if cell.value:
                col_letter = get_column_letter(col_idx)
                col_width = ws.column_dimensions[col_letter].width or 25
                text_length = len(str(cell.value))
                estimated_lines = max(1, (text_length / max(col_width * 0.8, 1)))
                estimated_height = estimated_lines * 15
                max_height = max(max_height, min(estimated_height, 60))

        ws.row_dimensions[current_row].height = max_height
        current_row += 1

    # Auto-adjust column widths berdasarkan konten aktual
    for col_idx, field_info in enumerate(all_fields, 1):
        col_letter = get_column_letter(col_idx)

        if col_letter == 'A':
            continue  # Skip kolom A (Nama Kegiatan) - sudah di-set

        max_length = 0
        header_cell = ws.cell(row=3, column=col_idx)
        if header_cell.value:
            max_length = len(str(header_cell.value))

        for row in range(4, current_row):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value:
                text = str(cell.value)
                text_length = len(text)
                if text_length > 50:
                    lines = text.split('\n') if '\n' in text else [text]
                    max_line_length = max(len(line) for line in lines)
                    max_length = max(max_length, min(max_line_length, 40))
                else:
                    max_length = max(max_length, text_length)

        optimal_width = max(12, min(max_length + 2, 40))
        if 'Alamat' in field_info['label']:
            optimal_width = max(optimal_width, min(max_length + 2, 45))
        ws.column_dimensions[col_letter].width = optimal_width

    # Freeze panes untuk memudahkan scrolling (freeze header)
    ws.freeze_panes = 'A4'

    # Enable auto filter untuk header
    ws.auto_filter.ref = f'A3:{last_col_letter}{current_row - 1}'

    # Set print settings
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # Set row heights untuk header
    ws.row_dimensions[1].height = 25

    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Refresh session
    session.permanent = True
    session.modified = True

    # Generate filename
    filename = f"Rekap_Tahun_{selected_year}"
    if bulan_awal and bulan_akhir:
        if bulan_awal == bulan_akhir:
            filename += f"_{bulan_names[bulan_awal]}"
        else:
            filename += f"_{bulan_names[bulan_awal]}_{bulan_names[bulan_akhir]}"
    filename += ".xlsx"

    return Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
    )

@app.route('/admin/kegiatan/<path:nama_kegiatan>')
@admin_required
def admin_detail_kegiatan(nama_kegiatan):
    """Halaman detail peserta per kegiatan berdasarkan kabupaten/kota"""
    from urllib.parse import unquote

    # Decode URL encoding
    nama_kegiatan = unquote(nama_kegiatan)
    kabupaten_data = {}
    user_id = get_user_id()
    user_role = get_user_role()

    connection = get_db_connection()
    if connection is None:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_kegiatan'))

    try:
        cursor = connection.cursor()

        # Jika operator, cek apakah kegiatan ini termasuk yang dia ikuti
        if user_role == 'operator':
            cursor.execute("""
                SELECT k.id
                FROM kegiatan_master k
                INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                WHERE ok.user_id = ?
                    AND TRIM(k.nama_kegiatan) = TRIM(?)
            """, (user_id, nama_kegiatan))
            if not cursor.fetchone():
                flash('Anda tidak memiliki akses ke kegiatan ini!', 'error')
                return redirect(url_for('admin_kegiatan'))

        # Ambil semua biodata untuk kegiatan yang dipilih
        cursor.execute("""
            SELECT nik, nama_lengkap, nip_nippk, instansi, peran, kabupaten_kota, nama_kegiatan
            FROM biodata_kegiatan
            WHERE TRIM(nama_kegiatan) = TRIM(?)
            ORDER BY kabupaten_kota ASC, nama_lengkap ASC
        """, (nama_kegiatan,))
        rows = cursor.fetchall()
        all_biodata = [row_to_dict(row) for row in rows]

        # Group by kabupaten_kota
        for biodata in all_biodata:
            kabupaten = biodata['kabupaten_kota'] or 'Tidak Diketahui'
            if kabupaten not in kabupaten_data:
                kabupaten_data[kabupaten] = []
            kabupaten_data[kabupaten].append(biodata)

    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan saat mengambil data: {str(e)}', 'error')
    finally:
        if connection:
            cursor.close()
            connection.close()

    return render_template(
        'admin/admin-detail-kegiatan.html',
        kabupaten_data=kabupaten_data,
        selected_kegiatan=nama_kegiatan,
        username=get_username()
    )

@app.route('/api/kabupaten-summary', methods=['GET'])
@admin_required
def api_kabupaten_summary():
    """API untuk popup rekap per kabupaten di dashboard. Operator hanya melihat data kegiatan yang dipegang."""
    all_kabupaten_list = [
        'BANGGAI', 'BANGGAI KEPULAUAN', 'BANGGAI LAUT', 'BUOL', 'DONGGALA',
        'MOROWALI', 'MOROWALI UTARA', 'PALU', 'PARIGI MOUTONG', 'POSO',
        'SIGI', 'TOJO UNA-UNA', 'TOLI-TOLI'
    ]
    all_kabupaten_upper = {k.upper().strip() for k in all_kabupaten_list}
    kabupaten_summary = []
    connection = get_db_connection()
    if not connection:
        return jsonify(kabupaten_summary)
    cursor = None
    try:
        cursor = connection.cursor()
        user_role = get_user_role()
        user_id = get_user_id()

        if user_role == 'operator' and user_id:
            cursor.execute("""
                SELECT bk.kabupaten_kota, COUNT(*) as jumlah_peserta
                FROM biodata_kegiatan bk
                INNER JOIN kegiatan_master k ON TRIM(k.nama_kegiatan) = TRIM(bk.nama_kegiatan)
                INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                WHERE ok.user_id = ? AND TRIM(bk.kabupaten_kota) != '' AND bk.kabupaten_kota IS NOT NULL
                GROUP BY bk.kabupaten_kota
            """, (user_id,))
        else:
            cursor.execute("""
                SELECT kabupaten_kota, COUNT(*) as jumlah_peserta
                FROM biodata_kegiatan
                WHERE TRIM(kabupaten_kota) != '' AND kabupaten_kota IS NOT NULL
                GROUP BY kabupaten_kota
            """)
        counts = {row['kabupaten_kota']: row['jumlah_peserta'] for row in cursor.fetchall()}
        lainnya_count = sum(c for k, c in counts.items() if (k or '').strip().upper() not in all_kabupaten_upper)
        kabupaten_summary = [{'nama': k, 'jumlah_peserta': counts.get(k, 0)} for k in all_kabupaten_list]
        kabupaten_summary.append({'nama': 'LAINNYA', 'jumlah_peserta': lainnya_count})
        # Pastikan "LAINNYA" selalu paling terakhir
        kabupaten_summary.sort(key=lambda x: ((x.get('nama') or '').strip().upper() == 'LAINNYA', x.get('nama') or ''))
    except sqlite3.Error:
        pass
    finally:
        if cursor:
            cursor.close()
        connection.close()
    return jsonify(kabupaten_summary)


@app.route('/api/get-kegiatan/<path:nama_kegiatan>', methods=['GET'])
def get_kegiatan_by_nama(nama_kegiatan):
    """API endpoint untuk mendapatkan data kegiatan berdasarkan nama kegiatan (public, tidak perlu login)"""
    from urllib.parse import unquote

    # Decode URL encoding
    nama_kegiatan = unquote(nama_kegiatan)

    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Koneksi database gagal!'}), 500

        try:
            cursor = connection.cursor()
            cursor.execute("""
                SELECT nama_kegiatan, waktu_pelaksanaan, tempat_pelaksanaan
                FROM kegiatan_master
                WHERE TRIM(nama_kegiatan) = TRIM(?)
                    AND (is_hidden IS NULL OR is_hidden = 0)
                LIMIT 1
            """, (nama_kegiatan,))
            kegiatan_row = cursor.fetchone()

            if kegiatan_row:
                kegiatan = row_to_dict(kegiatan_row)
                return jsonify({
                    'success': True,
                    'data': {
                        'nama_kegiatan': kegiatan.get('nama_kegiatan'),
                        'waktu_pelaksanaan': kegiatan.get('waktu_pelaksanaan'),
                        'tempat_pelaksanaan': kegiatan.get('tempat_pelaksanaan')
                    }
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'Kegiatan tidak ditemukan'
                })
        except sqlite3.Error as e:
            return jsonify({'error': f'Terjadi kesalahan: {str(e)}'}), 500
        finally:
            if connection is not None:
                cursor.close()
                connection.close()
    except Exception as e:
        return jsonify({'error': f'Terjadi kesalahan tidak terduga: {str(e)}'}), 500

@app.route('/check-nik', methods=['POST'])
@csrf.exempt
def check_nik():
    """API endpoint untuk mengecek apakah NIK sudah terdaftar (public API untuk auto-fill)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'available': False, 'message': 'Data tidak valid'}), 400

        nik = data.get('nik', '').strip()
        if not nik:
            return jsonify({'available': False, 'message': 'NIK tidak boleh kosong'}), 400

        # Validasi NIK harus tepat 16 digit
        if not nik.isdigit() or len(nik) != 16:
            return jsonify({'available': False, 'message': 'NIK harus tepat 16 digit angka'}), 400

        # Public API, user_id bisa None
        user_id = session.get('user_id')
        connection = get_db_connection()

        if not connection:
            return jsonify({'available': False, 'message': 'Koneksi database gagal!'}), 500

        try:
            cursor = connection.cursor()

            # Cek apakah user sudah punya data (jika ada user_id)
            if user_id:
                cursor.execute("SELECT nik FROM biodata_kegiatan WHERE user_id = ?", (user_id,))
                existing = cursor.fetchone()

                if existing and nik == existing[0]:
                    return jsonify({'available': True, 'message': 'NIK valid'}), 200

            # NIK bisa digunakan untuk beberapa kegiatan berbeda, jadi tidak perlu memblokir
            # Hanya validasi format NIK (16 digit), bukan keberadaannya
            return jsonify({'available': True, 'message': 'NIK valid - dapat digunakan untuk kegiatan baru'}), 200

        except sqlite3.Error as e:
            print(f"Error checking NIK: {e}")
            return jsonify({'available': False, 'message': 'Terjadi kesalahan saat mengecek NIK'}), 500
        finally:
            if connection is not None:
                cursor.close()
                connection.close()

    except Exception as e:
        print(f"Unexpected error in check_nik: {e}")
        return jsonify({'available': False, 'message': 'Terjadi kesalahan tidak terduga'}), 500

@app.route('/api/get-latest-by-nik', methods=['POST'])
@csrf.exempt
def get_latest_by_nik():
    """API endpoint untuk mengambil data biodata terakhir berdasarkan NIK (tidak perlu login)"""
    try:
        # Log request info untuk debugging
        print(f"🔍 Debug get_latest_by_nik - Content-Type: {request.content_type}")
        print(f"🔍 Debug get_latest_by_nik - Request method: {request.method}")
        print(f"🔍 Debug get_latest_by_nik - Is JSON: {request.is_json}")
        print(f"🔍 Debug get_latest_by_nik - Raw data: {request.data}")

        # Coba ambil JSON data dengan force=True untuk bypass Content-Type check
        data = None
        try:
            data = request.get_json(force=True)
        except Exception as e:
            print(f"❌ Error parsing JSON with force=True: {e}")
            # Coba tanpa force
            try:
                data = request.get_json()
            except Exception as e2:
                print(f"❌ Error parsing JSON: {e2}")
                return jsonify({'success': False, 'message': f'Error parsing JSON: {str(e2)}'}), 400

        print(f"🔍 Debug get_latest_by_nik - Data received: {data}, Type: {type(data)}")

        if not data:
            print(f"⚠️ Data kosong atau None")
            # Coba ambil dari form data sebagai fallback
            nik_from_form = request.form.get('nik')
            if nik_from_form:
                print(f"🔍 Found NIK in form data: {nik_from_form}")
                data = {'nik': nik_from_form}
            else:
                return jsonify({'success': False, 'message': 'Data tidak valid - pastikan mengirim JSON dengan field "nik"'}), 400

        # Ambil NIK, handle jika berupa number atau string
        nik_value = data.get('nik')
        if nik_value is None:
            return jsonify({'success': False, 'message': 'NIK tidak ditemukan dalam request. Pastikan request body berisi {"nik": "1234567890123456"}'}), 400

        # Convert ke string jika number
        if isinstance(nik_value, (int, float)):
            nik = str(int(nik_value)).strip()
        else:
            nik = str(nik_value).strip()

        print(f"🔍 Debug get_latest_by_nik - NIK extracted: '{nik}', length: {len(nik) if nik else 0}, type: {type(nik)}")

        if not nik:
            return jsonify({'success': False, 'message': 'NIK tidak boleh kosong'}), 400

        # Validasi NIK harus tepat 16 digit
        if not nik.isdigit() or len(nik) != 16:
            return jsonify({'success': False, 'message': f'NIK harus tepat 16 digit angka. Diterima: {len(nik)} digit'}), 400

        connection = get_db_connection()

        if not connection:
            return jsonify({'success': False, 'message': 'Koneksi database gagal!'}), 500

        try:
            cursor = connection.cursor()

            # Ambil data terakhir berdasarkan NIK (urutan berdasarkan created_at DESC)
            cursor.execute("""
                SELECT * FROM biodata_kegiatan
                WHERE nik = ?
                ORDER BY created_at DESC
                LIMIT 1
            """, (nik,))

            biodata_row = cursor.fetchone()

            if biodata_row:
                biodata = row_to_dict(biodata_row)
                # Log untuk debugging - lihat apa yang dikembalikan dari database
                print(f"Debug get_latest_by_nik - biodata dari DB: keys={list(biodata.keys())}")
                print(f"Debug get_latest_by_nik - tanggal_lahir dari DB: {biodata.get('tanggal_lahir')}, type={type(biodata.get('tanggal_lahir'))}")

                # Konversi biodata dict ke format yang bisa di-serialize ke JSON
                # Convert datetime objects ke string, tapi tetap simpan value asli untuk field penting
                biodata_dict = {}
                for key, value in biodata.items():
                    if value is None:
                        biodata_dict[key] = None
                    elif isinstance(value, datetime):
                        # Format tanggal lahir sebagai date, created_at sebagai datetime
                        if 'tanggal_lahir' in key.lower():
                            biodata_dict[key] = value.strftime('%Y-%m-%d')
                        elif 'created_at' in key.lower() or 'updated_at' in key.lower():
                            biodata_dict[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                        else:
                            biodata_dict[key] = value.strftime('%Y-%m-%d') if 'date' in key.lower() else value.strftime('%Y-%m-%d %H:%M:%S')
                    elif isinstance(value, (bytes, bytearray)):
                        biodata_dict[key] = value.decode('utf-8', errors='ignore')
                    elif isinstance(value, (int, float)):
                        biodata_dict[key] = value
                    elif isinstance(value, bool):
                        biodata_dict[key] = value
                    else:
                        # Tangani date object (dari MySQL DATE type bisa jadi date bukan datetime)
                        # Juga tangani string yang sudah dalam format date
                        if 'tanggal_lahir' in key.lower():
                            # Pastikan tanggal_lahir selalu dalam format YYYY-MM-DD
                            if hasattr(value, 'strftime'):
                                # Date/datetime object
                                biodata_dict[key] = value.strftime('%Y-%m-%d')
                            elif isinstance(value, str):
                                # String - coba parse dan format ulang jika perlu
                                if len(value) >= 10:
                                    # Ambil 10 karakter pertama (YYYY-MM-DD)
                                    biodata_dict[key] = value[:10]
                                else:
                                    biodata_dict[key] = value
                            else:
                                biodata_dict[key] = str(value)
                        else:
                            # String dan tipe lainnya tetap sebagai string
                            biodata_dict[key] = value

                # Normalisasi path buku tabungan dan buat URL lengkap
                if biodata_dict.get('buku_tabungan_path'):
                    normalized_path = normalize_buku_tabungan_path(biodata_dict['buku_tabungan_path'])
                    # Buat URL untuk akses gambar (relative path dari static)
                    if normalized_path.startswith('static/'):
                        biodata_dict['buku_tabungan_path'] = normalized_path
                    elif normalized_path.startswith('uploads/'):
                        biodata_dict['buku_tabungan_path'] = 'static/' + normalized_path
                    else:
                        biodata_dict['buku_tabungan_path'] = 'static/uploads/' + normalized_path

                # Normalisasi path tanda tangan jika ada dan buat URL lengkap
                if biodata_dict.get('tanda_tangan'):
                    tanda_tangan_val = str(biodata_dict['tanda_tangan'])
                    # Jika sudah base64, biarkan (untuk kompatibilitas)
                    if tanda_tangan_val.startswith('data:image'):
                        biodata_dict['tanda_tangan'] = tanda_tangan_val
                    else:
                        # Normalisasi path dan buat URL
                        if 'uploads/' in tanda_tangan_val or tanda_tangan_val.startswith('static/'):
                            normalized_ttd = normalize_buku_tabungan_path(tanda_tangan_val)
                        else:
                            normalized_ttd = tanda_tangan_val

                        if normalized_ttd.startswith('static/'):
                            biodata_dict['tanda_tangan'] = normalized_ttd
                        elif normalized_ttd.startswith('uploads/'):
                            biodata_dict['tanda_tangan'] = 'static/' + normalized_ttd
                        else:
                            biodata_dict['tanda_tangan'] = 'static/uploads/' + normalized_ttd

                # Log untuk debugging tanggal_lahir
                print(f"Debug get_latest_by_nik - tanggal_lahir dalam biodata_dict: {biodata_dict.get('tanggal_lahir')}")
                print(f"Debug get_latest_by_nik - Keys dalam biodata_dict: {list(biodata_dict.keys())}")

                return jsonify({
                    'success': True,
                    'data': biodata_dict
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'Tidak ada data ditemukan untuk NIK tersebut'
                })

        except sqlite3.Error as e:
            import traceback
            print(f"❌ Error fetching biodata by NIK: {e}")
            traceback.print_exc()
            return jsonify({'success': False, 'message': f'Terjadi kesalahan saat mengambil data: {str(e)}'}), 500
        finally:
            if connection and connection is not None:
                cursor.close()
                connection.close()

    except Exception as e:
        import traceback
        print(f"❌ Unexpected error in get_latest_by_nik: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Terjadi kesalahan tidak terduga: {str(e)}'}), 500

@app.route('/check-kegiatan', methods=['POST'])
@csrf.exempt
def check_kegiatan():
    """API endpoint untuk mengecek apakah user sudah memiliki data untuk kegiatan tertentu (public API untuk form tambah/edit)"""

    data = request.get_json()
    nama_kegiatan = data.get('nama_kegiatan', '').strip()
    exclude_kegiatan = data.get('exclude_kegiatan', '').strip()  # Untuk exclude kegiatan yang sedang diedit
    user_id = get_user_id()

    if not nama_kegiatan:
        return jsonify({'has_data': False, 'message': 'Nama kegiatan tidak boleh kosong'})

    # Jika tidak ada user_id, return False (belum ada data)
    if not user_id:
        return jsonify({'has_data': False, 'message': 'Belum ada data'})

    connection = get_db_connection()
    if not connection:
        return jsonify({'has_data': False, 'message': 'Koneksi database gagal'})

    cursor = None
    try:
        cursor = connection.cursor()
        # Cek apah peserta sudah memiliki data untuk kegiatan ini
        # Jika exclude_kegiatan diberikan, exclude kegiatan tersebut dari pengecekan
        if exclude_kegiatan:
            cursor.execute("""
                SELECT id FROM biodata_kegiatan
                WHERE user_id = ? AND TRIM(nama_kegiatan) = TRIM(?) AND TRIM(nama_kegiatan) != TRIM(?)
                LIMIT 1
            """, (user_id, nama_kegiatan, exclude_kegiatan))
        else:
            cursor.execute("""
                SELECT id FROM biodata_kegiatan
                WHERE user_id = ? AND TRIM(nama_kegiatan) = TRIM(?)
                LIMIT 1
            """, (user_id, nama_kegiatan))
        existing = cursor.fetchone()

        if existing:
            return jsonify({
                'has_data': True,
                'message': f'Anda sudah memiliki data untuk kegiatan "{nama_kegiatan}".  '
            })
        else:
            return jsonify({'has_data': False, 'message': 'Kegiatan tersedia'})
    except Exception as e:
        print(f"Error checking kegiatan: {e}")
        return jsonify({'has_data': False, 'message': f'Terjadi kesalahan: {str(e)}'})
    finally:
        if cursor:
            try:
                cursor.close()
            except:
                pass
        if connection and connection is not None:
            connection.close()

@app.route('/admin/kegiatan')
@admin_required
def admin_kegiatan():
    """Halaman admin untuk melihat Kegiatan (menggunakan DataTables untuk pagination dan sorting)"""

    connection = get_db_connection()
    if connection is None:
        flash('Koneksi database gagal!', 'error')
        return render_template('admin/admin-kegiatan.html', kegiatan_list=[], username=get_username(), user_role=get_user_role())

    user_id = get_user_id()
    user_role = get_user_role()

    try:
        cursor = connection.cursor()

        # Jika operator, hanya tampilkan kegiatan yang dia ikuti
        if user_role == 'operator':
            cursor.execute("""
                SELECT
                    k.nama_kegiatan,
                    COALESCE(COUNT(b.id), 0) as jumlah_peserta,
                    k.id as kegiatan_id,
                    COALESCE(k.is_hidden, 0) as is_hidden
                FROM kegiatan_master k
                INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                LEFT JOIN biodata_kegiatan b ON TRIM(k.nama_kegiatan) = TRIM(b.nama_kegiatan)
                WHERE ok.user_id = ?
                    AND TRIM(k.nama_kegiatan) != ''
                GROUP BY k.id, k.nama_kegiatan, k.is_hidden
                ORDER BY k.id DESC
            """, (user_id,))
        else:
            # Jika admin, tampilkan semua kegiatan
            cursor.execute("""
                SELECT
                    k.nama_kegiatan,
                    COALESCE(COUNT(b.id), 0) as jumlah_peserta,
                    k.id as kegiatan_id,
                    COALESCE(k.is_hidden, 0) as is_hidden
                FROM kegiatan_master k
                LEFT JOIN biodata_kegiatan b ON TRIM(k.nama_kegiatan) = TRIM(b.nama_kegiatan)
                WHERE TRIM(k.nama_kegiatan) != ''
                GROUP BY k.id, k.nama_kegiatan, k.is_hidden
                ORDER BY k.id DESC
            """)
        kegiatan_list = cursor.fetchall()
    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan saat mengambil data: {str(e)}', 'error')
        kegiatan_list = []
    finally:
        if connection:
            cursor.close()
            connection.close()

    return render_template('admin/admin-kegiatan.html', kegiatan_list=kegiatan_list, username=get_username(), user_role=get_user_role())

@app.route('/admin/toggle-hide-kegiatan/<int:kegiatan_id>', methods=['POST'])
@admin_required
def toggle_hide_kegiatan(kegiatan_id):
    """Toggle hide/show status kegiatan"""
    connection = get_db_connection()
    if connection is None:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_kegiatan'))
    
    user_role = get_user_role()
    if user_role == 'operator':
        flash('Anda tidak memiliki izin untuk melakukan aksi ini!', 'error')
        return redirect(url_for('admin_kegiatan'))
    
    try:
        cursor = connection.cursor()
        # Ambil status is_hidden saat ini
        cursor.execute("SELECT is_hidden, nama_kegiatan FROM kegiatan_master WHERE id = ?", (kegiatan_id,))
        result = cursor.fetchone()
        
        if not result:
            flash('Kegiatan tidak ditemukan!', 'error')
            return redirect(url_for('admin_kegiatan'))
        
        current_hidden = result[0] if result[0] is not None else 0
        nama_kegiatan = result[1]
        new_hidden = 1 if current_hidden == 0 else 0
        
        # Update status is_hidden
        cursor.execute("""
            UPDATE kegiatan_master
            SET is_hidden = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (new_hidden, kegiatan_id))
        connection.commit()
        
        status_text = "disembunyikan" if new_hidden == 1 else "ditampilkan"
        flash(f'Kegiatan "{nama_kegiatan}" berhasil {status_text}!', 'success')
    except sqlite3.Error as e:
        connection.rollback()
        flash(f'Terjadi kesalahan saat mengubah status: {str(e)}', 'error')
    finally:
        if connection:
            cursor.close()
            connection.close()
    
    return redirect(url_for('admin_kegiatan'))

@app.route('/api/get-peserta-kegiatan/<path:nama_kegiatan>', methods=['GET'])
@admin_required
def get_peserta_kegiatan(nama_kegiatan):
    """API endpoint untuk mendapatkan daftar peserta berdasarkan nama kegiatan"""
    from urllib.parse import unquote

    # Decode URL encoding
    nama_kegiatan = unquote(nama_kegiatan)

    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Koneksi database gagal!'}), 500

        try:
            cursor = connection.cursor()
            cursor.execute("""
                SELECT nama_lengkap, nik
                FROM biodata_kegiatan
                WHERE TRIM(nama_kegiatan) = TRIM(?)
                ORDER BY nama_lengkap ASC
            """, (nama_kegiatan,))
            peserta_list = cursor.fetchall()

            return jsonify({
                'success': True,
                'data': peserta_list,
                'jumlah': len(peserta_list)
            })
        except sqlite3.Error as e:
            return jsonify({'error': f'Terjadi kesalahan: {str(e)}'}), 500
        finally:
            if connection is not None:
                cursor.close()
                connection.close()
    except Exception as e:
        return jsonify({'error': f'Terjadi kesalahan tidak terduga: {str(e)}'}), 500

@app.route('/api/get-kegiatan-id/<path:nama_kegiatan>', methods=['GET'])
@admin_required
def get_kegiatan_id(nama_kegiatan):
    """API endpoint untuk mendapatkan ID kegiatan berdasarkan nama kegiatan"""
    from urllib.parse import unquote

    # Decode URL encoding
    nama_kegiatan = unquote(nama_kegiatan)

    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Koneksi database gagal!'}), 500

        try:
            cursor = connection.cursor()
            cursor.execute("""
                SELECT id FROM kegiatan_master
                WHERE TRIM(nama_kegiatan) = TRIM(?)
                LIMIT 1
            """, (nama_kegiatan,))
            kegiatan = cursor.fetchone()

            if kegiatan:
                return jsonify({
                    'success': True,
                    'id': kegiatan['id']
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'Kegiatan tidak ditemukan di kegiatan_master'
                })
        except sqlite3.Error as e:
            return jsonify({'error': f'Terjadi kesalahan: {str(e)}'}), 500
        finally:
            if connection is not None:
                cursor.close()
                connection.close()
    except Exception as e:
        return jsonify({'error': f'Terjadi kesalahan tidak terduga: {str(e)}'}), 500

def parse_waktu_pelaksanaan(waktu_pelaksanaan):
    """
    Parse waktu_pelaksanaan string menjadi datetime object untuk sorting.
    Format yang didukung: "15-20 Januari 2025", "10 Februari 2025", dll.
    Returns: datetime object atau datetime.min jika parsing gagal
    """
    if not waktu_pelaksanaan:
        return datetime.min

    # Mapping nama bulan Indonesia ke angka
    bulan_map = {
        'januari': 1, 'februari': 2, 'maret': 3, 'april': 4,
        'mei': 5, 'juni': 6, 'juli': 7, 'agustus': 8,
        'september': 9, 'oktober': 10, 'november': 11, 'desember': 12
    }

    try:
        # Normalize: lowercase dan hapus spasi berlebih
        text = waktu_pelaksanaan.lower().strip()

        # Extract tahun (4 digit)
        tahun_match = re.search(r'\b(19|20)\d{2}\b', text)
        if not tahun_match:
            return datetime.min
        tahun = int(tahun_match.group(0))  # Ambil seluruh match, bukan hanya group

        # Extract bulan
        bulan = None
        for bulan_nama, bulan_angka in bulan_map.items():
            if bulan_nama in text:
                bulan = bulan_angka
                break

        if not bulan:
            return datetime.min

        # Extract tanggal (ambil angka pertama, bisa range seperti "15-20")
        tanggal_match = re.search(r'(\d+)(?:\s*-\s*\d+)?', text)
        if not tanggal_match:
            return datetime.min
        tanggal = int(tanggal_match.group(1))

        # Validasi tanggal
        if tanggal < 1 or tanggal > 31:
            return datetime.min

        return datetime(tahun, bulan, tanggal)
    except:
        return datetime.min


@app.route('/admin/tambah-kegiatan', methods=['GET', 'POST'])
@admin_required
def admin_tambah_kegiatan():
    """Halaman admin untuk menambahkan kegiatan baru"""
    # Operator tidak bisa mengakses halaman tambah kegiatan
    if get_user_role() == 'operator':
        flash('Operator tidak memiliki akses ke halaman ini!', 'error')
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        nama_kegiatan = request.form.get('nama_kegiatan', '').strip()
        waktu_pelaksanaan = request.form.get('waktu_pelaksanaan', '').strip()
        tempat_pelaksanaan = request.form.get('tempat_pelaksanaan', '').strip()

        # Validasi
        if not nama_kegiatan:
            flash('Nama kegiatan wajib diisi!', 'error')
        elif not waktu_pelaksanaan:
            flash('Waktu pelaksanaan wajib diisi!', 'error')
        elif not tempat_pelaksanaan:
            flash('Tempat pelaksanaan wajib diisi!', 'error')
        else:
            connection = get_db_connection()
            if connection is None:
                flash('Koneksi database gagal!', 'error')
            else:
                try:
                    cursor = connection.cursor()
                    # Cek apakah nama kegiatan sudah ada (case-insensitive dan trim whitespace)
                    cursor.execute("""
                        SELECT id FROM kegiatan_master
                        WHERE TRIM(LOWER(nama_kegiatan)) = TRIM(LOWER(?))
                        LIMIT 1
                    """, (nama_kegiatan,))
                    existing_kegiatan = cursor.fetchone()

                    if existing_kegiatan:
                        flash(f'Nama kegiatan "{nama_kegiatan}" sudah ada! Silakan gunakan nama kegiatan yang berbeda.', 'error')
                    else:
                        # Insert data kegiatan baru
                        cursor.execute("""
                            INSERT INTO kegiatan_master (nama_kegiatan, waktu_pelaksanaan, tempat_pelaksanaan)
                            VALUES (?, ?, ?)
                        """, (nama_kegiatan, waktu_pelaksanaan, tempat_pelaksanaan))
                        connection.commit()
                        flash('Kegiatan berhasil ditambahkan!', 'success')
                        return redirect(url_for('admin_kegiatan'))
                except sqlite3.Error as e:
                    connection.rollback()
                    flash(f'Terjadi kesalahan saat menyimpan data: {str(e)}', 'error')
                finally:
                    if connection is not None:
                        cursor.close()
                        connection.close()

    # Ambil daftar semua kegiatan yang sudah ditambahkan
    connection = get_db_connection()
    kegiatan_list = []
    if connection:
        try:
            cursor = connection.cursor()
            cursor.execute("""
                SELECT id, nama_kegiatan, waktu_pelaksanaan, tempat_pelaksanaan
                FROM kegiatan_master
            """)
            rows = cursor.fetchall()

            # Convert Row objects to dicts
            kegiatan_list = [row_to_dict(row) for row in rows]

            # Urutkan berdasarkan waktu pelaksanaan (parse tanggal)
            kegiatan_list.sort(key=lambda x: parse_waktu_pelaksanaan(x.get('waktu_pelaksanaan', '')))
        except sqlite3.Error as e:
            flash(f'Terjadi kesalahan saat mengambil data: {str(e)}', 'error')
        finally:
            if connection is not None:
                cursor.close()
                connection.close()

    return render_template('admin/admin-tambah-kegiatan.html', kegiatan_list=kegiatan_list, username=get_username(), user_role=get_user_role())

@app.route('/operator/edit-kegiatan/<int:kegiatan_id>', methods=['GET', 'POST'])
def operator_edit_kegiatan(kegiatan_id):
    """Halaman operator untuk mengedit kegiatan"""
    # Hanya operator yang bisa akses route ini
    if not is_logged_in():
        flash('Silakan login terlebih dahulu!', 'error')
        return redirect(url_for('login'))

    if get_user_role() != 'operator':
        flash('Hanya operator yang dapat mengakses halaman ini!', 'error')
        return redirect(url_for('admin_dashboard'))

    return _edit_kegiatan(kegiatan_id, is_operator=True)

@app.route('/admin/edit-kegiatan/<int:kegiatan_id>', methods=['GET', 'POST'])
@admin_required
def admin_edit_kegiatan(kegiatan_id):
    """Halaman admin untuk mengedit kegiatan"""
    return _edit_kegiatan(kegiatan_id, is_operator=False)

def _edit_kegiatan(kegiatan_id, is_operator=False):
    """Fungsi helper untuk edit kegiatan (digunakan oleh admin dan operator)"""
    connection = get_db_connection()
    if connection is None:
        flash('Koneksi database gagal!', 'error')
        if is_operator:
            return redirect(url_for('admin_kegiatan'))
        else:
            return redirect(url_for('admin_tambah_kegiatan'))

    # Ambil data kegiatan berdasarkan ID
    try:
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM kegiatan_master WHERE id = ?", (kegiatan_id,))
        kegiatan_row = cursor.fetchone()
        kegiatan = row_to_dict(kegiatan_row) if kegiatan_row else None

        if not kegiatan:
            flash('Kegiatan tidak ditemukan!', 'error')
            if connection is not None:
                cursor.close()
                connection.close()
            if is_operator:
                return redirect(url_for('admin_kegiatan'))
            else:
                return redirect(url_for('admin_tambah_kegiatan'))

        if request.method == 'POST':
            nama_kegiatan = request.form.get('nama_kegiatan', '').strip()
            waktu_pelaksanaan = request.form.get('waktu_pelaksanaan', '').strip()
            tempat_pelaksanaan = request.form.get('tempat_pelaksanaan', '').strip()

            # Validasi
            if not nama_kegiatan:
                flash('Nama kegiatan wajib diisi!', 'error')
            elif not waktu_pelaksanaan:
                flash('Waktu pelaksanaan wajib diisi!', 'error')
            elif not tempat_pelaksanaan:
                flash('Tempat pelaksanaan wajib diisi!', 'error')
            else:
                try:
                    # Simpan nama_kegiatan lama untuk update biodata
                    old_nama_kegiatan = (kegiatan.get('nama_kegiatan') or '').strip()
                    new_nama_kegiatan = nama_kegiatan.strip()

                    # Cek apakah nama kegiatan baru sudah ada di kegiatan lain (case-insensitive dan trim whitespace)
                    # Exclude kegiatan yang sedang diedit
                    cursor.execute("""
                        SELECT id FROM kegiatan_master
                        WHERE TRIM(LOWER(nama_kegiatan)) = TRIM(LOWER(?))
                        AND id != ?
                        LIMIT 1
                    """, (nama_kegiatan, kegiatan_id))
                    existing_kegiatan = cursor.fetchone()

                    if existing_kegiatan:
                        flash(f'Nama kegiatan "{nama_kegiatan}" sudah ada! Silakan gunakan nama kegiatan yang berbeda.', 'error')
                    else:
                        # Update data kegiatan
                        cursor.execute("""
                            UPDATE kegiatan_master
                            SET nama_kegiatan = ?,
                                waktu_pelaksanaan = ?,
                                tempat_pelaksanaan = ?
                            WHERE id = ?
                        """, (nama_kegiatan, waktu_pelaksanaan, tempat_pelaksanaan, kegiatan_id))

                        # Jika nama_kegiatan berubah, update semua biodata yang terkait
                        if old_nama_kegiatan != new_nama_kegiatan:
                            # Update semua biodata_kegiatan yang memiliki nama_kegiatan lama
                            cursor.execute("""
                                UPDATE biodata_kegiatan
                                SET nama_kegiatan = ?,
                                    waktu_pelaksanaan = ?,
                                    tempat_pelaksanaan = ?
                                WHERE TRIM(nama_kegiatan) = TRIM(?)
                            """, (new_nama_kegiatan, waktu_pelaksanaan, tempat_pelaksanaan, old_nama_kegiatan))
                            jumlah_terupdate = cursor.rowcount

                            connection.commit()
                            if jumlah_terupdate > 0:
                                flash(f'Kegiatan berhasil diperbarui! {jumlah_terupdate} data biodata terkait juga telah diperbarui.', 'success')
                            else:
                                flash('Kegiatan berhasil diperbarui!', 'success')
                        else:
                            # Jika hanya waktu atau tempat yang berubah, update juga biodata
                            cursor.execute("""
                                UPDATE biodata_kegiatan
                                SET waktu_pelaksanaan = ?,
                                    tempat_pelaksanaan = ?
                                WHERE TRIM(nama_kegiatan) = TRIM(?)
                            """, (waktu_pelaksanaan, tempat_pelaksanaan, new_nama_kegiatan))
                            jumlah_terupdate = cursor.rowcount

                            connection.commit()
                            if jumlah_terupdate > 0:
                                flash(f'Kegiatan berhasil diperbarui! {jumlah_terupdate} data biodata terkait juga telah diperbarui.', 'success')
                            else:
                                flash('Kegiatan berhasil diperbarui!', 'success')

                        if connection is not None:
                            cursor.close()
                            connection.close()
                        if is_operator:
                            return redirect(url_for('admin_kegiatan'))
                        else:
                            return redirect(url_for('admin_kegiatan'))
                except sqlite3.Error as e:
                    if connection:
                        connection.rollback()
                    flash(f'Terjadi kesalahan saat memperbarui data: {str(e)}', 'error')
                    import traceback
                    traceback.print_exc()
                except Exception as e:
                    if connection:
                        connection.rollback()
                    flash(f'Terjadi kesalahan tidak terduga: {str(e)}', 'error')
                    import traceback
                    traceback.print_exc()
    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    except Exception as e:
        flash(f'Terjadi kesalahan tidak terduga: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    finally:
        if connection:
            try:
                if cursor:
                    cursor.close()
            except:
                pass
            connection.close()

    return render_template('admin/admin-edit-kegiatan.html', kegiatan=kegiatan, username=get_username(), user_role=get_user_role())

@app.route('/admin/edit-biodata/<nik>/<path:nama_kegiatan>', methods=['GET', 'POST'])
@admin_required
def admin_edit_biodata(nik, nama_kegiatan):
    """Edit biodata - redirect sesuai halaman asal (rekap kabupaten atau detail kegiatan)Halaman admin untuk mengedit biodata user berdasarkan NIK dan nama kegiatan"""
    from urllib.parse import unquote

    current_year = datetime.now(ZoneInfo("Asia/Makassar")).year
    nama_kegiatan = unquote(nama_kegiatan)

    # Ambil data biodata berdasarkan NIK dan nama_kegiatan
    connection = get_db_connection()
    if connection is None:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_kegiatan'))

    biodata = None
    try:
        cursor = connection.cursor()
        cursor.execute("""
            SELECT * FROM biodata_kegiatan
            WHERE nik = ? AND TRIM(nama_kegiatan) = TRIM(?)
            LIMIT 1
        """, (nik, nama_kegiatan))
        biodata_row = cursor.fetchone()
        biodata = row_to_dict(biodata_row) if biodata_row else None

        if not biodata:
            flash('Data tidak ditemukan!', 'error')
            return redirect(url_for('admin_kegiatan'))

        # Normalisasi path buku tabungan
        if biodata and biodata.get('buku_tabungan_path'):
            biodata['buku_tabungan_path'] = normalize_buku_tabungan_path(biodata['buku_tabungan_path'])
    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan saat mengambil data: {str(e)}', 'error')
        return redirect(url_for('admin_kegiatan'))
    finally:
        if connection:
            cursor.close()
            connection.close()

    # Ambil daftar kegiatan untuk dropdown: hanya yang tidak disembunyikan.
    # Kecuali kegiatan biodata yang sedang diedit — selalu tampilkan agar form tetap valid.
    kegiatan_list = []
    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor()
            cursor.execute("""
                SELECT DISTINCT nama_kegiatan
                FROM kegiatan_master
                WHERE (is_hidden IS NULL OR is_hidden = 0)
                ORDER BY nama_kegiatan ASC
            """)
            rows = cursor.fetchall()
            kegiatan_list = [row_to_dict(row) for row in rows]
            nama_set = {r['nama_kegiatan'] for r in kegiatan_list}
            current_nama = (biodata or {}).get('nama_kegiatan')
            if current_nama and current_nama not in nama_set:
                kegiatan_list.append({'nama_kegiatan': current_nama})
                kegiatan_list.sort(key=lambda x: (x.get('nama_kegiatan') or ''))
        except sqlite3.Error as e:
            print(f"Error fetching kegiatan: {e}")
        finally:
            if connection is not None:
                cursor.close()
                connection.close()

    # Handle POST request
    if request.method == 'POST':
        try:
            form_data = get_form_data()
            user_id = biodata['user_id']  # Gunakan user_id dari biodata yang ditemukan

            # Handle file upload
            buku_tabungan_file = request.files.get('buku_tabungan')
            buku_tabungan_path = None
            if buku_tabungan_file and buku_tabungan_file.filename:
                # File baru diupload
                buku_tabungan_path = save_uploaded_file(buku_tabungan_file, form_data['nik'])

            # Handle tanda tangan - simpan ke file static dan database
            tanda_tangan_base64 = form_data.get('tanda_tangan')
            if tanda_tangan_base64:
                # Admin membuat tanda tangan baru, simpan sebagai file
                print(f"DEBUG admin_edit_biodata: Processing new tanda_tangan (length: {len(tanda_tangan_base64)})")
                # Cek apakah sudah berupa path file atau masih base64
                if 'uploads/' in str(tanda_tangan_base64) or str(tanda_tangan_base64).startswith('static/'):
                    # Sudah berupa path file, normalisasi saja
                    form_data['tanda_tangan'] = normalize_buku_tabungan_path(tanda_tangan_base64)
                    print(f"DEBUG admin_edit_biodata: Tanda tangan sudah berupa path: {form_data['tanda_tangan']}")
                else:
                    # Masih base64, simpan sebagai file
                    print(f"DEBUG admin_edit_biodata: Tanda tangan adalah base64, menyimpan sebagai file...")
                    tanda_tangan_path = save_tanda_tangan_file(tanda_tangan_base64, form_data['nik'])
                    if tanda_tangan_path:
                        form_data['tanda_tangan'] = tanda_tangan_path
                        print(f"DEBUG admin_edit_biodata: Tanda tangan disimpan sebagai file: {tanda_tangan_path}")
                    else:
                        flash('Gagal menyimpan tanda tangan!', 'error')
                        return render_template('admin/admin-edit-biodata.html', biodata=biodata, kegiatan_list=kegiatan_list, username=get_username(), nik=nik, nama_kegiatan=nama_kegiatan, current_year=current_year)
            elif biodata and biodata.get('tanda_tangan'):
                # Admin tidak mengubah tanda tangan, gunakan tanda tangan yang sudah ada
                existing_ttd = biodata.get('tanda_tangan')
                print(f"DEBUG admin_edit_biodata: Using existing tanda_tangan (length: {len(existing_ttd) if existing_ttd else 0})")
                # Cek apakah sudah berupa path file atau masih base64
                if existing_ttd and ('uploads/' in str(existing_ttd) or str(existing_ttd).startswith('static/')):
                    # Sudah berupa path file, normalisasi saja
                    form_data['tanda_tangan'] = normalize_buku_tabungan_path(existing_ttd)
                    print(f"DEBUG admin_edit_biodata: Existing tanda tangan adalah path: {form_data['tanda_tangan']}")
                elif existing_ttd:
                    # Masih base64, simpan sebagai file
                    print(f"DEBUG admin_edit_biodata: Existing tanda tangan adalah base64, menyimpan sebagai file...")
                    tanda_tangan_path = save_tanda_tangan_file(existing_ttd, form_data['nik'])
                    if tanda_tangan_path:
                        form_data['tanda_tangan'] = tanda_tangan_path
                        print(f"DEBUG admin_edit_biodata: Existing tanda tangan disimpan sebagai file: {tanda_tangan_path}")
                    else:
                        # Jika gagal, tetap gunakan base64 yang ada
                        form_data['tanda_tangan'] = existing_ttd
                        print(f"DEBUG admin_edit_biodata: Gagal menyimpan existing tanda tangan, menggunakan base64")
                else:
                    form_data['tanda_tangan'] = existing_ttd
            else:
                print(f"DEBUG admin_edit_biodata: WARNING - No tanda_tangan provided and no existing tanda_tangan!")

            # Validasi
            if not form_data['nik']:
                flash('NIK wajib diisi!', 'error')
                return render_template('admin/admin-edit-biodata.html', biodata=biodata, kegiatan_list=kegiatan_list, username=get_username(), nik=nik, nama_kegiatan=nama_kegiatan, current_year=current_year)

            # Validasi NIK harus tepat 16 digit
            if not form_data['nik'].isdigit() or len(form_data['nik']) != 16:
                flash('NIK harus tepat 16 digit angka!', 'error')
                return render_template('admin/admin-edit-biodata.html', biodata=biodata, kegiatan_list=kegiatan_list, username=get_username(), nik=nik, nama_kegiatan=nama_kegiatan, current_year=current_year)

            if not validate_required_fields(form_data):
                flash('Semua field wajib harus diisi!', 'error')
                return render_template('admin/admin-edit-biodata.html', biodata=biodata, kegiatan_list=kegiatan_list, username=get_username(), nik=nik, nama_kegiatan=nama_kegiatan, current_year=current_year)

            # Simpan data (update) - gunakan fungsi khusus admin berdasarkan NIK dan nama_kegiatan
            # nama_kegiatan parameter adalah old_nama_kegiatan untuk identifikasi data yang akan diupdate
            success, message = admin_update_biodata(form_data, nik, nama_kegiatan, buku_tabungan_path)

            if success:
                flash(f'Biodata kegiatan untuk "{form_data["nama_lengkap"]}" {message}', 'success')
                # Cek dari mana user datang untuk menentukan redirect
                # Bisa dari query parameter atau form data
                from_page = request.form.get('from_page') or request.args.get('from', '')
                if from_page == 'rekap-filter':
                    return redirect(url_for('admin_rekap_filter'))
                else:
                    # Default: redirect ke detail kegiatan
                    from urllib.parse import quote
                    return redirect(url_for('admin_detail_kegiatan', nama_kegiatan=form_data["nama_kegiatan"]))
            else:
                flash(message, 'error')
                return render_template('admin/admin-edit-biodata.html', biodata=biodata, kegiatan_list=kegiatan_list, username=get_username(), nik=nik, nama_kegiatan=nama_kegiatan, current_year=current_year)

        except Exception as e:
            print(f"Unexpected error in admin_edit_biodata: {e}")
            import traceback
            traceback.print_exc()
            flash(f'Terjadi kesalahan tidak terduga: {str(e)}', 'error')
            return render_template('admin/admin-edit-biodata.html', biodata=biodata, kegiatan_list=kegiatan_list, username=get_username(), nik=nik, nama_kegiatan=nama_kegiatan, current_year=current_year)

    # Handle GET request - tampilkan form dengan data existing
    # Ambil parameter from untuk diteruskan ke template
    from_page = request.args.get('from', '')
    return render_template('admin/admin-edit-biodata.html', biodata=biodata, kegiatan_list=kegiatan_list, username=get_username(), nik=nik, nama_kegiatan=nama_kegiatan, from_page=from_page, current_year=current_year)

@app.route('/operator/hapus-kegiatan/<int:kegiatan_id>', methods=['POST'])
def operator_hapus_kegiatan(kegiatan_id):
    """Hapus kegiatan berdasarkan ID (untuk operator)"""
    # Hanya operator yang bisa akses route ini
    if not is_logged_in():
        flash('Silakan login terlebih dahulu!', 'error')
        return redirect(url_for('login'))

    if get_user_role() != 'operator':
        flash('Hanya operator yang dapat mengakses halaman ini!', 'error')
        return redirect(url_for('admin_dashboard'))

    return _hapus_kegiatan(kegiatan_id, is_operator=True)

@app.route('/admin/hapus-kegiatan/<int:kegiatan_id>', methods=['POST'])
@admin_required
def admin_hapus_kegiatan(kegiatan_id):
    """Hapus kegiatan berdasarkan ID (untuk admin)"""
    return _hapus_kegiatan(kegiatan_id, is_operator=False)

def _hapus_kegiatan(kegiatan_id, is_operator=False):
    """Fungsi helper untuk hapus kegiatan (digunakan oleh admin dan operator)"""
    connection = get_db_connection()
    if connection is None:
        flash('Koneksi database gagal!', 'error')
        if is_operator:
            return redirect(url_for('admin_kegiatan'))
        else:
            # Cek referer untuk menentukan redirect
            referer = request.headers.get('Referer', '')
            if 'tambah-kegiatan' in referer:
                return redirect(url_for('admin_tambah_kegiatan'))
            else:
                return redirect(url_for('admin_kegiatan'))

    try:
        cursor = connection.cursor()
        # Ambil nama_kegiatan sebelum dihapus
        cursor.execute("SELECT nama_kegiatan FROM kegiatan_master WHERE id = ?", (kegiatan_id,))
        kegiatan = cursor.fetchone()

        if not kegiatan:
            flash('Kegiatan tidak ditemukan!', 'error')
        else:
            nama_kegiatan = kegiatan['nama_kegiatan']

            # Update semua biodata_kegiatan yang memiliki nama_kegiatan yang sama
            # Set nama_kegiatan, waktu_pelaksanaan, dan tempat_pelaksanaan menjadi string kosong (data user lainnya tetap utuh)
            cursor.execute("""
                UPDATE biodata_kegiatan
                SET nama_kegiatan = '',
                    waktu_pelaksanaan = '',
                    tempat_pelaksanaan = ''
                WHERE TRIM(nama_kegiatan) = TRIM(?)
            """, (nama_kegiatan,))

            jumlah_terpengaruh = cursor.rowcount

            # Hapus relasi operator_kegiatan yang terkait
            cursor.execute("DELETE FROM operator_kegiatan WHERE kegiatan_id = ?", (kegiatan_id,))

            # Hapus kegiatan dari kegiatan_master
            cursor.execute("DELETE FROM kegiatan_master WHERE id = ?", (kegiatan_id,))
            connection.commit()

            if jumlah_terpengaruh > 0:
                flash(f'Kegiatan "{nama_kegiatan}" berhasil dihapus! Nama kegiatan, waktu pelaksanaan, dan tempat pelaksanaan pada {jumlah_terpengaruh} data biodata telah dihapus (data user lainnya tetap utuh).', 'success')
            else:
                flash(f'Kegiatan "{nama_kegiatan}" berhasil dihapus!', 'success')
    except sqlite3.Error as e:
        connection.rollback()
        flash(f'Terjadi kesalahan saat menghapus data: {str(e)}', 'error')
    finally:
        if connection:
            cursor.close()
            connection.close()

    # Cek referer untuk menentukan redirect
    referer = request.headers.get('Referer', '')
    if is_operator:
        return redirect(url_for('admin_kegiatan'))
    elif 'tambah-kegiatan' in referer:
        return redirect(url_for('admin_tambah_kegiatan'))
    else:
        return redirect(url_for('admin_kegiatan'))

@app.route('/admin/hapus-biodata/<path:nik>/<path:nama_kegiatan>', methods=['POST'])
@admin_required
def admin_hapus_biodata(nik, nama_kegiatan):
    """Hapus biodata berdasarkan NIK dan nama_kegiatan"""
    from urllib.parse import unquote

    # Decode URL encoding
    nik = unquote(nik)
    nama_kegiatan = unquote(nama_kegiatan)

    connection = get_db_connection()
    if not connection:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_rekap_filter'))

    try:
        cursor = connection.cursor()

        # Ambil data biodata untuk mendapatkan nama_lengkap (untuk pesan)
        # Handle kasus nama_kegiatan kosong
        if nama_kegiatan and nama_kegiatan.strip():
            cursor.execute("""
                SELECT nama_lengkap FROM biodata_kegiatan
                WHERE nik = ? AND TRIM(COALESCE(nama_kegiatan, '')) = TRIM(?)
                LIMIT 1
            """, (nik, nama_kegiatan))
        else:
            cursor.execute("""
                SELECT nama_lengkap FROM biodata_kegiatan
                WHERE nik = ? AND (nama_kegiatan IS NULL OR TRIM(COALESCE(nama_kegiatan, '')) = '')
                LIMIT 1
            """, (nik,))
        biodata = cursor.fetchone()

        if not biodata:
            flash('Data tidak ditemukan!', 'error')
            return redirect(url_for('admin_rekap_filter'))

        nama_lengkap = biodata['nama_lengkap']

        # Hapus data biodata
        # Handle kasus nama_kegiatan kosong dengan menggunakan COALESCE atau kondisi khusus
        if nama_kegiatan and nama_kegiatan.strip():
            cursor.execute("""
                DELETE FROM biodata_kegiatan
                WHERE nik = ? AND TRIM(COALESCE(nama_kegiatan, '')) = TRIM(?)
            """, (nik, nama_kegiatan))
        else:
            # Jika nama_kegiatan kosong, hapus berdasarkan NIK saja dengan nama_kegiatan yang NULL atau kosong
            cursor.execute("""
                DELETE FROM biodata_kegiatan
                WHERE nik = ? AND (nama_kegiatan IS NULL OR TRIM(COALESCE(nama_kegiatan, '')) = '')
            """, (nik,))

        connection.commit()
        kegiatan_display = nama_kegiatan.strip() if nama_kegiatan and nama_kegiatan.strip() else '(Kegiatan Kosong)'
        flash(f'Data biodata untuk "{nama_lengkap}" (NIK: {nik}) pada kegiatan "{kegiatan_display}" berhasil dihapus!', 'success')

    except sqlite3.Error as e:
        if connection:
            connection.rollback()
        print(f"Error deleting biodata: {e}")
        flash(f'Terjadi kesalahan saat menghapus data: {str(e)}', 'error')
    finally:
        if connection:
            cursor.close()
            connection.close()

    # Redirect kembali ke rekap kabupaten
    return redirect(url_for('admin_rekap_filter'))

@app.route('/admin/users')
@admin_required
def admin_users():
    """Halaman admin untuk melihat daftar operator"""
    # Operator tidak bisa mengakses halaman daftar operator
    if get_user_role() == 'operator':
        flash('Operator tidak memiliki akses ke halaman ini!', 'error')
        return redirect(url_for('admin_dashboard'))

    connection = get_db_connection()
    if connection is None:
        flash('Koneksi database gagal!', 'error')
        return render_template('admin/admin-users.html', operators_list=[], kegiatan_list=[], all_selected_kegiatan_ids=[], username=get_username())

    try:
        cursor = connection.cursor()

        # Ambil semua operator (DataTables client-side)
        cursor.execute("""
            SELECT
                u.id,
                u.nama,
                u.username
            FROM users u
            WHERE u.role = 'operator'
            ORDER BY u.id ASC
        """)
        rows = cursor.fetchall()
        operators_list = [row_to_dict(row) for row in rows]

        # Ambil kegiatan untuk setiap operator
        for operator in operators_list:
            cursor.execute("""
                SELECT k.nama_kegiatan, k.id as kegiatan_id
                FROM operator_kegiatan ok
                INNER JOIN kegiatan_master k ON ok.kegiatan_id = k.id
                WHERE ok.user_id = ?
                ORDER BY k.nama_kegiatan ASC
            """, (operator['id'],))
            rows = cursor.fetchall()
            operator['kegiatan_list'] = [row_to_dict(row) for row in rows]

        # Ambil semua kegiatan yang tersedia untuk dropdown
        cursor.execute("""
            SELECT id, nama_kegiatan
            FROM kegiatan_master
            WHERE TRIM(nama_kegiatan) != ''
            ORDER BY nama_kegiatan ASC
        """)
        rows = cursor.fetchall()
        kegiatan_list = [row_to_dict(row) for row in rows]

        # Ambil semua kegiatan yang sudah dipilih oleh operator manapun (untuk filter global)
        cursor.execute("""
            SELECT DISTINCT kegiatan_id
            FROM operator_kegiatan
        """)
        all_selected_kegiatan_ids = [row['kegiatan_id'] for row in cursor.fetchall()]
    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan saat mengambil data: {str(e)}', 'error')
        operators_list = []
        kegiatan_list = []
        all_selected_kegiatan_ids = []
    finally:
        if connection:
            cursor.close()
            connection.close()

    return render_template('admin/admin-users.html', operators_list=operators_list, kegiatan_list=kegiatan_list, all_selected_kegiatan_ids=all_selected_kegiatan_ids, username=get_username())

@app.route('/admin/tambah-operator', methods=['GET', 'POST'])
@admin_required
def admin_tambah_operator():
    """Halaman admin untuk menambahkan operator baru"""
    # Operator tidak bisa mengakses halaman tambah operator
    if get_user_role() == 'operator':
        flash('Operator tidak memiliki akses ke halaman ini!', 'error')
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        nama = request.form.get('nama', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        # Validasi
        if not nama:
            flash('Nama wajib diisi!', 'error')
        elif not username:
            flash('Username wajib diisi!', 'error')
        elif not password:
            flash('Password wajib diisi!', 'error')
        else:
            connection = get_db_connection()
            if connection is None:
                flash('Koneksi database gagal!', 'error')
            else:
                try:
                    cursor = connection.cursor()
                    # Cek apakah username sudah ada
                    cursor.execute("SELECT id FROM users WHERE username = ?", (username,))
                    if cursor.fetchone():
                        flash('Username sudah digunakan!', 'error')
                    else:
                        # Insert operator baru
                        cursor.execute("""
                            INSERT INTO users (nama, username, password, role)
                            VALUES (?, ?, ?, 'operator')
                        """, (nama, username, password))

                        connection.commit()
                        flash('Operator berhasil ditambahkan!', 'success')
                        # Setelah menambahkan operator, kembali ke daftar operator
                        return redirect(url_for('admin_users'))
                except sqlite3.Error as e:
                    connection.rollback()
                    flash(f'Terjadi kesalahan saat menyimpan data: {str(e)}', 'error')
                finally:
                    if connection is not None:
                        cursor.close()
                        connection.close()

    # Ambil daftar semua operator untuk ditampilkan di bawah form
    connection = get_db_connection()
    operators_list = []
    if connection:
        try:
            cursor = connection.cursor()
            cursor.execute("""
                SELECT
                    u.id,
                    u.nama,
                    u.username
                FROM users u
                WHERE u.role = 'operator'
                ORDER BY u.id ASC
            """)
            operators_list = cursor.fetchall()
        except sqlite3.Error as e:
            print(f"Error fetching operators list: {e}")
        finally:
            if connection is not None:
                cursor.close()
                connection.close()

    return render_template('admin/admin-tambah-operator.html',
                         username=get_user_display_name(),
                         operators_list=operators_list)

@app.route('/admin/edit-operator/<int:operator_id>', methods=['GET', 'POST'])
@admin_required
def admin_edit_operator(operator_id):
    """Halaman admin untuk mengedit operator"""
    # Operator tidak bisa mengakses halaman edit operator
    if get_user_role() == 'operator':
        flash('Operator tidak memiliki akses ke halaman ini!', 'error')
        return redirect(url_for('admin_dashboard'))

    connection = get_db_connection()
    if connection is None:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_users'))

    try:
        cursor = connection.cursor()

        # Cek apakah operator ada dan adalah operator
        cursor.execute("SELECT id, nama, username FROM users WHERE id = ? AND role = 'operator'", (operator_id,))
        operator_row = cursor.fetchone()
        operator = row_to_dict(operator_row) if operator_row else None

        if not operator:
            flash('Operator tidak ditemukan!', 'error')
            return redirect(url_for('admin_users'))

        if request.method == 'POST':
            nama = request.form.get('nama', '').strip()

            # Validasi
            if not nama:
                flash('Nama wajib diisi!', 'error')
            else:
                try:
                    # Update nama operator
                    cursor.execute("UPDATE users SET nama = ? WHERE id = ?", (nama, operator_id))

                    connection.commit()
                    flash('Data operator berhasil diperbarui!', 'success')
                    # Setelah simpan, kembali ke daftar operator
                    return redirect(url_for('admin_users'))
                except sqlite3.Error as e:
                    connection.rollback()
                    flash(f'Terjadi kesalahan saat menyimpan data: {str(e)}', 'error')

    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan: {str(e)}', 'error')
        return redirect(url_for('admin_users'))
    finally:
        if connection:
            cursor.close()
            connection.close()

    return render_template('admin/admin-edit-operator.html',
                         operator=operator,
                         username=get_user_display_name())

@app.route('/admin/reset-password-operator/<int:operator_id>', methods=['POST'])
@admin_required
def admin_reset_password_operator(operator_id):
    """Reset password operator menjadi default"""
    # Operator tidak bisa reset password operator lain
    if get_user_role() == 'operator':
        flash('Operator tidak memiliki akses!', 'error')
        return redirect(url_for('admin_dashboard'))

    connection = get_db_connection()
    if connection is None:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_users'))

    try:
        cursor = connection.cursor()

        # Cek apakah operator ada dan adalah operator
        cursor.execute("SELECT id, nama, username FROM users WHERE id = ? AND role = 'operator'", (operator_id,))
        operator_row = cursor.fetchone()

        if not operator_row:
            flash('Operator tidak ditemukan!', 'error')
            return redirect(url_for('admin_users'))

        # Convert Row to dict
        operator = row_to_dict(operator_row)

        # Reset password menjadi default
        default_password = 'operator123'
        cursor.execute("UPDATE users SET password = ? WHERE id = ?", (default_password, operator_id))
        connection.commit()

        operator_name = operator.get('nama') or operator.get('username')
        flash(f'Password operator "{operator_name}" berhasil direset menjadi: {default_password}', 'success')

    except sqlite3.Error as e:
        connection.rollback()
        flash(f'Terjadi kesalahan saat mereset password: {str(e)}', 'error')
    finally:
        if connection:
            cursor.close()
            connection.close()

    return redirect(url_for('admin_users'))

@app.route('/admin/operator/<int:operator_id>/tambah-kegiatan', methods=['POST'])
@admin_required
def admin_tambah_kegiatan_operator(operator_id):
    """Menambahkan kegiatan ke operator"""
    # Operator tidak bisa menambahkan kegiatan ke operator lain
    if get_user_role() == 'operator':
        flash('Operator tidak memiliki akses!', 'error')
        return redirect(url_for('admin_dashboard'))

    kegiatan_id = request.form.get('kegiatan_id', '').strip()

    if not kegiatan_id:
        flash('Pilih kegiatan yang ingin ditambahkan!', 'error')
        return redirect(url_for('admin_users'))

    connection = get_db_connection()
    if connection is None:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_users'))

    try:
        cursor = connection.cursor()

        # Cek apakah operator ada dan adalah operator
        cursor.execute("SELECT id, nama, username FROM users WHERE id = ? AND role = 'operator'", (operator_id,))
        operator_row = cursor.fetchone()

        if not operator_row:
            flash('Operator tidak ditemukan!', 'error')
            return redirect(url_for('admin_users'))

        # Convert Row to dict
        operator = row_to_dict(operator_row)

        # Cek apakah kegiatan sudah terdaftar untuk operator ini
        cursor.execute("""
            SELECT id FROM operator_kegiatan
            WHERE user_id = ? AND kegiatan_id = ?
        """, (operator_id, int(kegiatan_id)))
        existing = cursor.fetchone()

        if existing:
            # Ambil nama kegiatan untuk pesan error
            cursor.execute("SELECT nama_kegiatan FROM kegiatan_master WHERE id = ?", (int(kegiatan_id),))
            kegiatan_row = cursor.fetchone()
            if kegiatan_row:
                kegiatan = row_to_dict(kegiatan_row)
                kegiatan_name = kegiatan.get('nama_kegiatan', 'kegiatan ini')
            else:
                kegiatan_name = 'kegiatan ini'
            flash(f'Kegiatan "{kegiatan_name}" sudah terdaftar untuk operator ini!', 'error')
        else:
            # Tambahkan kegiatan ke operator
            cursor.execute("""
                INSERT INTO operator_kegiatan (user_id, kegiatan_id)
                VALUES (?, ?)
            """, (operator_id, int(kegiatan_id)))
            connection.commit()

            # Ambil nama kegiatan untuk pesan success
            cursor.execute("SELECT nama_kegiatan FROM kegiatan_master WHERE id = ?", (int(kegiatan_id),))
            kegiatan_row = cursor.fetchone()
            if kegiatan_row:
                kegiatan = row_to_dict(kegiatan_row)
                kegiatan_name = kegiatan.get('nama_kegiatan', 'kegiatan')
            else:
                kegiatan_name = 'kegiatan'
            operator_name = operator.get('nama') or operator.get('username')
            flash(f'Kegiatan "{kegiatan_name}" berhasil ditambahkan ke operator "{operator_name}"!', 'success')

    except sqlite3.Error as e:
        connection.rollback()
        flash(f'Terjadi kesalahan saat menambahkan kegiatan: {str(e)}', 'error')
    finally:
        if connection:
            cursor.close()
            connection.close()

    return redirect(url_for('admin_users'))

@app.route('/admin/operator/<int:operator_id>/hapus-kegiatan/<int:kegiatan_id>', methods=['POST'])
@admin_required
def admin_hapus_kegiatan_operator(operator_id, kegiatan_id):
    """Menghapus kegiatan dari operator"""
    # Operator tidak bisa menghapus kegiatan operator lain
    if get_user_role() == 'operator':
        flash('Operator tidak memiliki akses!', 'error')
        return redirect(url_for('admin_dashboard'))

    connection = get_db_connection()
    if connection is None:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_users'))

    try:
        cursor = connection.cursor()

        # Cek apakah operator ada dan adalah operator
        cursor.execute("SELECT id, nama, username FROM users WHERE id = ? AND role = 'operator'", (operator_id,))
        operator_row = cursor.fetchone()

        if not operator_row:
            flash('Operator tidak ditemukan!', 'error')
            return redirect(url_for('admin_users'))

        # Convert Row to dict
        operator = row_to_dict(operator_row)

        # Cek apakah relasi kegiatan ada
        cursor.execute("""
            SELECT id FROM operator_kegiatan
            WHERE user_id = ? AND kegiatan_id = ?
        """, (operator_id, kegiatan_id))
        existing = cursor.fetchone()

        if not existing:
            flash('Kegiatan tidak ditemukan untuk operator ini!', 'error')
        else:
            # Hapus relasi kegiatan
            cursor.execute("""
                DELETE FROM operator_kegiatan
                WHERE user_id = ? AND kegiatan_id = ?
            """, (operator_id, kegiatan_id))
            connection.commit()

            # Ambil nama kegiatan untuk pesan success
            cursor.execute("SELECT nama_kegiatan FROM kegiatan_master WHERE id = ?", (kegiatan_id,))
            kegiatan_row = cursor.fetchone()
            if kegiatan_row:
                kegiatan = row_to_dict(kegiatan_row)
                kegiatan_name = kegiatan.get('nama_kegiatan', 'kegiatan')
            else:
                kegiatan_name = 'kegiatan'
            operator_name = operator.get('nama') or operator.get('username')
            flash(f'Kegiatan "{kegiatan_name}" berhasil dihapus dari operator "{operator_name}"!', 'success')

    except sqlite3.Error as e:
        connection.rollback()
        flash(f'Terjadi kesalahan saat menghapus kegiatan: {str(e)}', 'error')
    finally:
        if connection:
            cursor.close()
            connection.close()

    return redirect(url_for('admin_users'))

@app.route('/admin/hapus-operator/<int:operator_id>', methods=['POST'])
@admin_required
def admin_hapus_operator(operator_id):
    """Hapus operator"""

    # Operator tidak bisa hapus operator lain
    if get_user_role() == 'operator':
        flash('Operator tidak memiliki akses!', 'error')
        return redirect(url_for('admin_dashboard'))

    connection = get_db_connection()
    if connection is None:
        flash('Koneksi database gagal!', 'error')
        # Cek referer untuk menentukan redirect
        referer = request.headers.get('Referer', '')
        if 'tambah-operator' in referer:
            return redirect(url_for('admin_tambah_operator'))
        else:
            return redirect(url_for('admin_users'))

    try:
        cursor = connection.cursor()

        # Cek apakah operator ada dan adalah operator
        cursor.execute("SELECT id, nama, username FROM users WHERE id = ? AND role = 'operator'", (operator_id,))
        operator_row = cursor.fetchone()

        if not operator_row:
            flash('Operator tidak ditemukan!', 'error')
            # Cek referer untuk menentukan redirect
            referer = request.headers.get('Referer', '')
            if 'tambah-operator' in referer:
                return redirect(url_for('admin_tambah_operator'))
            else:
                return redirect(url_for('admin_users'))

        # Convert Row to dict
        operator = row_to_dict(operator_row)
        operator_name = operator.get('nama') or operator.get('username')

        # Hapus operator (relasi operator_kegiatan akan terhapus otomatis karena CASCADE)
        cursor.execute("DELETE FROM users WHERE id = ? AND role = 'operator'", (operator_id,))
        connection.commit()

        flash(f'Operator "{operator_name}" berhasil dihapus!', 'success')

    except sqlite3.Error as e:
        connection.rollback()
        flash(f'Terjadi kesalahan saat menghapus operator: {str(e)}', 'error')
    except Exception as e:
        flash(f'Terjadi kesalahan: {str(e)}', 'error')
    finally:
        if connection:
            cursor.close()
            connection.close()

    # Cek referer untuk menentukan redirect
    referer = request.headers.get('Referer', '')
    if 'tambah-operator' in referer:
        return redirect(url_for('admin_tambah_operator'))
    else:
        return redirect(url_for('admin_users'))

@app.route('/admin/export-all-pdf/<path:nama_kegiatan>')
@admin_required
def export_all_pdf_kegiatan(nama_kegiatan):
    """Export semua biodata per kegiatan ke PDF - 1 user 1 halaman"""
    from urllib.parse import unquote

    if not is_admin():
        flash('Anda tidak memiliki akses!', 'error')
        return redirect(url_for('dashboard'))

    # Decode URL encoding
    nama_kegiatan = unquote(nama_kegiatan)
    user_id = get_user_id()
    user_role = get_user_role()

    # Ambil semua biodata untuk kegiatan ini (opsional filter kabupaten_kota via query string)
    connection = get_db_connection()
    if not connection:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_kegiatan'))

    # Jika operator, cek apakah kegiatan ini termasuk yang dia ikuti
    if user_role == 'operator':
        try:
            cursor = connection.cursor()
            cursor.execute("""
                SELECT k.id
                FROM kegiatan_master k
                INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                WHERE ok.user_id = ?
                    AND TRIM(k.nama_kegiatan) = TRIM(?)
            """, (user_id, nama_kegiatan))
            if not cursor.fetchone():
                flash('Anda tidak memiliki akses ke kegiatan ini!', 'error')
                return redirect(url_for('admin_kegiatan'))
        except sqlite3.Error as e:
            flash('Terjadi kesalahan saat memverifikasi akses!', 'error')
            return redirect(url_for('admin_kegiatan'))

    all_biodata = []
    try:
        cursor = connection.cursor()
        params = [nama_kegiatan]
        kabupaten_filter = request.args.get('kabupaten_kota', '').strip()

        base_query = """
            SELECT * FROM biodata_kegiatan
            WHERE TRIM(nama_kegiatan) = TRIM(?)
        """
        if kabupaten_filter:
            base_query += " AND TRIM(kabupaten_kota) = TRIM(?)"
            params.append(kabupaten_filter)

        base_query += " ORDER BY kabupaten_kota ASC, nama_lengkap ASC"

        # Pastikan semua kolom termasuk tanda_tangan diambil
        cursor.execute(base_query, params)
        rows = cursor.fetchall()
        all_biodata = [row_to_dict(row) for row in rows]

        if not all_biodata:
            flash('Tidak ada data untuk kegiatan ini!', 'error')
            return redirect(url_for('admin_detail_kegiatan', nama_kegiatan=nama_kegiatan))

        # Limit maksimal untuk export (5000 rows)
        MAX_EXPORT_ROWS = 5000
        if len(all_biodata) > MAX_EXPORT_ROWS:
            flash(f'Data terlalu besar ({len(all_biodata)} rows). Maksimal {MAX_EXPORT_ROWS} rows untuk export. Silakan gunakan filter yang lebih spesifik.', 'error')
            return redirect(url_for('admin_detail_kegiatan', nama_kegiatan=nama_kegiatan))

        # Warning jika data > 1000 rows
        if len(all_biodata) > 1000:
            print(f"WARNING: Export PDF dengan {len(all_biodata)} rows - mungkin memakan waktu lama")

        # Debug: Check tanda_tangan in database results
        print(f"DEBUG export_all_pdf: ===== DATABASE QUERY RESULTS =====")
        print(f"DEBUG export_all_pdf: Total biodata retrieved: {len(all_biodata)}")
        print(f"DEBUG export_all_pdf: Nama kegiatan: {nama_kegiatan}")
        for idx, biodata in enumerate(all_biodata):
            print(f"DEBUG export_all_pdf: --- Biodata {idx} ---")
            print(f"DEBUG export_all_pdf: NIK: {biodata.get('nik')}, Nama: {biodata.get('nama_lengkap')}")
            print(f"DEBUG export_all_pdf: Fields in biodata: {list(biodata.keys())}")
            print(f"DEBUG export_all_pdf: tanda_tangan field exists: {'tanda_tangan' in biodata}")
            print(f"DEBUG export_all_pdf: tanda_tangan is not None: {biodata.get('tanda_tangan') is not None}")
            if biodata.get('tanda_tangan'):
                tanda_tangan_val = biodata.get('tanda_tangan')
                print(f"DEBUG export_all_pdf: tanda_tangan type: {type(tanda_tangan_val)}")
                if isinstance(tanda_tangan_val, str):
                    print(f"DEBUG export_all_pdf: tanda_tangan length: {len(tanda_tangan_val)}")
                    if len(tanda_tangan_val) > 0:
                        print(f"DEBUG export_all_pdf: tanda_tangan starts with data:image: {tanda_tangan_val.startswith('data:image')}")
                        print(f"DEBUG export_all_pdf: tanda_tangan preview (first 100 chars): {tanda_tangan_val[:100]}")
            else:
                print(f"DEBUG export_all_pdf: WARNING: tanda_tangan is None or empty for this biodata!")

        # Normalisasi path buku tabungan untuk setiap biodata
        for biodata in all_biodata:
            if biodata.get('buku_tabungan_path'):
                biodata['buku_tabungan_path'] = normalize_buku_tabungan_path(biodata['buku_tabungan_path'])
    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan saat mengambil data: {str(e)}', 'error')
        return redirect(url_for('admin_detail_kegiatan', nama_kegiatan=nama_kegiatan))
    finally:
        if connection:
            cursor.close()
            connection.close()

    # Load logo untuk header
    logo_path = os.path.join(BASE_DIR, 'static', 'Logo_BGTK.png')
    logo_img = None
    logo_width = 0
    logo_height = 0
    if os.path.exists(logo_path):
        try:
            logo_pil = Image.open(logo_path)
            # Resize logo untuk header (max height 0.6 inch)
            max_logo_height = 0.6 * inch
            logo_ratio = logo_pil.width / logo_pil.height
            logo_width = max_logo_height * logo_ratio
            logo_height = max_logo_height

            # Convert PIL Image to reportlab Image
            logo_buffer = io.BytesIO()
            logo_pil.save(logo_buffer, format='PNG')
            logo_buffer.seek(0)
            logo_img = RLImage(logo_buffer, width=logo_width, height=logo_height)
        except Exception as e:
            print(f"Error loading logo: {e}")
            logo_img = None

    # Load logo Pendidikan Bermutu dengan background transparan
    pendidikan_bermutu_path = os.path.join(BASE_DIR, 'static', 'Pendidikan Bermutu untuk Semua.png')
    pendidikan_bermutu_processed_path = None
    pendidikan_bermutu_width = 0
    pendidikan_bermutu_height = 0
    pendidikan_bermutu_temp_file = None
    if os.path.exists(pendidikan_bermutu_path):
        try:
            pendidikan_bermutu_pil = Image.open(pendidikan_bermutu_path)
            # Convert ke RGBA jika belum
            if pendidikan_bermutu_pil.mode != 'RGBA':
                pendidikan_bermutu_pil = pendidikan_bermutu_pil.convert('RGBA')

            # Hapus background hitam (buat transparan)
            pixels = pendidikan_bermutu_pil.load()
            width, height = pendidikan_bermutu_pil.size
            black_threshold = 30  # Threshold untuk mendeteksi warna hitam

            for x in range(width):
                for y in range(height):
                    r, g, b, a = pixels[x, y]
                    # Jika pixel sangat gelap (background hitam), buat transparan
                    if r < black_threshold and g < black_threshold and b < black_threshold:
                        pixels[x, y] = (r, g, b, 0)

            # Hitung ukuran untuk header (max height 0.5 inch - diperkecil)
            # Tapi jangan resize dengan PIL, biarkan reportlab yang handle resize untuk kualitas lebih baik
            max_logo_height = 0.5 * inch
            pendidikan_bermutu_ratio = pendidikan_bermutu_pil.width / pendidikan_bermutu_pil.height
            pendidikan_bermutu_width = max_logo_height * pendidikan_bermutu_ratio
            pendidikan_bermutu_height = max_logo_height

            # Simpan gambar asli (tanpa resize) ke temporary file dengan kualitas tinggi
            # Biarkan reportlab yang handle resize dengan preserveAspectRatio untuk kualitas lebih tajam
            pendidikan_bermutu_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            pendidikan_bermutu_pil.save(pendidikan_bermutu_temp_file.name, format='PNG', optimize=False)
            pendidikan_bermutu_processed_path = pendidikan_bermutu_temp_file.name
        except Exception as e:
            print(f"Error loading logo Pendidikan Bermutu: {e}")

    # Load logo Ramah dengan background transparan
    ramah_path = os.path.join(BASE_DIR, 'static', 'Ramah.png')
    ramah_processed_path = None
    ramah_width = 0
    ramah_height = 0
    ramah_temp_file = None
    if os.path.exists(ramah_path):
        try:
            ramah_pil = Image.open(ramah_path)
            # Convert ke RGBA jika belum
            if ramah_pil.mode != 'RGBA':
                ramah_pil = ramah_pil.convert('RGBA')

            # Hapus background hitam (buat transparan)
            pixels = ramah_pil.load()
            width, height = ramah_pil.size
            black_threshold = 30  # Threshold untuk mendeteksi warna hitam

            for x in range(width):
                for y in range(height):
                    r, g, b, a = pixels[x, y]
                    # Jika pixel sangat gelap (background hitam), buat transparan
                    if r < black_threshold and g < black_threshold and b < black_threshold:
                        pixels[x, y] = (r, g, b, 0)

            # Hitung ukuran untuk header (max height 0.5 inch - diperkecil)
            # Tapi jangan resize dengan PIL, biarkan reportlab yang handle resize untuk kualitas lebih baik
            max_logo_height = 0.5 * inch
            ramah_ratio = ramah_pil.width / ramah_pil.height
            ramah_width = max_logo_height * ramah_ratio
            ramah_height = max_logo_height

            # Simpan gambar asli (tanpa resize) ke temporary file dengan kualitas tinggi
            # Biarkan reportlab yang handle resize dengan preserveAspectRatio untuk kualitas lebih tajam
            ramah_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            ramah_pil.save(ramah_temp_file.name, format='PNG', optimize=False)
            ramah_processed_path = ramah_temp_file.name
        except Exception as e:
            print(f"Error loading logo Ramah: {e}")

    # Fungsi untuk header dengan logo dan footer
    def add_header_footer(canvas, doc):
        canvas.saveState()

        # Hitung tinggi logo terbesar untuk alignment vertikal (hanya logo BGTK di header)
        max_logo_h = logo_height if logo_height > 0 else 0

        # Header - Logo BGTK di kiri
        if logo_img and logo_height > 0:
            try:
                logo_x = 25
                logo_y = F4_SIZE[1] - 25 - max_logo_h
                canvas.drawImage(logo_path, logo_x, logo_y, width=logo_width, height=logo_height, preserveAspectRatio=True)
            except Exception as e:
                print(f"Error drawing logo: {e}")

        # Garis header
        canvas.setStrokeColor(colors.HexColor('#067ac1'))
        canvas.setLineWidth(1.5)
        max_header_height = logo_height if logo_height > 0 else 0
        header_line_y = F4_SIZE[1] - 25 - max_header_height - 5
        canvas.line(25, header_line_y, F4_SIZE[0] - 25, header_line_y)

        # Footer - Logo Pendidikan Bermutu dan Ramah di kanan bawah
        max_footer_logo_height = max(pendidikan_bermutu_height if pendidikan_bermutu_height > 0 else 0,
                                    ramah_height if ramah_height > 0 else 0)
        footer_logo_y = 45  # Posisi Y untuk logo footer

        if max_footer_logo_height > 0:
            # Hitung total lebar kedua logo + spacing
            total_footer_logo_width = (pendidikan_bermutu_width if pendidikan_bermutu_height > 0 else 0) + \
                                     (ramah_width if ramah_height > 0 else 0) + 10  # 10 pt spacing antara logo
            # Posisi X untuk logo di kanan (rata kanan dengan margin 25)
            footer_logo_start_x = F4_SIZE[0] - 25 - total_footer_logo_width

            # Logo Pendidikan Bermutu di kiri (dalam footer)
            if pendidikan_bermutu_processed_path and os.path.exists(pendidikan_bermutu_processed_path) and pendidikan_bermutu_height > 0:
                try:
                    pendidikan_bermutu_footer_x = footer_logo_start_x
                    pendidikan_bermutu_footer_y = footer_logo_y
                    canvas.drawImage(pendidikan_bermutu_processed_path, pendidikan_bermutu_footer_x, pendidikan_bermutu_footer_y,
                                   width=pendidikan_bermutu_width, height=pendidikan_bermutu_height, preserveAspectRatio=True, mask='auto')
                except Exception as e:
                    print(f"Error drawing logo Pendidikan Bermutu di footer: {e}")

            # Logo Ramah di kanan (dalam footer)
            if ramah_processed_path and os.path.exists(ramah_processed_path) and ramah_height > 0:
                try:
                    ramah_footer_x = footer_logo_start_x + (pendidikan_bermutu_width if pendidikan_bermutu_height > 0 else 0) + 10
                    ramah_footer_y = footer_logo_y
                    canvas.drawImage(ramah_processed_path, ramah_footer_x, ramah_footer_y,
                                   width=ramah_width, height=ramah_height, preserveAspectRatio=True, mask='auto')
                except Exception as e:
                    print(f"Error drawing logo Ramah di footer: {e}")

        # Garis footer (di bawah logo)
        canvas.setStrokeColor(colors.HexColor('#067ac1'))
        canvas.setLineWidth(1)
        footer_line_y = footer_logo_y - 10  # Garis di bawah logo dengan spacing 10
        canvas.line(25, footer_line_y, F4_SIZE[0] - 25, footer_line_y)

        # Footer - Tanggal dan waktu export WITA (di bawah garis)
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.black)
        # WITA = UTC+8
        wita_time = datetime.utcnow() + timedelta(hours=8)
        footer_text = f"Dicetak pada: {wita_time.strftime('%d/%m/%Y %H:%M')} WITA"
        footer_width = canvas.stringWidth(footer_text, 'Helvetica', 8)
        footer_x = 25  # Rata kiri dengan margin 25 pt
        footer_y = footer_line_y - 15  # Tanggal di bawah garis
        canvas.drawString(footer_x, footer_y, footer_text)

        canvas.restoreState()

    # Gunakan kertas F4 (8.27 x 13 inch) dan beri ruang header
    buffer = io.BytesIO()
    # Gunakan tinggi logo terbesar untuk margin (hanya logo BGTK di header)
    max_header_logo_height = logo_height if logo_height > 0 else 0
    top_margin_with_logo = 25 + max_header_logo_height + 15
    doc = SimpleDocTemplate(
        buffer,
        pagesize=F4_SIZE,
        rightMargin=25,
        leftMargin=25,
        topMargin=top_margin_with_logo,
        bottomMargin=40
    )

    # Container untuk elemen PDF
    elements = []

    # Styles - lebih compact untuk muat di 1 halaman
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.black,
        spaceAfter=12,
        spaceBefore=0,
        alignment=1,  # Center
        fontName='Helvetica-Bold',
        leading=24
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.black,
        alignment=0  # Left
    )
    value_style = ParagraphStyle(
        'ValueStyle',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.black,
        alignment=0,  # Left
        leftIndent=0,
        rightIndent=0
    )

    # Field mapping untuk label yang lebih readable dan profesional
    field_labels = {
        'nik': 'NIK',
        'nama_lengkap': 'Nama Lengkap',
        'nip_nippk': 'NIP/NIPPK',
        'tempat_lahir': 'Tempat Lahir',
        'tanggal_lahir': 'Tanggal Lahir',
        'jenis_kelamin': 'Jenis Kelamin',
        'agama': 'Agama',
        'pendidikan_terakhir': 'Pendidikan Terakhir',
        'jurusan': 'Jurusan',
        'status_asn': 'Status Kepegawaian',
        'pangkat_golongan': 'Pangkat / Golongan',
        'jabatan': 'Jabatan',
        'instansi': 'Nama Instansi',
        'alamat_instansi': 'Alamat Instansi',
        'alamat_domisili': 'Alamat Domisili',
        'kabupaten_kota': 'Kabupaten/Kota',
        'kabko_lainnya': 'Kabupaten/Kota Lainnya',
        'peran': 'Peran dalam Kegiatan',
        'no_hp': 'Nomor HP',
        'alamat_email': 'Alamat Email',
        'npwp': 'NPWP',
        'nama_bank': 'Nama Bank',
        'nama_bank_lainnya': 'Nama Bank Lainnya',
        'no_rekening': 'Nomor Rekening',
        'nama_pemilik_rekening': 'Nama Pemilik Rekening'
    }

    # Urutan field yang diinginkan
    field_order = [
        'nik', 'nama_lengkap', 'nip_nippk', 'tempat_lahir', 'tanggal_lahir',
        'jenis_kelamin', 'agama', 'pendidikan_terakhir', 'jurusan',
        'status_asn', 'pangkat_golongan', 'jabatan', 'instansi',
        'alamat_instansi', 'alamat_domisili', 'kabupaten_kota', 'kabko_lainnya',
        'peran',
        'no_hp', 'alamat_email', 'npwp',
        'nama_bank', 'nama_bank_lainnya', 'no_rekening', 'nama_pemilik_rekening'
    ]

    # Exclude fields yang tidak perlu ditampilkan
    exclude_fields = ['id', 'user_id', 'buku_tabungan_path', 'tanda_tangan', 'created_at', 'updated_at', 'nama_kegiatan', 'waktu_pelaksanaan', 'tempat_pelaksanaan']

    # List untuk menyimpan temporary files tanda tangan untuk cleanup
    tanda_tangan_temp_files = []

    # Loop untuk setiap user - buat 1 halaman per user
    for user_idx, biodata in enumerate(all_biodata):
        # Tambahkan page break kecuali untuk user pertama
        if user_idx > 0:
            elements.append(PageBreak())

        # Title
        title = Paragraph("BIODATA KEGIATAN", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.15*inch))

        # Ambil data kegiatan untuk header tabel
        biodata_nama_kegiatan = biodata.get('nama_kegiatan', '-')
        waktu_pelaksanaan = biodata.get('waktu_pelaksanaan', '-')
        tempat_pelaksanaan = biodata.get('tempat_pelaksanaan', '-')

        # Buat header informasi kegiatan di atas tabel dengan titik dua yang sejajar
        if biodata_nama_kegiatan or waktu_pelaksanaan or tempat_pelaksanaan:
            info_style = ParagraphStyle(
                'InfoStyle',
                parent=styles['Normal'],
                fontSize=10,
                leading=14,
                textColor=colors.black,
                alignment=0,  # Left align
                fontName='Helvetica-Bold'
            )

            # Buat tabel kecil untuk info kegiatan agar titik dua sejajar
            # Pisahkan label dan titik dua menjadi kolom terpisah
            info_table_data = []
            if biodata_nama_kegiatan and str(biodata_nama_kegiatan).strip() and str(biodata_nama_kegiatan).strip() != '-':
                info_table_data.append([
                    Paragraph("<b>Nama Kegiatan</b>", info_style),
                    Paragraph(":", info_style),
                    Paragraph(str(biodata_nama_kegiatan), info_style)
                ])
            if waktu_pelaksanaan and str(waktu_pelaksanaan).strip() and str(waktu_pelaksanaan).strip() != '-':
                info_table_data.append([
                    Paragraph("<b>Waktu Pelaksanaan</b>", info_style),
                    Paragraph(":", info_style),
                    Paragraph(str(waktu_pelaksanaan), info_style)
                ])
            if tempat_pelaksanaan and str(tempat_pelaksanaan).strip() and str(tempat_pelaksanaan).strip() != '-':
                info_table_data.append([
                    Paragraph("<b>Tempat Pelaksanaan</b>", info_style),
                    Paragraph(":", info_style),
                    Paragraph(str(tempat_pelaksanaan), info_style)
                ])

            if info_table_data:
                # Hitung lebar yang sama dengan tabel utama agar sejajar dengan border biru
                available_width = (8.27 * inch) - (25 * 2)
                # Gunakan 3 kolom: label, titik dua, value
                # Lebar kolom disesuaikan agar titik dua sejajar dan lebar total sama dengan tabel utama
                info_table = Table(info_table_data, colWidths=[2.2*inch, 0.15*inch, available_width - 2.2*inch - 0.15*inch])
                info_table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),   # Left align untuk kolom label agar sejajar dengan border kiri
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),   # Left align untuk titik dua
                    ('ALIGN', (2, 0), (2, -1), 'LEFT'),   # Left align untuk kolom value
                    ('LEFTPADDING', (0, 0), (0, -1), 0),  # Tidak ada padding kiri untuk kolom label agar sejajar dengan border kiri tabel
                    ('LEFTPADDING', (1, 0), (1, -1), 0),  # Tidak ada padding untuk titik dua
                    ('LEFTPADDING', (2, 0), (2, -1), 0),  # Tidak ada padding untuk kolom value
                    ('RIGHTPADDING', (0, 0), (0, -1), 4),  # Padding kanan untuk kolom label
                    ('RIGHTPADDING', (1, 0), (1, -1), 0),  # Tidak ada padding untuk titik dua
                    ('RIGHTPADDING', (2, 0), (2, -1), 6),  # Padding kanan sama dengan tabel utama
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    # Border dihilangkan sesuai permintaan
                ]))
                elements.append(info_table)
                elements.append(Spacer(1, 0.1*inch))

        # Semua data dalam 1 tabel
        all_data = []

        # Collect all data in order
        for key in field_order:
            if key in biodata and key not in exclude_fields:
                label = field_labels.get(key, key.replace('_', ' ').title())
                value = biodata[key]
                display_value = str(value) if value and str(value).strip() else '-'
                all_data.append([label, display_value])

        # Add any remaining fields not in field_order
        for key, value in biodata.items():
            if key not in exclude_fields and key not in field_order:
                label = field_labels.get(key, key.replace('_', ' ').title())
                display_value = str(value) if value and str(value).strip() else '-'
                all_data.append([label, display_value])

        # Process tanda tangan menggunakan helper function
        nik_user = biodata.get('nik', 'N/A')
        nama_user = biodata.get('nama_lengkap', 'N/A')
        print(f"DEBUG export_all_pdf: ===== Processing tanda tangan for user {user_idx} =====")
        print(f"DEBUG export_all_pdf: NIK: {nik_user}, Nama: {nama_user}")

        tanda_tangan_raw = biodata.get('tanda_tangan')
        print(f"DEBUG export_all_pdf: Tanda tangan exists in biodata: {tanda_tangan_raw is not None}")
        if tanda_tangan_raw:
            print(f"DEBUG export_all_pdf: Tanda tangan type: {type(tanda_tangan_raw)}")
            if isinstance(tanda_tangan_raw, str):
                print(f"DEBUG export_all_pdf: Tanda tangan length: {len(tanda_tangan_raw)}")
                print(f"DEBUG export_all_pdf: Tanda tangan preview: {tanda_tangan_raw[:50]}...")

        tanda_tangan_img, error_msg = process_tanda_tangan_for_pdf(tanda_tangan_raw, tanda_tangan_temp_files)
        if error_msg:
            print(f"DEBUG export_all_pdf ERROR: {error_msg}")
        else:
            print(f"DEBUG export_all_pdf: Tanda tangan processed successfully")
        print(f"DEBUG export_all_pdf: Tanda tangan image object: {tanda_tangan_img is not None}")

        # Buat 1 tabel untuk semua data - compact untuk muat di 1 halaman
        if all_data:
            # Prepare table data with Paragraph for text wrapping
            table_data = []
            # Header dihapus - langsung ke data rows
            # Data rows dengan Paragraph untuk text wrapping
            for row in all_data:
                field_label = Paragraph(f"<b>{row[0]}</b>", value_style)
                value_text = str(row[1]) if row[1] else '-'
                value_text = value_text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                if len(value_text) > 50:
                    words = value_text.split()
                    lines = []
                    current_line = []
                    current_length = 0
                    for word in words:
                        if current_length + len(word) + 1 > 50:
                            if current_line:
                                lines.append(' '.join(current_line))
                            current_line = [word]
                            current_length = len(word)
                        else:
                            current_line.append(word)
                            current_length += len(word) + 1
                    if current_line:
                        lines.append(' '.join(current_line))
                    value_text = '<br/>'.join(lines)
                value_para = Paragraph(value_text, value_style)
                table_data.append([field_label, value_para])

            # Tambahkan tanda tangan ke tabel jika ada
            print(f"DEBUG export_all_pdf: Adding tanda tangan to table...")
            print(f"DEBUG export_all_pdf: tanda_tangan_img is not None: {tanda_tangan_img is not None}")
            print(f"DEBUG export_all_pdf: biodata has tanda_tangan: {biodata.get('tanda_tangan') is not None}")

            if tanda_tangan_img:
                print(f"DEBUG export_all_pdf: Adding tanda tangan image to table")
                field_label = Paragraph("<b>Tanda Tangan</b>", value_style)
                table_data.append([field_label, tanda_tangan_img])
                print(f"DEBUG export_all_pdf: Tanda tangan added to table successfully")
            elif biodata.get('tanda_tangan'):
                # Jika ada data tanda tangan tapi gagal diproses, tampilkan pesan
                print(f"DEBUG export_all_pdf: Adding error message to table (tanda tangan exists but failed to process)")
                field_label = Paragraph("<b>Tanda Tangan</b>", value_style)
                error_msg = Paragraph("<i>Tanda tangan tidak tersedia</i>", value_style)
                table_data.append([field_label, error_msg])
            else:
                print(f"DEBUG export_all_pdf: No tanda tangan data, skipping table row")

            # Calculate table width (F4 width minus margins)
            available_width = (8.27 * inch) - (25 * 2)
            table = Table(table_data, colWidths=[2.2*inch, available_width - 2.2*inch])
            table.setStyle(TableStyle([
                # Tidak ada header, langsung ke data rows
                ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                # Padding lebih kecil agar teks rata lebih kiri
                ('LEFTPADDING', (0, 0), (0, -1), 4),
                ('LEFTPADDING', (1, 0), (1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
                ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#067ac1')),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.08*inch))

    # Build PDF
    try:
        doc.build(elements, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
        buffer.seek(0)

        # Refresh session sebelum return response untuk mencegah logout
        session.permanent = True
        session.modified = True

        # Return PDF
        return Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename=biodata_{nama_kegiatan.replace(" ", "_")}.pdf'
            }
        )
    finally:
        # Cleanup temporary files
        try:
            if pendidikan_bermutu_temp_file and os.path.exists(pendidikan_bermutu_processed_path):
                os.unlink(pendidikan_bermutu_processed_path)
        except Exception as e:
            print(f"Error cleaning up Pendidikan Bermutu temp file: {e}")

        try:
            if ramah_temp_file and os.path.exists(ramah_processed_path):
                os.unlink(ramah_processed_path)
        except Exception as e:
            print(f"Error cleaning up Ramah temp file: {e}")

        # Cleanup temporary files tanda tangan
        for temp_file_path in tanda_tangan_temp_files:
            try:
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)
            except Exception as e:
                print(f"Error cleaning up tanda tangan temp file {temp_file_path}: {e}")

@app.route('/admin/export-all-excel/<path:nama_kegiatan>')
@admin_required
def export_all_excel_kegiatan(nama_kegiatan):
    """Export semua biodata per kegiatan ke Excel - dengan semua field detail lengkap"""
    from urllib.parse import unquote
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not is_admin():
        flash('Anda tidak memiliki akses!', 'error')
        return redirect(url_for('dashboard'))

    # Decode URL encoding
    nama_kegiatan = unquote(nama_kegiatan)
    user_id = get_user_id()
    user_role = get_user_role()

    # Ambil semua biodata untuk kegiatan ini (opsional filter kabupaten_kota via query string)
    connection = get_db_connection()
    if not connection:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_kegiatan'))

    # Jika operator, cek apakah kegiatan ini termasuk yang dia ikuti
    if user_role == 'operator':
        try:
            cursor = connection.cursor()
            cursor.execute("""
                SELECT k.id
                FROM kegiatan_master k
                INNER JOIN operator_kegiatan ok ON k.id = ok.kegiatan_id
                WHERE ok.user_id = ?
                    AND TRIM(k.nama_kegiatan) = TRIM(?)
            """, (user_id, nama_kegiatan))
            if not cursor.fetchone():
                flash('Anda tidak memiliki akses ke kegiatan ini!', 'error')
                return redirect(url_for('admin_kegiatan'))
        except sqlite3.Error as e:
            flash('Terjadi kesalahan saat memverifikasi akses!', 'error')
            return redirect(url_for('admin_kegiatan'))

    all_biodata = []
    try:
        cursor = connection.cursor()
        params = [nama_kegiatan]
        kabupaten_filter = request.args.get('kabupaten_kota', '').strip()

        base_query = """
            SELECT * FROM biodata_kegiatan
            WHERE TRIM(nama_kegiatan) = TRIM(?)
        """
        if kabupaten_filter:
            base_query += " AND TRIM(kabupaten_kota) = TRIM(?)"
            params.append(kabupaten_filter)

        base_query += " ORDER BY kabupaten_kota ASC, nama_lengkap ASC"

        cursor.execute(base_query, params)
        rows = cursor.fetchall()
        all_biodata = [row_to_dict(row) for row in rows]

        if not all_biodata:
            flash('Tidak ada data untuk kegiatan ini!', 'error')
            return redirect(url_for('admin_detail_kegiatan', nama_kegiatan=nama_kegiatan))

        # Limit maksimal untuk export (5000 rows)
        MAX_EXPORT_ROWS = 5000
        if len(all_biodata) > MAX_EXPORT_ROWS:
            flash(f'Data terlalu besar ({len(all_biodata)} rows). Maksimal {MAX_EXPORT_ROWS} rows untuk export. Silakan gunakan filter yang lebih spesifik.', 'error')
            return redirect(url_for('admin_detail_kegiatan', nama_kegiatan=nama_kegiatan))

        # Warning jika data > 1000 rows
        if len(all_biodata) > 1000:
            print(f"WARNING: Export Excel dengan {len(all_biodata)} rows - mungkin memakan waktu lama")

        # Normalisasi path buku tabungan untuk setiap biodata
        for biodata in all_biodata:
            if biodata.get('buku_tabungan_path'):
                biodata['buku_tabungan_path'] = normalize_buku_tabungan_path(biodata['buku_tabungan_path'])
    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan saat mengambil data: {str(e)}', 'error')
        return redirect(url_for('admin_detail_kegiatan', nama_kegiatan=nama_kegiatan))
    finally:
        if connection:
            cursor.close()
            connection.close()

    # Field mapping untuk label yang lebih readable dan profesional
    field_labels = {
        'nik': 'NIK',
        'nama_lengkap': 'Nama Lengkap',
        'nip_nippk': 'NIP/NIPPK',
        'tempat_lahir': 'Tempat Lahir',
        'tanggal_lahir': 'Tanggal Lahir',
        'jenis_kelamin': 'Jenis Kelamin',
        'agama': 'Agama',
        'pendidikan_terakhir': 'Pendidikan Terakhir',
        'jurusan': 'Jurusan',
        'status_asn': 'Status Kepegawaian',
        'pangkat_golongan': 'Pangkat / Golongan',
        'jabatan': 'Jabatan',
        'instansi': 'Nama Instansi',
        'alamat_instansi': 'Alamat Instansi',
        'alamat_domisili': 'Alamat Domisili',
        'kabupaten_kota': 'Kabupaten/Kota',
        'kabko_lainnya': 'Kabupaten/Kota Lainnya',
        'peran': 'Peran dalam Kegiatan',
        'no_hp': 'Nomor HP',
        'alamat_email': 'Alamat Email',
        'npwp': 'NPWP',
        'nama_bank': 'Nama Bank',
        'nama_bank_lainnya': 'Nama Bank Lainnya',
        'no_rekening': 'Nomor Rekening',
        'nama_pemilik_rekening': 'Nama Pemilik Rekening'
    }

    # Urutan field yang diinginkan
    field_order = [
        'nik', 'nama_lengkap', 'nip_nippk', 'tempat_lahir', 'tanggal_lahir',
        'jenis_kelamin', 'agama', 'pendidikan_terakhir', 'jurusan',
        'status_asn', 'pangkat_golongan', 'jabatan', 'instansi',
        'alamat_instansi', 'alamat_domisili', 'kabupaten_kota', 'kabko_lainnya',
        'peran',
        'no_hp', 'alamat_email', 'npwp',
        'nama_bank', 'nama_bank_lainnya', 'no_rekening', 'nama_pemilik_rekening'
    ]

    # Exclude fields yang tidak perlu ditampilkan
    exclude_fields = ['id', 'user_id', 'buku_tabungan_path', 'tanda_tangan', 'created_at', 'updated_at', 'nama_kegiatan', 'waktu_pelaksanaan', 'tempat_pelaksanaan']

    # Buat workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Biodata"

    # Styles
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_font = Font(bold=True, color="000000", size=11)
    title_font = Font(bold=True, size=12, color="000000")
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Buat list semua field yang akan ditampilkan (dalam urutan yang diinginkan)
    all_fields = []
    for key in field_order:
        if key not in exclude_fields:
            label = field_labels.get(key, key.replace('_', ' ').title())
            all_fields.append({'key': key, 'label': label})

    # Tambahkan field tanda tangan
    all_fields.append({'key': 'tanda_tangan', 'label': 'Tanda Tangan'})

    # Ambil waktu dan tempat pelaksanaan dari data pertama (semua biodata dalam satu kegiatan memiliki waktu dan tempat yang sama)
    nama_kegiatan = all_biodata[0].get('nama_kegiatan', '-') if all_biodata else '-'
    waktu_pelaksanaan = all_biodata[0].get('waktu_pelaksanaan', '-') if all_biodata else '-'
    tempat_pelaksanaan = all_biodata[0].get('tempat_pelaksanaan', '-') if all_biodata else '-'

    # Header kegiatan
    num_cols = len(all_fields)
    last_col_letter = get_column_letter(num_cols)

    # Gunakan 2 kolom: kolom A untuk label (diperlebar), kolom B sampai akhir untuk titik dua + value (di-merge)
    # Ini memungkinkan label tidak terpotong dan value tidak terpotong
    ws['A1'] = "NAMA KEGIATAN"
    ws['B1'] = f": {nama_kegiatan}"
    ws['A2'] = "WAKTU PELAKSANAAN"
    ws['B2'] = f": {waktu_pelaksanaan}"
    ws['A3'] = "TEMPAT PELAKSANAAN"
    ws['B3'] = f": {tempat_pelaksanaan}"

    # Merge kolom B sampai akhir untuk titik dua + value
    ws.merge_cells(f'B1:{last_col_letter}1')
    ws.merge_cells(f'B2:{last_col_letter}2')
    ws.merge_cells(f'B3:{last_col_letter}3')

    # Set font, alignment, dan fill untuk setiap cell header
    # Buat font untuk header dengan warna hitam dan ukuran lebih kecil
    header_title_font = Font(bold=True, size=11, color="000000")
    for row in [1, 2, 3]:
        # Cell label (kolom A) - rata kiri
        label_cell = ws[f'A{row}']
        label_cell.font = header_title_font
        label_cell.alignment = Alignment(horizontal='left', vertical='center')
        label_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Cell titik dua + value (kolom B yang di-merge) - rata kiri
        value_cell = ws[f'B{row}']
        value_cell.font = header_title_font
        value_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        value_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Tambahkan informasi tanggal export (definisikan dulu sebelum digunakan)
    from datetime import datetime
    export_date = datetime.now().strftime("%d %B %Y, %H:%M:%S")

    # Set lebar kolom A untuk label - diperlebar agar tidak terpotong
    # Hitung lebar optimal berdasarkan konten di row 1-3 (header kegiatan)
    max_label_length = max(
        len("NAMA KEGIATAN"),
        len("WAKTU PELAKSANAAN"),
        len("TEMPAT PELAKSANAAN"),
        len(f"Tanggal Export: {export_date}")
    )
    # Set width berdasarkan konten dengan padding, minimal 25, maksimal 35
    col_a_width = max(20, min(max_label_length + 3, 23))
    ws.column_dimensions['A'].width = col_a_width
    # Tinggi baris header dibuat lebih tinggi agar teks panjang tidak terpotong
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 32
    ws['A4'] = f"Tanggal Export: {export_date}"
    ws.merge_cells(f'A4:{last_col_letter}4')
    ws['A4'].font = Font(size=10, italic=True)
    ws['A4'].alignment = left_align
    ws.row_dimensions[4].height = 18

    # Buat header kolom (semua field sebagai kolom)
    for col_idx, field_info in enumerate(all_fields, 1):
        cell = ws.cell(row=5, column=col_idx, value=field_info['label'])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_style
    ws.row_dimensions[5].height = 22

    # Set lebar semua kolom menjadi 25 agar konsisten
    for col_idx in range(1, len(all_fields) + 1):
        col_letter = get_column_letter(col_idx)
        # Skip kolom A - sudah di-set sebelumnya dengan width yang lebih lebar
        if col_letter != 'A':
            ws.column_dimensions[col_letter].width = 25

    # Data untuk setiap user (setiap user = 1 baris)
    current_row = 6
    for user_idx, biodata in enumerate(all_biodata):
        for col_idx, field_info in enumerate(all_fields, 1):
            field_key = field_info['key']

            # Ambil nilai field
            if field_key == 'tanda_tangan':
                display_value = "[Tersedia]" if biodata.get('tanda_tangan') else "-"
            else:
                value = biodata.get(field_key, '')
                display_value = str(value) if value and str(value).strip() else '-'

            # Tulis ke cell
            cell = ws.cell(row=current_row, column=col_idx, value=display_value)
            cell.border = border_style

            # Alternating row colors untuk readability
            if user_idx % 2 == 0:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # Alignment dengan wrap text untuk semua kolom agar teks tidak terpotong
            if field_key in ['nik', 'nip_nippk', 'no_hp', 'npwp', 'no_rekening']:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            elif field_key in ['tanggal_lahir']:
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            else:
                # Semua kolom lain juga menggunakan wrap_text
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Auto adjust row height berdasarkan konten
        max_height = 18  # Minimum height
        for col_idx, field_info in enumerate(all_fields, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            if cell.value:
                # Estimasi tinggi berdasarkan panjang teks dan lebar kolom
                col_letter = get_column_letter(col_idx)
                col_width = ws.column_dimensions[col_letter].width or 25
                text_length = len(str(cell.value))
                # Estimasi: setiap ~col_width karakter = 1 baris, setiap baris = 15 points
                estimated_lines = max(1, (text_length / max(col_width * 0.8, 1)))
                estimated_height = estimated_lines * 15
                max_height = max(max_height, min(estimated_height, 60))  # Max 60 points

        ws.row_dimensions[current_row].height = max_height
        current_row += 1

    # Auto-adjust column widths berdasarkan konten aktual (header dan data)
    for col_idx, field_info in enumerate(all_fields, 1):
        col_letter = get_column_letter(col_idx)

        # Skip kolom A - sudah di-set sebelumnya dan tidak perlu di-adjust
        if col_letter == 'A':
            continue

        max_length = 0

        # Cek panjang header
        header_cell = ws.cell(row=5, column=col_idx)
        if header_cell.value:
            max_length = len(str(header_cell.value))

        # Cek panjang konten data di kolom ini
        for row in range(6, current_row):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value:
                # Untuk teks panjang, ambil panjang maksimal per baris (dengan wrap text)
                text = str(cell.value)
                # Estimasi: jika teks panjang, mungkin akan wrap, jadi ambil panjang maksimal
                # Tapi untuk perhitungan lebar, gunakan panjang total dibagi estimasi wrap
                text_length = len(text)
                # Jika teks sangat panjang, estimasi lebar berdasarkan wrap
                if text_length > 50:
                    # Estimasi: setiap 15 karakter = 1 baris, ambil panjang maksimal per baris
                    lines = text.split('\n') if '\n' in text else [text]
                    max_line_length = max(len(line) for line in lines)
                    max_length = max(max_length, min(max_line_length, 40))  # Maksimal 40 karakter per baris
                else:
                    max_length = max(max_length, text_length)

        # Set lebar kolom: panjang maksimal + 2 (padding), minimal 12, maksimal 40
        optimal_width = max(12, min(max_length + 2, 40))
        # Untuk kolom alamat, boleh lebih lebar
        if 'Alamat' in field_info['label']:
            optimal_width = max(optimal_width, min(max_length + 2, 45))
        ws.column_dimensions[col_letter].width = optimal_width

    # Pastikan kolom A tetap memiliki width yang cukup (jika ter-overwrite, set ulang)
    if ws.column_dimensions['A'].width < 20:
        ws.column_dimensions['A'].width = col_a_width

    # Freeze panes untuk memudahkan scrolling (freeze header)
    ws.freeze_panes = 'A6'

    # Enable auto filter untuk header
    ws.auto_filter.ref = f'A5:{last_col_letter}{current_row - 1}'

    # Set print settings untuk lebih rapi saat print
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # Save to BytesIO
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Refresh session sebelum return response untuk mencegah logout
    session.permanent = True
    session.modified = True

    # Return Excel file
    return Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename=biodata_{nama_kegiatan.replace(" ", "_")}.xlsx'
        }
    )

@app.route('/admin/export-pdf/<nik>/<path:nama_kegiatan>')
@admin_required
def export_biodata_pdf(nik, nama_kegiatan):
    """Export biodata ke PDF"""
    from urllib.parse import unquote

    if not is_admin():
        flash('Anda tidak memiliki akses!', 'error')
        return redirect(url_for('dashboard'))

    # Decode URL encoding
    nik = unquote(nik)
    nama_kegiatan = unquote(nama_kegiatan)

    # Ambil data biodata dari database
    connection = get_db_connection()
    if not connection:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('admin_kegiatan'))

    biodata = None
    try:
        cursor = connection.cursor()
        # Pastikan semua kolom termasuk tanda_tangan diambil
        cursor.execute("""
            SELECT * FROM biodata_kegiatan
            WHERE nik = ? AND TRIM(nama_kegiatan) = TRIM(?)
            LIMIT 1
        """, (nik, nama_kegiatan))
        biodata_row = cursor.fetchone()

        if not biodata_row:
            flash('Data tidak ditemukan!', 'error')
            return redirect(url_for('admin_kegiatan'))

        # Convert Row to dict
        biodata = row_to_dict(biodata_row)

        # Normalisasi path buku tabungan
        if biodata.get('buku_tabungan_path'):
            biodata['buku_tabungan_path'] = normalize_buku_tabungan_path(biodata['buku_tabungan_path'])
    except sqlite3.Error as e:
        flash(f'Terjadi kesalahan saat mengambil data: {str(e)}', 'error')
        return redirect(url_for('admin_kegiatan'))
    finally:
        if connection:
            cursor.close()
            connection.close()

    # Load logo untuk header
    logo_path = os.path.join(BASE_DIR, 'static', 'Logo_BGTK.png')
    logo_img = None
    logo_width = 0
    logo_height = 0
    if os.path.exists(logo_path):
        try:
            logo_pil = Image.open(logo_path)
            # Resize logo untuk header (max height 0.6 inch)
            max_logo_height = 0.6 * inch
            logo_ratio = logo_pil.width / logo_pil.height
            logo_width = max_logo_height * logo_ratio
            logo_height = max_logo_height

            # Convert PIL Image to reportlab Image
            logo_buffer = io.BytesIO()
            logo_pil.save(logo_buffer, format='PNG')
            logo_buffer.seek(0)
            logo_img = RLImage(logo_buffer, width=logo_width, height=logo_height)
        except Exception as e:
            print(f"Error loading logo: {e}")
            logo_img = None

    # Load logo Pendidikan Bermutu dengan background transparan
    pendidikan_bermutu_path = os.path.join(BASE_DIR, 'static', 'Pendidikan Bermutu untuk Semua.png')
    pendidikan_bermutu_processed_path = None
    pendidikan_bermutu_width = 0
    pendidikan_bermutu_height = 0
    pendidikan_bermutu_temp_file = None
    if os.path.exists(pendidikan_bermutu_path):
        try:
            pendidikan_bermutu_pil = Image.open(pendidikan_bermutu_path)
            # Convert ke RGBA jika belum
            if pendidikan_bermutu_pil.mode != 'RGBA':
                pendidikan_bermutu_pil = pendidikan_bermutu_pil.convert('RGBA')

            # Hapus background hitam (buat transparan)
            pixels = pendidikan_bermutu_pil.load()
            width, height = pendidikan_bermutu_pil.size
            black_threshold = 30  # Threshold untuk mendeteksi warna hitam

            for x in range(width):
                for y in range(height):
                    r, g, b, a = pixels[x, y]
                    # Jika pixel sangat gelap (background hitam), buat transparan
                    if r < black_threshold and g < black_threshold and b < black_threshold:
                        pixels[x, y] = (r, g, b, 0)

            # Hitung ukuran untuk header (max height 0.5 inch - diperkecil)
            # Tapi jangan resize dengan PIL, biarkan reportlab yang handle resize untuk kualitas lebih baik
            max_logo_height = 0.5 * inch
            pendidikan_bermutu_ratio = pendidikan_bermutu_pil.width / pendidikan_bermutu_pil.height
            pendidikan_bermutu_width = max_logo_height * pendidikan_bermutu_ratio
            pendidikan_bermutu_height = max_logo_height

            # Simpan gambar asli (tanpa resize) ke temporary file dengan kualitas tinggi
            # Biarkan reportlab yang handle resize dengan preserveAspectRatio untuk kualitas lebih tajam
            pendidikan_bermutu_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            pendidikan_bermutu_pil.save(pendidikan_bermutu_temp_file.name, format='PNG', optimize=False)
            pendidikan_bermutu_processed_path = pendidikan_bermutu_temp_file.name
        except Exception as e:
            print(f"Error loading logo Pendidikan Bermutu: {e}")

    # Load logo Ramah dengan background transparan
    ramah_path = os.path.join(BASE_DIR, 'static', 'Ramah.png')
    ramah_processed_path = None
    ramah_width = 0
    ramah_height = 0
    ramah_temp_file = None
    if os.path.exists(ramah_path):
        try:
            ramah_pil = Image.open(ramah_path)
            # Convert ke RGBA jika belum
            if ramah_pil.mode != 'RGBA':
                ramah_pil = ramah_pil.convert('RGBA')

            # Hapus background hitam (buat transparan)
            pixels = ramah_pil.load()
            width, height = ramah_pil.size
            black_threshold = 30  # Threshold untuk mendeteksi warna hitam

            for x in range(width):
                for y in range(height):
                    r, g, b, a = pixels[x, y]
                    # Jika pixel sangat gelap (background hitam), buat transparan
                    if r < black_threshold and g < black_threshold and b < black_threshold:
                        pixels[x, y] = (r, g, b, 0)

            # Hitung ukuran untuk header (max height 0.5 inch - diperkecil)
            # Tapi jangan resize dengan PIL, biarkan reportlab yang handle resize untuk kualitas lebih baik
            max_logo_height = 0.5 * inch
            ramah_ratio = ramah_pil.width / ramah_pil.height
            ramah_width = max_logo_height * ramah_ratio
            ramah_height = max_logo_height

            # Simpan gambar asli (tanpa resize) ke temporary file dengan kualitas tinggi
            # Biarkan reportlab yang handle resize dengan preserveAspectRatio untuk kualitas lebih tajam
            ramah_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            ramah_pil.save(ramah_temp_file.name, format='PNG', optimize=False)
            ramah_processed_path = ramah_temp_file.name
        except Exception as e:
            print(f"Error loading logo Ramah: {e}")

    # Fungsi untuk header dengan logo dan footer
    def add_header_footer(canvas, doc):
        canvas.saveState()

        # Hitung tinggi logo terbesar untuk alignment vertikal (hanya logo BGTK di header)
        max_logo_h = logo_height if logo_height > 0 else 0

        # Header - Logo BGTK di kiri
        if logo_img and logo_height > 0:
            try:
                logo_x = 25
                logo_y = F4_SIZE[1] - 25 - max_logo_h
                canvas.drawImage(logo_path, logo_x, logo_y, width=logo_width, height=logo_height, preserveAspectRatio=True)
            except Exception as e:
                print(f"Error drawing logo: {e}")

        # Garis header
        canvas.setStrokeColor(colors.HexColor('#067ac1'))
        canvas.setLineWidth(1.5)
        max_header_height = logo_height if logo_height > 0 else 0
        header_line_y = F4_SIZE[1] - 25 - max_header_height - 5
        canvas.line(25, header_line_y, F4_SIZE[0] - 25, header_line_y)

        # Footer - Logo Pendidikan Bermutu dan Ramah di kanan bawah
        max_footer_logo_height = max(pendidikan_bermutu_height if pendidikan_bermutu_height > 0 else 0,
                                    ramah_height if ramah_height > 0 else 0)
        footer_logo_y = 45  # Posisi Y untuk logo footer

        if max_footer_logo_height > 0:
            # Hitung total lebar kedua logo + spacing
            total_footer_logo_width = (pendidikan_bermutu_width if pendidikan_bermutu_height > 0 else 0) + \
                                     (ramah_width if ramah_height > 0 else 0) + 10  # 10 pt spacing antara logo
            # Posisi X untuk logo di kanan (rata kanan dengan margin 25)
            footer_logo_start_x = F4_SIZE[0] - 25 - total_footer_logo_width

            # Logo Pendidikan Bermutu di kiri (dalam footer)
            if pendidikan_bermutu_processed_path and os.path.exists(pendidikan_bermutu_processed_path) and pendidikan_bermutu_height > 0:
                try:
                    pendidikan_bermutu_footer_x = footer_logo_start_x
                    pendidikan_bermutu_footer_y = footer_logo_y
                    canvas.drawImage(pendidikan_bermutu_processed_path, pendidikan_bermutu_footer_x, pendidikan_bermutu_footer_y,
                                   width=pendidikan_bermutu_width, height=pendidikan_bermutu_height, preserveAspectRatio=True, mask='auto')
                except Exception as e:
                    print(f"Error drawing logo Pendidikan Bermutu di footer: {e}")

            # Logo Ramah di kanan (dalam footer)
            if ramah_processed_path and os.path.exists(ramah_processed_path) and ramah_height > 0:
                try:
                    ramah_footer_x = footer_logo_start_x + (pendidikan_bermutu_width if pendidikan_bermutu_height > 0 else 0) + 10
                    ramah_footer_y = footer_logo_y
                    canvas.drawImage(ramah_processed_path, ramah_footer_x, ramah_footer_y,
                                   width=ramah_width, height=ramah_height, preserveAspectRatio=True, mask='auto')
                except Exception as e:
                    print(f"Error drawing logo Ramah di footer: {e}")

        # Garis footer (di bawah logo)
        canvas.setStrokeColor(colors.HexColor('#067ac1'))
        canvas.setLineWidth(1)
        footer_line_y = footer_logo_y - 10  # Garis di bawah logo dengan spacing 10
        canvas.line(25, footer_line_y, F4_SIZE[0] - 25, footer_line_y)

        # Footer - Tanggal dan waktu export WITA (di bawah garis)
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.black)
        # WITA = UTC+8
        wita_time = datetime.utcnow() + timedelta(hours=8)
        footer_text = f"Dicetak pada: {wita_time.strftime('%d/%m/%Y %H:%M')} WITA"
        footer_width = canvas.stringWidth(footer_text, 'Helvetica', 8)
        footer_x = 25  # Rata kiri dengan margin 25 pt
        footer_y = footer_line_y - 15  # Tanggal di bawah garis
        canvas.drawString(footer_x, footer_y, footer_text)

        canvas.restoreState()

    # Gunakan kertas F4 (8.27 x 13 inch) dan beri ruang header
    buffer = io.BytesIO()
    # Gunakan tinggi logo terbesar untuk margin (hanya logo BGTK di header)
    max_header_logo_height = logo_height if logo_height > 0 else 0
    top_margin_with_logo = 25 + max_header_logo_height + 15
    doc = SimpleDocTemplate(
        buffer,
        pagesize=F4_SIZE,
        rightMargin=25,
        leftMargin=25,
        topMargin=top_margin_with_logo,
        bottomMargin=40
    )

    # Container untuk elemen PDF
    elements = []

    # Styles - lebih compact untuk muat di 1 halaman
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.black,
        spaceAfter=12,
        spaceBefore=0,
        alignment=1,  # Center
        fontName='Helvetica-Bold',
        leading=24
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.black,
        alignment=0  # Left
    )
    value_style = ParagraphStyle(
        'ValueStyle',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.black,
        alignment=0,  # Left
        leftIndent=0,
        rightIndent=0
    )

    # Title
    title = Paragraph("BIODATA KEGIATAN", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.15*inch))

    # Ambil data kegiatan untuk header tabel
    nama_kegiatan = biodata.get('nama_kegiatan', '-')
    waktu_pelaksanaan = biodata.get('waktu_pelaksanaan', '-')
    tempat_pelaksanaan = biodata.get('tempat_pelaksanaan', '-')

    # Buat header informasi kegiatan di atas tabel dengan titik dua yang sejajar
    if nama_kegiatan or waktu_pelaksanaan or tempat_pelaksanaan:
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            leading=14,
            textColor=colors.black,
            alignment=0,  # Left align
            fontName='Helvetica-Bold'
        )

        # Buat tabel kecil untuk info kegiatan agar titik dua sejajar
        # Pisahkan label dan titik dua menjadi kolom terpisah
        info_table_data = []
        if nama_kegiatan and str(nama_kegiatan).strip() and str(nama_kegiatan).strip() != '-':
            info_table_data.append([
                Paragraph("<b>Nama Kegiatan</b>", info_style),
                Paragraph(":", info_style),
                Paragraph(str(nama_kegiatan), info_style)
            ])
        if waktu_pelaksanaan and str(waktu_pelaksanaan).strip() and str(waktu_pelaksanaan).strip() != '-':
            info_table_data.append([
                Paragraph("<b>Waktu Pelaksanaan</b>", info_style),
                Paragraph(":", info_style),
                Paragraph(str(waktu_pelaksanaan), info_style)
            ])
        if tempat_pelaksanaan and str(tempat_pelaksanaan).strip() and str(tempat_pelaksanaan).strip() != '-':
            info_table_data.append([
                Paragraph("<b>Tempat Pelaksanaan</b>", info_style),
                Paragraph(":", info_style),
                Paragraph(str(tempat_pelaksanaan), info_style)
            ])

        if info_table_data:
            # Hitung lebar yang sama dengan tabel utama agar sejajar dengan border biru
            available_width = (8.27 * inch) - (25 * 2)
            # Gunakan 3 kolom: label, titik dua, value
            # Lebar kolom disesuaikan agar titik dua sejajar dan lebar total sama dengan tabel utama
            info_table = Table(info_table_data, colWidths=[2.2*inch, 0.15*inch, available_width - 2.2*inch - 0.15*inch])
            info_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),   # Left align untuk kolom label agar sejajar dengan border kiri
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),   # Left align untuk titik dua
                ('ALIGN', (2, 0), (2, -1), 'LEFT'),   # Left align untuk kolom value
                ('LEFTPADDING', (0, 0), (0, -1), 0),  # Tidak ada padding kiri untuk kolom label agar sejajar dengan border kiri tabel
                ('LEFTPADDING', (1, 0), (1, -1), 0),  # Tidak ada padding untuk titik dua
                ('LEFTPADDING', (2, 0), (2, -1), 0),  # Tidak ada padding untuk kolom value
                ('RIGHTPADDING', (0, 0), (0, -1), 4),  # Padding kanan untuk kolom label
                ('RIGHTPADDING', (1, 0), (1, -1), 0),  # Tidak ada padding untuk titik dua
                ('RIGHTPADDING', (2, 0), (2, -1), 6),  # Padding kanan sama dengan tabel utama
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                # Border dihilangkan sesuai permintaan
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 0.1*inch))

    # Semua data dalam 1 tabel
    all_data = []

    # Field mapping untuk label yang lebih readable - urut sesuai prioritas
    field_labels = {
        'nik': 'NIK',
        'nama_lengkap': 'Nama Lengkap',
        'nip_nippk': 'NIP/NIPPK',
        'tempat_lahir': 'Tempat Lahir',
        'tanggal_lahir': 'Tanggal Lahir',
        'jenis_kelamin': 'Jenis Kelamin',
        'agama': 'Agama',
        'pendidikan_terakhir': 'Pendidikan Terakhir',
        'jurusan': 'Jurusan',
        'status_asn': 'Status ASN',
        'pangkat_golongan': 'Pangkat/Golongan',
        'jabatan': 'Jabatan',
        'instansi': 'Instansi',
        'alamat_instansi': 'Alamat Instansi',
        'alamat_domisili': 'Alamat Domisili',
        'kabupaten_kota': 'Kabupaten/Kota',
        'kabko_lainnya': 'Kabupaten/Kota Lainnya',
        'peran': 'Peran',
        'no_hp': 'No. HP',
        'alamat_email': 'Email',
        'npwp': 'NPWP',
        'nama_bank': 'Nama Bank',
        'nama_bank_lainnya': 'Nama Bank Lainnya',
        'no_rekening': 'No. Rekening',
        'nama_pemilik_rekening': 'Nama Pemilik Rekening'
    }

    # Urutan field yang diinginkan (tanpa nama_kegiatan, waktu_pelaksanaan, tempat_pelaksanaan)
    field_order = [
        'nik', 'nama_lengkap', 'nip_nippk', 'tempat_lahir', 'tanggal_lahir',
        'jenis_kelamin', 'agama', 'pendidikan_terakhir', 'jurusan',
        'status_asn', 'pangkat_golongan', 'jabatan', 'instansi',
        'alamat_instansi', 'alamat_domisili', 'kabupaten_kota', 'kabko_lainnya',
        'peran',
        'no_hp', 'alamat_email', 'npwp',
        'nama_bank', 'nama_bank_lainnya', 'no_rekening', 'nama_pemilik_rekening'
    ]

    # Exclude fields yang tidak perlu ditampilkan (termasuk nama_kegiatan, waktu_pelaksanaan, tempat_pelaksanaan)
    exclude_fields = ['id', 'user_id', 'buku_tabungan_path', 'tanda_tangan', 'created_at', 'updated_at', 'nama_kegiatan', 'waktu_pelaksanaan', 'tempat_pelaksanaan']

    # Collect all data in order
    for key in field_order:
        if key in biodata and key not in exclude_fields:
            label = field_labels.get(key, key.replace('_', ' ').title())
            value = biodata[key]
            display_value = str(value) if value and str(value).strip() else '-'
            all_data.append([label, display_value])

    # Add any remaining fields not in field_order
    for key, value in biodata.items():
        if key not in exclude_fields and key not in field_order:
            label = field_labels.get(key, key.replace('_', ' ').title())
            display_value = str(value) if value and str(value).strip() else '-'
            all_data.append([label, display_value])

    # Process tanda tangan menggunakan helper function
    print(f"DEBUG export_biodata_pdf: ===== Processing tanda tangan for NIK: {nik} =====")
    nama_user = biodata.get('nama_lengkap', 'N/A')
    print(f"DEBUG export_biodata_pdf: Nama: {nama_user}")

    tanda_tangan_temp_files = []
    tanda_tangan_raw = biodata.get('tanda_tangan')
    print(f"DEBUG export_biodata_pdf: Tanda tangan exists in biodata: {tanda_tangan_raw is not None}")
    if tanda_tangan_raw:
        print(f"DEBUG export_biodata_pdf: Tanda tangan type: {type(tanda_tangan_raw)}")
        if isinstance(tanda_tangan_raw, str):
            print(f"DEBUG export_biodata_pdf: Tanda tangan length: {len(tanda_tangan_raw)}")
            print(f"DEBUG export_biodata_pdf: Tanda tangan preview: {tanda_tangan_raw[:50]}...")

    tanda_tangan_img, error_msg = process_tanda_tangan_for_pdf(tanda_tangan_raw, tanda_tangan_temp_files)
    if error_msg:
        print(f"DEBUG export_biodata_pdf ERROR: {error_msg}")
    else:
        print(f"DEBUG export_biodata_pdf: Tanda tangan processed successfully")
    print(f"DEBUG export_biodata_pdf: Tanda tangan image object: {tanda_tangan_img is not None}")

    buku_tabungan_img = None

    # Buat 1 tabel untuk semua data - compact untuk muat di 1 halaman
    if all_data:
        # Prepare table data with Paragraph for text wrapping
        table_data = []
        # Header dihapus - langsung ke data rows
        # Data rows dengan Paragraph untuk text wrapping
        for row in all_data:
            field_label = Paragraph(f"<b>{row[0]}</b>", value_style)
            # Wrap long text values
            value_text = str(row[1]) if row[1] else '-'
            # Escape HTML characters
            value_text = value_text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            # Break long words - lebih pendek untuk compact
            if len(value_text) > 50:
                words = value_text.split()
                lines = []
                current_line = []
                current_length = 0
                for word in words:
                    if current_length + len(word) + 1 > 50:
                        if current_line:
                            lines.append(' '.join(current_line))
                        current_line = [word]
                        current_length = len(word)
                    else:
                        current_line.append(word)
                        current_length += len(word) + 1
                if current_line:
                    lines.append(' '.join(current_line))
                value_text = '<br/>'.join(lines)
            value_para = Paragraph(value_text, value_style)
            table_data.append([field_label, value_para])

        # Tambahkan tanda tangan ke tabel jika ada
        print(f"DEBUG export_biodata_pdf: Adding tanda tangan to table...")
        print(f"DEBUG export_biodata_pdf: tanda_tangan_img is not None: {tanda_tangan_img is not None}")
        print(f"DEBUG export_biodata_pdf: biodata has tanda_tangan: {biodata.get('tanda_tangan') is not None}")

        if tanda_tangan_img:
            print(f"DEBUG export_biodata_pdf: Adding tanda tangan image to table")
            field_label = Paragraph("<b>Tanda Tangan</b>", value_style)
            table_data.append([field_label, tanda_tangan_img])
            print(f"DEBUG export_biodata_pdf: Tanda tangan added to table successfully")
        elif biodata.get('tanda_tangan'):
            # Jika ada data tanda tangan tapi gagal diproses, tampilkan pesan
            print(f"DEBUG export_biodata_pdf: Adding error message to table (tanda tangan exists but failed to process)")
            field_label = Paragraph("<b>Tanda Tangan</b>", value_style)
            error_msg = Paragraph("<i>Tanda tangan tidak tersedia</i>", value_style)
            table_data.append([field_label, error_msg])
        else:
            print(f"DEBUG export_biodata_pdf: No tanda tangan data, skipping table row")

        # Calculate table width (F4 width minus margins)
        available_width = (8.27 * inch) - (25 * 2)
        table = Table(table_data, colWidths=[2.2*inch, available_width - 2.2*inch])
        table.setStyle(TableStyle([
            # Tidak ada header, langsung ke data rows
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            # Padding lebih kecil agar teks rata lebih kiri
            ('LEFTPADDING', (0, 0), (0, -1), 4),
            ('LEFTPADDING', (1, 0), (1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#067ac1')),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.08*inch))

    # Build PDF
    try:
        doc.build(elements, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
        buffer.seek(0)
        pdf_data = buffer.getvalue()
        buffer.close()

        # Generate filename
        filename = f"Biodata_{biodata.get('nama_lengkap', 'Unknown').replace(' ', '_')}_{nik}.pdf"

        # Refresh session sebelum return response untuk mencegah logout
        session.permanent = True
        session.modified = True

        # Return PDF as response
        return Response(
            pdf_data,
            mimetype='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )
    finally:
        # Cleanup temporary files tanda tangan
        for temp_file_path in tanda_tangan_temp_files:
            try:
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)
            except Exception as e:
                print(f"Error cleaning up tanda tangan temp file {temp_file_path}: {e}")

        # Cleanup temporary files logo
        try:
            if 'pendidikan_bermutu_processed_path' in locals() and os.path.exists(pendidikan_bermutu_processed_path):
                os.unlink(pendidikan_bermutu_processed_path)
        except Exception as e:
            print(f"Error cleaning up Pendidikan Bermutu temp file: {e}")

        try:
            if 'ramah_processed_path' in locals() and os.path.exists(ramah_processed_path):
                os.unlink(ramah_processed_path)
        except Exception as e:
            print(f"Error cleaning up Ramah temp file: {e}")

@app.route('/user/hapus-biodata/<path:nik>/<path:nama_kegiatan>', methods=['POST'])
@login_required
def user_hapus_biodata(nik, nama_kegiatan):
    """Hapus biodata user berdasarkan NIK dan nama_kegiatan"""
    from urllib.parse import unquote

    # Pastikan hanya user yang memiliki data tersebut yang bisa menghapus
    user_id = get_user_id()
    if not user_id:
        flash('Session tidak valid! Silakan login kembali.', 'error')
        return redirect(url_for('login'))

    # Decode URL encoding
    nik = unquote(nik)
    nama_kegiatan = unquote(nama_kegiatan)

    connection = get_db_connection()
    if not connection:
        flash('Koneksi database gagal!', 'error')
        return redirect(url_for('daftar_kegiatan'))

    try:
        cursor = connection.cursor()

        # Cek apakah data tersebut milik user yang sedang login
        cursor.execute("""
            SELECT nama_lengkap, user_id FROM biodata_kegiatan
            WHERE nik = ? AND TRIM(nama_kegiatan) = TRIM(?)
            LIMIT 1
        """, (nik, nama_kegiatan))
        biodata = cursor.fetchone()

        if not biodata:
            flash('Data tidak ditemukan!', 'error')
            return redirect(url_for('daftar_kegiatan'))


        # Pastikan user hanya bisa menghapus data miliknya sendiri
        if biodata['user_id'] != user_id:
            flash('Anda tidak memiliki izin untuk menghapus data ini!', 'error')
            return redirect(url_for('daftar_kegiatan'))

        nama_lengkap = biodata['nama_lengkap']

        # Hapus data biodata
        cursor.execute("""
            DELETE FROM biodata_kegiatan
            WHERE nik = ? AND TRIM(nama_kegiatan) = TRIM(?) AND user_id = ?
        """, (nik, nama_kegiatan, user_id))

        connection.commit()
        flash(f'Data biodata untuk "{nama_lengkap}" (NIK: {nik}) pada kegiatan "{nama_kegiatan}" berhasil dihapus!', 'success')

    except sqlite3.Error as e:
        if connection:
            connection.rollback()
        print(f"Error deleting biodata: {e}")
        flash(f'Terjadi kesalahan saat menghapus data: {str(e)}', 'error')
    finally:
        if connection:
            cursor.close()
            connection.close()

    return redirect(url_for('daftar_kegiatan'))

@app.route('/init-db', methods=['GET', 'POST'])
def init_db_route():
    """Route untuk inisialisasi database secara manual"""
    if request.method == 'POST':
        try:
            result = init_database()
            if result:
                flash('Database berhasil diinisialisasi!', 'success')
            else:
                flash('Gagal menginisialisasi database! Periksa file database dan permission folder.', 'error')
        except Exception as e:
            flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('index'))

    return '''
    <html>
        <head><title>Inisialisasi Database</title></head>
        <body>
            <h1>Inisialisasi Database</h1>
            <p>Klik tombol di bawah untuk menginisialisasi database:</p>
            <form method="POST">
                <button type="submit">Inisialisasi Database</button>
            </form>
            <br>
            <a href="/">Kembali ke Halaman Utama</a>
        </body>
    </html>
    '''

@app.route('/logout')
def logout():
    """Logout user"""
    clear_session()
    flash('Anda telah logout!', 'info')
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)

